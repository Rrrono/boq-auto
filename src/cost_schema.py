"""Normalized cost-data schema and lightweight SQLite repository."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
import json
import sqlite3
from typing import Any
from uuid import uuid4

from openpyxl import Workbook, load_workbook

from src.ingestion import ALIASES_HEADERS, RATE_LIBRARY_HEADERS, append_row, ensure_sheet_headers
from src.normalizer import normalize_text, normalize_unit
from src.utils import ensure_parent, safe_float


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def schema_database_path(database_path: str | Path) -> Path:
    path = Path(database_path)
    if path.suffix.lower() == ".sqlite":
        return path
    return path.with_suffix(".sqlite")


@dataclass(slots=True)
class CostSource:
    id: str
    name: str
    version: str
    file_path: str
    uploaded_at: str


@dataclass(slots=True)
class CostItem:
    id: str
    code: str
    description: str
    normalized_description: str
    unit: str
    domain: str
    work_family: str
    category: str
    subcategory: str
    item_kind: str
    project_context: str
    material: str
    keywords: list[str]
    rate: float
    source_id: str
    created_at: str


@dataclass(slots=True)
class CostAlias:
    id: str
    item_id: str
    alias: str


@dataclass(slots=True)
class IngestionLog:
    id: str
    source_id: str
    status: str
    message: str
    created_at: str


@dataclass(slots=True)
class ItemEmbedding:
    item_id: str
    embedding: list[float]
    model: str
    created_at: str


@dataclass(slots=True)
class MatchFeedback:
    id: str
    query_text: str
    item_id: str
    action: str
    alternative_item_id: str
    timestamp: str

    @property
    def selected_item_id(self) -> str:
        return self.item_id

    @property
    def created_at(self) -> str:
        return self.timestamp


@dataclass(slots=True)
class RateObservation:
    id: str
    description: str
    canonical_description: str
    unit: str
    rate: float
    source: str
    reviewer: str
    status: str
    metadata_json: str
    created_at: str


@dataclass(slots=True)
class CandidateReviewRecord:
    id: str
    description: str
    suggested_description: str
    unit: str
    reason: str
    reviewer: str
    status: str
    metadata_json: str
    created_at: str


@dataclass(slots=True)
class AliasSuggestion:
    id: str
    alias: str
    canonical_term: str
    section_bias: str
    reviewer: str
    status: str
    metadata_json: str
    created_at: str


class CostDatabase:
    """SQLite-backed normalized cost-data store with Excel compatibility helpers."""

    def __init__(self, path: str | Path) -> None:
        self.path = schema_database_path(path)

    def initialize(self) -> Path:
        ensure_parent(self.path)
        with sqlite3.connect(self.path) as conn:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS sources (
                    id TEXT PRIMARY KEY,
                    name TEXT NOT NULL,
                    version TEXT NOT NULL,
                    file_path TEXT NOT NULL,
                    uploaded_at TEXT NOT NULL
                )
                """
            )
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS items (
                    id TEXT PRIMARY KEY,
                    code TEXT NOT NULL,
                    description TEXT NOT NULL,
                    normalized_description TEXT NOT NULL,
                    unit TEXT NOT NULL,
                    domain TEXT NOT NULL DEFAULT '',
                    work_family TEXT NOT NULL DEFAULT '',
                    category TEXT NOT NULL,
                    subcategory TEXT NOT NULL,
                    item_kind TEXT NOT NULL DEFAULT 'work_item',
                    project_context TEXT NOT NULL DEFAULT '',
                    material TEXT NOT NULL DEFAULT '',
                    keywords TEXT NOT NULL DEFAULT '[]',
                    rate REAL NOT NULL,
                    source_id TEXT NOT NULL,
                    created_at TEXT NOT NULL,
                    FOREIGN KEY(source_id) REFERENCES sources(id)
                )
                """
            )
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS aliases (
                    id TEXT PRIMARY KEY,
                    item_id TEXT NOT NULL,
                    alias TEXT NOT NULL,
                    FOREIGN KEY(item_id) REFERENCES items(id)
                )
                """
            )
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS ingestion_logs (
                    id TEXT PRIMARY KEY,
                    source_id TEXT NOT NULL,
                    status TEXT NOT NULL,
                    message TEXT NOT NULL,
                    created_at TEXT NOT NULL,
                    FOREIGN KEY(source_id) REFERENCES sources(id)
                )
                """
            )
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS item_embeddings (
                    item_id TEXT PRIMARY KEY,
                    embedding TEXT NOT NULL,
                    model TEXT NOT NULL,
                    created_at TEXT NOT NULL,
                    FOREIGN KEY(item_id) REFERENCES items(id)
                )
                """
            )
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS match_feedback (
                    id TEXT PRIMARY KEY,
                    query_text TEXT NOT NULL,
                    item_id TEXT NOT NULL,
                    action TEXT NOT NULL,
                    alternative_item_id TEXT NOT NULL DEFAULT '',
                    timestamp TEXT NOT NULL
                )
                """
            )
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS rate_observations (
                    id TEXT PRIMARY KEY,
                    description TEXT NOT NULL,
                    canonical_description TEXT NOT NULL,
                    unit TEXT NOT NULL,
                    rate REAL NOT NULL,
                    source TEXT NOT NULL,
                    reviewer TEXT NOT NULL,
                    status TEXT NOT NULL,
                    metadata_json TEXT NOT NULL DEFAULT '{}',
                    created_at TEXT NOT NULL
                )
                """
            )
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS candidate_review_records (
                    id TEXT PRIMARY KEY,
                    description TEXT NOT NULL,
                    suggested_description TEXT NOT NULL,
                    unit TEXT NOT NULL,
                    reason TEXT NOT NULL,
                    reviewer TEXT NOT NULL,
                    status TEXT NOT NULL,
                    metadata_json TEXT NOT NULL DEFAULT '{}',
                    created_at TEXT NOT NULL
                )
                """
            )
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS alias_suggestions (
                    id TEXT PRIMARY KEY,
                    alias TEXT NOT NULL,
                    canonical_term TEXT NOT NULL,
                    section_bias TEXT NOT NULL DEFAULT '',
                    reviewer TEXT NOT NULL,
                    status TEXT NOT NULL,
                    metadata_json TEXT NOT NULL DEFAULT '{}',
                    created_at TEXT NOT NULL
                )
                """
            )
            existing_columns = {row[1] for row in conn.execute("PRAGMA table_info(items)")}
            if "domain" not in existing_columns:
                conn.execute("ALTER TABLE items ADD COLUMN domain TEXT NOT NULL DEFAULT ''")
            if "work_family" not in existing_columns:
                conn.execute("ALTER TABLE items ADD COLUMN work_family TEXT NOT NULL DEFAULT ''")
            if "material" not in existing_columns:
                conn.execute("ALTER TABLE items ADD COLUMN material TEXT NOT NULL DEFAULT ''")
            if "item_kind" not in existing_columns:
                conn.execute("ALTER TABLE items ADD COLUMN item_kind TEXT NOT NULL DEFAULT 'work_item'")
            if "project_context" not in existing_columns:
                conn.execute("ALTER TABLE items ADD COLUMN project_context TEXT NOT NULL DEFAULT ''")
            if "keywords" not in existing_columns:
                conn.execute("ALTER TABLE items ADD COLUMN keywords TEXT NOT NULL DEFAULT '[]'")
            feedback_columns = {row[1] for row in conn.execute("PRAGMA table_info(match_feedback)")}
            if "query_text" not in feedback_columns and {"query", "selected_item_id", "rejected_item_ids", "created_at"} <= feedback_columns:
                existing_feedback = conn.execute(
                    "SELECT id, query, selected_item_id, rejected_item_ids, created_at FROM match_feedback"
                ).fetchall()
                conn.execute("ALTER TABLE match_feedback RENAME TO match_feedback_legacy")
                conn.execute(
                    """
                    CREATE TABLE match_feedback (
                        id TEXT PRIMARY KEY,
                        query_text TEXT NOT NULL,
                        item_id TEXT NOT NULL,
                        action TEXT NOT NULL,
                        alternative_item_id TEXT NOT NULL DEFAULT '',
                        timestamp TEXT NOT NULL
                    )
                    """
                )
                for row in existing_feedback:
                    conn.execute(
                        """
                        INSERT INTO match_feedback (id, query_text, item_id, action, alternative_item_id, timestamp)
                        VALUES (?, ?, ?, ?, ?, ?)
                        """,
                        (
                            str(row[0]),
                            str(row[1]),
                            str(row[2]),
                            "accepted",
                            "",
                            str(row[4]),
                        ),
                    )
                conn.execute("DROP TABLE match_feedback_legacy")
        return self.path

    def register_source(self, name: str, version: str, file_path: str) -> CostSource:
        self.initialize()
        source = CostSource(
            id=str(uuid4()),
            name=name.strip() or "Unnamed Source",
            version=version.strip() or "v1",
            file_path=file_path,
            uploaded_at=utc_now_iso(),
        )
        with sqlite3.connect(self.path) as conn:
            conn.execute(
                "INSERT INTO sources (id, name, version, file_path, uploaded_at) VALUES (?, ?, ?, ?, ?)",
                (source.id, source.name, source.version, source.file_path, source.uploaded_at),
            )
        return source

    def insert_items(self, items: list[CostItem], aliases_by_item: dict[str, list[str]] | None = None) -> int:
        self.initialize()
        aliases_by_item = aliases_by_item or {}
        with sqlite3.connect(self.path) as conn:
            for item in items:
                conn.execute(
                    """
                    INSERT INTO items (id, code, description, normalized_description, unit, domain, work_family, category, subcategory, item_kind, project_context, material, keywords, rate, source_id, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        item.id,
                        item.code,
                        item.description,
                        item.normalized_description,
                        item.unit,
                        item.domain,
                        item.work_family,
                        item.category,
                        item.subcategory,
                        item.item_kind,
                        item.project_context,
                        item.material,
                        json.dumps(item.keywords, ensure_ascii=True),
                        item.rate,
                        item.source_id,
                        item.created_at,
                    ),
                )
                for alias in aliases_by_item.get(item.id, []):
                    conn.execute(
                        "INSERT INTO aliases (id, item_id, alias) VALUES (?, ?, ?)",
                        (str(uuid4()), item.id, alias),
                    )
        return len(items)

    def log_ingestion(self, source_id: str, status: str, message: str) -> IngestionLog:
        self.initialize()
        record = IngestionLog(
            id=str(uuid4()),
            source_id=source_id,
            status=status,
            message=message,
            created_at=utc_now_iso(),
        )
        with sqlite3.connect(self.path) as conn:
            conn.execute(
                "INSERT INTO ingestion_logs (id, source_id, status, message, created_at) VALUES (?, ?, ?, ?, ?)",
                (record.id, record.source_id, record.status, record.message, record.created_at),
            )
        return record

    def fetch_items(self) -> list[CostItem]:
        self.initialize()
        with sqlite3.connect(self.path) as conn:
            rows = conn.execute(
                """
                SELECT id, code, description, normalized_description, unit, domain, work_family, category, subcategory, item_kind, project_context, material, keywords, rate, source_id, created_at
                FROM items
                ORDER BY created_at DESC
                """
            ).fetchall()
        return [
            CostItem(
                id=str(row[0]),
                code=str(row[1]),
                description=str(row[2]),
                normalized_description=str(row[3]),
                unit=str(row[4]),
                domain=str(row[5]),
                work_family=str(row[6]),
                category=str(row[7]),
                subcategory=str(row[8]),
                item_kind=str(row[9]),
                project_context=str(row[10]),
                material=str(row[11]),
                keywords=list(json.loads(row[12] or "[]")),
                rate=float(row[13]),
                source_id=str(row[14]),
                created_at=str(row[15]),
            )
            for row in rows
        ]

    def fetch_ingestion_logs(self) -> list[IngestionLog]:
        self.initialize()
        with sqlite3.connect(self.path) as conn:
            rows = conn.execute(
                """
                SELECT id, source_id, status, message, created_at
                FROM ingestion_logs
                ORDER BY created_at DESC
                """
            ).fetchall()
        return [IngestionLog(*row) for row in rows]

    def fetch_aliases(self) -> list[CostAlias]:
        self.initialize()
        with sqlite3.connect(self.path) as conn:
            rows = conn.execute("SELECT id, item_id, alias FROM aliases ORDER BY alias ASC").fetchall()
        return [CostAlias(*row) for row in rows]

    def save_embedding(self, item_id: str, embedding: list[float], model: str) -> None:
        self.initialize()
        with sqlite3.connect(self.path) as conn:
            conn.execute(
                """
                INSERT INTO item_embeddings (item_id, embedding, model, created_at)
                VALUES (?, ?, ?, ?)
                ON CONFLICT(item_id) DO UPDATE SET
                    embedding=excluded.embedding,
                    model=excluded.model,
                    created_at=excluded.created_at
                """,
                (item_id, json.dumps(embedding), model, utc_now_iso()),
            )

    def fetch_embedding_lookup(self) -> dict[str, list[float]]:
        self.initialize()
        with sqlite3.connect(self.path) as conn:
            rows = conn.execute(
                """
                SELECT items.code, item_embeddings.embedding
                FROM item_embeddings
                JOIN items ON items.id = item_embeddings.item_id
                """
            ).fetchall()
        return {str(code): list(json.loads(embedding_json)) for code, embedding_json in rows}

    def clear_embeddings(self) -> int:
        self.initialize()
        with sqlite3.connect(self.path) as conn:
            before = conn.execute("SELECT COUNT(*) FROM item_embeddings").fetchone()
            conn.execute("DELETE FROM item_embeddings")
        return int((before or [0])[0])

    def fetch_embedding_stats(self) -> dict[str, Any]:
        self.initialize()
        with sqlite3.connect(self.path) as conn:
            total_items = int((conn.execute("SELECT COUNT(*) FROM items").fetchone() or [0])[0])
            embedded_items = int((conn.execute("SELECT COUNT(*) FROM item_embeddings").fetchone() or [0])[0])
            last_updated = conn.execute("SELECT MAX(created_at) FROM item_embeddings").fetchone()
        return {
            "total_items": total_items,
            "embedded_items": embedded_items,
            "last_updated": str((last_updated or [""])[0] or ""),
        }

    def fetch_embedding_records(self, limit: int = 20) -> list[dict[str, Any]]:
        self.initialize()
        with sqlite3.connect(self.path) as conn:
            rows = conn.execute(
                """
                SELECT items.code, items.description, item_embeddings.item_id, item_embeddings.model, item_embeddings.created_at
                FROM item_embeddings
                JOIN items ON items.id = item_embeddings.item_id
                ORDER BY item_embeddings.created_at DESC
                LIMIT ?
                """,
                (int(limit),),
            ).fetchall()
        return [
            {
                "code": str(row[0] or ""),
                "description": str(row[1] or ""),
                "item_id": str(row[2] or ""),
                "model": str(row[3] or ""),
                "created_at": str(row[4] or ""),
            }
            for row in rows
        ]

    def log_match_feedback(
        self,
        query_text: str,
        item_id: str,
        action: str,
        alternative_item_id: str = "",
    ) -> MatchFeedback:
        self.initialize()
        record = MatchFeedback(
            id=str(uuid4()),
            query_text=query_text,
            item_id=item_id,
            action=action,
            alternative_item_id=alternative_item_id,
            timestamp=utc_now_iso(),
        )
        with sqlite3.connect(self.path) as conn:
            conn.execute(
                """
                INSERT INTO match_feedback (id, query_text, item_id, action, alternative_item_id, timestamp)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (record.id, record.query_text, record.item_id, record.action, record.alternative_item_id, record.timestamp),
            )
        return record

    def fetch_match_feedback(self) -> list[MatchFeedback]:
        self.initialize()
        with sqlite3.connect(self.path) as conn:
            rows = conn.execute(
                """
                SELECT id, query_text, item_id, action, alternative_item_id, timestamp
                FROM match_feedback
                ORDER BY timestamp DESC
                """
            ).fetchall()
        return [MatchFeedback(*[str(value or "") for value in row]) for row in rows]

    def clear_match_feedback(self) -> int:
        self.initialize()
        with sqlite3.connect(self.path) as conn:
            before = conn.execute("SELECT COUNT(*) FROM match_feedback").fetchone()
            conn.execute("DELETE FROM match_feedback")
        return int((before or [0])[0])

    def resolve_item_id(self, item_ref: str) -> str:
        self.initialize()
        with sqlite3.connect(self.path) as conn:
            row = conn.execute(
                "SELECT id FROM items WHERE id = ? OR code = ? ORDER BY created_at DESC LIMIT 1",
                (item_ref, item_ref),
            ).fetchone()
        return str(row[0]) if row else ""

    def record_rate_observation(
        self,
        description: str,
        canonical_description: str,
        unit: str,
        rate: float,
        *,
        source: str,
        reviewer: str,
        status: str = "approved",
        metadata: dict[str, Any] | None = None,
    ) -> RateObservation:
        self.initialize()
        record = RateObservation(
            id=str(uuid4()),
            description=description.strip(),
            canonical_description=canonical_description.strip() or description.strip(),
            unit=unit.strip(),
            rate=float(rate),
            source=source.strip() or "review_task",
            reviewer=reviewer.strip(),
            status=status.strip() or "approved",
            metadata_json=json.dumps(metadata or {}, ensure_ascii=True),
            created_at=utc_now_iso(),
        )
        with sqlite3.connect(self.path) as conn:
            conn.execute(
                """
                INSERT INTO rate_observations (id, description, canonical_description, unit, rate, source, reviewer, status, metadata_json, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    record.id,
                    record.description,
                    record.canonical_description,
                    record.unit,
                    record.rate,
                    record.source,
                    record.reviewer,
                    record.status,
                    record.metadata_json,
                    record.created_at,
                ),
            )
        return record

    def fetch_rate_observations(self) -> list[RateObservation]:
        self.initialize()
        with sqlite3.connect(self.path) as conn:
            rows = conn.execute(
                """
                SELECT id, description, canonical_description, unit, rate, source, reviewer, status, metadata_json, created_at
                FROM rate_observations
                ORDER BY created_at DESC
                """
            ).fetchall()
        return [
            RateObservation(
                id=str(row[0]),
                description=str(row[1]),
                canonical_description=str(row[2]),
                unit=str(row[3]),
                rate=float(row[4]),
                source=str(row[5]),
                reviewer=str(row[6]),
                status=str(row[7]),
                metadata_json=str(row[8]),
                created_at=str(row[9]),
            )
            for row in rows
        ]

    def record_candidate_review(
        self,
        description: str,
        suggested_description: str,
        unit: str,
        *,
        reason: str,
        reviewer: str,
        status: str = "pending",
        metadata: dict[str, Any] | None = None,
    ) -> CandidateReviewRecord:
        self.initialize()
        record = CandidateReviewRecord(
            id=str(uuid4()),
            description=description.strip(),
            suggested_description=suggested_description.strip(),
            unit=unit.strip(),
            reason=reason.strip() or "review_required",
            reviewer=reviewer.strip(),
            status=status.strip() or "pending",
            metadata_json=json.dumps(metadata or {}, ensure_ascii=True),
            created_at=utc_now_iso(),
        )
        with sqlite3.connect(self.path) as conn:
            conn.execute(
                """
                INSERT INTO candidate_review_records (id, description, suggested_description, unit, reason, reviewer, status, metadata_json, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    record.id,
                    record.description,
                    record.suggested_description,
                    record.unit,
                    record.reason,
                    record.reviewer,
                    record.status,
                    record.metadata_json,
                    record.created_at,
                ),
            )
        return record

    def fetch_candidate_reviews(self) -> list[CandidateReviewRecord]:
        self.initialize()
        with sqlite3.connect(self.path) as conn:
            rows = conn.execute(
                """
                SELECT id, description, suggested_description, unit, reason, reviewer, status, metadata_json, created_at
                FROM candidate_review_records
                ORDER BY created_at DESC
                """
            ).fetchall()
        return [
            CandidateReviewRecord(
                id=str(row[0]),
                description=str(row[1]),
                suggested_description=str(row[2]),
                unit=str(row[3]),
                reason=str(row[4]),
                reviewer=str(row[5]),
                status=str(row[6]),
                metadata_json=str(row[7]),
                created_at=str(row[8]),
            )
            for row in rows
        ]

    def record_alias_suggestion(
        self,
        alias: str,
        canonical_term: str,
        *,
        section_bias: str = "",
        reviewer: str,
        status: str = "pending",
        metadata: dict[str, Any] | None = None,
    ) -> AliasSuggestion:
        self.initialize()
        record = AliasSuggestion(
            id=str(uuid4()),
            alias=alias.strip(),
            canonical_term=canonical_term.strip(),
            section_bias=section_bias.strip(),
            reviewer=reviewer.strip(),
            status=status.strip() or "pending",
            metadata_json=json.dumps(metadata or {}, ensure_ascii=True),
            created_at=utc_now_iso(),
        )
        with sqlite3.connect(self.path) as conn:
            conn.execute(
                """
                INSERT INTO alias_suggestions (id, alias, canonical_term, section_bias, reviewer, status, metadata_json, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    record.id,
                    record.alias,
                    record.canonical_term,
                    record.section_bias,
                    record.reviewer,
                    record.status,
                    record.metadata_json,
                    record.created_at,
                ),
            )
        return record

    def fetch_alias_suggestions(self) -> list[AliasSuggestion]:
        self.initialize()
        with sqlite3.connect(self.path) as conn:
            rows = conn.execute(
                """
                SELECT id, alias, canonical_term, section_bias, reviewer, status, metadata_json, created_at
                FROM alias_suggestions
                ORDER BY created_at DESC
                """
            ).fetchall()
        return [
            AliasSuggestion(
                id=str(row[0]),
                alias=str(row[1]),
                canonical_term=str(row[2]),
                section_bias=str(row[3]),
                reviewer=str(row[4]),
                status=str(row[5]),
                metadata_json=str(row[6]),
                created_at=str(row[7]),
            )
            for row in rows
        ]

    def export_to_excel(self, excel_path: str | Path) -> Path:
        self.initialize()
        target = Path(excel_path)
        workbook = load_workbook(target) if target.exists() else Workbook()
        if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) == 1:
            del workbook["Sheet"]

        rate_sheet = ensure_sheet_headers(workbook, "RateLibrary", RATE_LIBRARY_HEADERS)
        alias_sheet = ensure_sheet_headers(workbook, "Aliases", ALIASES_HEADERS)
        for sheet in (rate_sheet, alias_sheet):
            if sheet.max_row > 1:
                sheet.delete_rows(2, sheet.max_row - 1)

        item_lookup = {item.id: item for item in self.fetch_items()}
        for item in item_lookup.values():
            append_row(
                rate_sheet,
                RATE_LIBRARY_HEADERS,
                {
                    "item_code": item.code,
                    "description": item.description,
                    "normalized_description": item.normalized_description,
                    "section": item.category,
                    "subsection": item.subcategory,
                    "unit": item.unit,
                    "rate": item.rate,
                    "currency": "KES",
                    "region": "",
                    "source": "Normalized Cost Schema",
                    "source_sheet": "items",
                    "source_page": "",
                    "basis": "Exported from normalized schema",
                    "crew_type": "",
                    "plant_type": "",
                    "material_type": "",
                    "keywords": ", ".join(item.keywords),
                    "alias_group": "",
                    "build_up_recipe_id": "",
                    "confidence_hint": 0.0,
                    "notes": " | ".join(part for part in [item.domain, item.work_family, item.item_kind, item.project_context] if part),
                    "active": True,
                },
            )
        for alias in self.fetch_aliases():
            item = item_lookup.get(alias.item_id)
            if item is None:
                continue
            append_row(alias_sheet, ALIASES_HEADERS, {"alias": alias.alias, "canonical_term": item.description, "section_bias": item.category, "notes": ""})
        ensure_parent(target)
        workbook.save(target)
        return target

    def import_excel_database(self, excel_path: str | Path) -> int:
        self.initialize()
        workbook = load_workbook(excel_path, data_only=True)
        if "RateLibrary" not in workbook.sheetnames:
            return 0
        rate_sheet = workbook["RateLibrary"]
        headers = [str(cell.value or "").strip() for cell in next(rate_sheet.iter_rows(min_row=1, max_row=1))]
        source = self.register_source(Path(excel_path).stem, "excel-import", str(excel_path))
        items: list[CostItem] = []
        for row in rate_sheet.iter_rows(min_row=2, values_only=True):
            values = dict(zip(headers, row))
            description = str(values.get("description") or "").strip()
            if not description:
                continue
            taxonomy = derive_taxonomy_fields(
                description=description,
                category=str(values.get("section") or ""),
                subcategory=str(values.get("subsection") or ""),
                material=str(values.get("material_type") or ""),
            )
            items.append(
                CostItem(
                    id=str(uuid4()),
                    code=str(values.get("item_code") or ""),
                    description=description,
                    normalized_description=normalize_text(str(values.get("normalized_description") or description)),
                    unit=normalize_unit(str(values.get("unit") or "")),
                    domain=taxonomy["domain"],
                    work_family=taxonomy["work_family"],
                    category=str(values.get("section") or ""),
                    subcategory=str(values.get("subsection") or ""),
                    item_kind=taxonomy["item_kind"],
                    project_context=taxonomy["project_context"],
                    material=str(values.get("material_type") or ""),
                    keywords=[value.strip() for value in str(values.get("keywords") or "").split(",") if value.strip()],
                    rate=safe_float(values.get("rate")) or 0.0,
                    source_id=source.id,
                    created_at=utc_now_iso(),
                )
            )
        inserted = self.insert_items(items)
        self.log_ingestion(source.id, "completed", f"Imported {inserted} item(s) from Excel database.")
        return inserted


def build_cost_item(
    code: str,
    description: str,
    unit: str,
    category: str,
    subcategory: str,
    material: str,
    keywords: list[str] | None,
    rate: float,
    source_id: str,
    *,
    domain: str = "",
    work_family: str = "",
    item_kind: str = "",
    project_context: str = "",
) -> CostItem:
    taxonomy = derive_taxonomy_fields(
        description=description,
        category=category,
        subcategory=subcategory,
        material=material,
        domain=domain,
        work_family=work_family,
        item_kind=item_kind,
        project_context=project_context,
    )
    return CostItem(
        id=str(uuid4()),
        code=code.strip(),
        description=description.strip(),
        normalized_description=normalize_text(description),
        unit=normalize_unit(unit),
        domain=taxonomy["domain"],
        work_family=taxonomy["work_family"],
        category=category.strip(),
        subcategory=subcategory.strip(),
        item_kind=taxonomy["item_kind"],
        project_context=taxonomy["project_context"],
        material=material.strip(),
        keywords=list(keywords or []),
        rate=rate,
        source_id=source_id,
        created_at=utc_now_iso(),
    )


def composed_embedding_text(item: CostItem) -> str:
    """Build the canonical embedding text for a cost item."""

    return " | ".join(
        [
            item.domain.strip(),
            item.work_family.strip(),
            item.category.strip(),
            item.item_kind.strip(),
            item.project_context.strip(),
            item.material.strip(),
            item.description.strip(),
            item.unit.strip(),
        ]
    )


DOMAIN_HINTS: dict[str, tuple[str, ...]] = {
    "preliminaries": ("preliminar", "office", "ablution", "temporary", "starlink", "containerized"),
    "structures": ("concrete", "reinforcement", "rebar", "formwork", "footing", "masonry"),
    "roads": ("road", "grader", "roller", "bitumen", "pavement", "kerb", "furniture"),
    "drainage_utilities": ("pipe", "culvert", "manhole", "drain", "sewer", "water supply"),
    "electrical_lighting": ("electrical", "light", "lighting", "cable", "socket", "pole", "transformer"),
    "survey": ("survey", "beacon", "leveling", "total station", "setting out", "theodolite"),
    "lab_testing": ("laboratory", "lab", "testing", "cube test", "soil test"),
    "plant_transport": ("plant", "excavator", "tipper", "truck", "roller", "grader", "dozer", "pump", "mixer"),
    "accommodation_furnishings": ("furniture", "bed", "chair", "table", "housing", "wardrobe", "office equipment"),
}

WORK_FAMILY_HINTS: dict[str, tuple[str, ...]] = {
    "earthworks": ("excavat", "fill", "backfill", "compaction", "soil", "hardcore"),
    "concrete": ("concrete", "blinding", "reinforcement", "formwork", "footing"),
    "plant": ("excavator", "tipper", "truck", "roller", "grader", "dozer", "pump", "mixer", "vibrator"),
    "preliminaries": ("preliminar", "office", "temporary", "site establishment"),
    "piping": ("pipe", "culvert", "manhole", "drain", "sewer"),
    "roads": ("road", "bitumen", "pavement", "kerb", "furniture"),
    "survey": ("survey", "beacon", "total station", "setting out"),
    "lab_testing": ("laboratory", "lab", "testing"),
    "electrical": ("electrical", "light", "lighting", "cable", "transformer", "pole"),
}

EQUIPMENT_HINTS: tuple[str, ...] = (
    "excavator",
    "tipper",
    "truck",
    "roller",
    "grader",
    "dozer",
    "pump",
    "mixer",
    "compressor",
    "vibrator",
    "total station",
    "theodolite",
)


def derive_taxonomy_fields(
    *,
    description: str,
    category: str,
    subcategory: str,
    material: str,
    domain: str = "",
    work_family: str = "",
    item_kind: str = "",
    project_context: str = "",
) -> dict[str, str]:
    text = normalize_text(" ".join(part for part in [description, category, subcategory, material] if part))

    resolved_domain = domain.strip()
    if not resolved_domain:
        for candidate, terms in DOMAIN_HINTS.items():
            if any(term in text for term in terms):
                resolved_domain = candidate
                break

    resolved_work_family = work_family.strip()
    if not resolved_work_family:
        normalized_category = normalize_text(category)
        normalized_subcategory = normalize_text(subcategory)
        for candidate, terms in WORK_FAMILY_HINTS.items():
            if candidate == normalized_category or candidate == normalized_subcategory or any(term in text for term in terms):
                resolved_work_family = candidate
                break

    if not resolved_domain:
        if resolved_work_family in {"earthworks", "concrete"}:
            resolved_domain = "civil_works"
        elif resolved_work_family == "preliminaries":
            resolved_domain = "preliminaries"
        elif resolved_work_family == "piping":
            resolved_domain = "drainage_utilities"
        elif resolved_work_family == "electrical":
            resolved_domain = "electrical_lighting"
        elif resolved_work_family == "survey":
            resolved_domain = "survey"
        elif resolved_work_family == "lab_testing":
            resolved_domain = "lab_testing"
        elif resolved_work_family == "roads":
            resolved_domain = "roads"
        elif resolved_work_family == "plant":
            resolved_domain = "plant_transport"

    resolved_item_kind = item_kind.strip()
    if not resolved_item_kind:
        if any(term in text for term in EQUIPMENT_HINTS) or normalize_text(subcategory) == "plant":
            resolved_item_kind = "equipment"
        elif "service" in text or "testing" in text or "survey" in text:
            resolved_item_kind = "service"
        else:
            resolved_item_kind = "work_item"

    resolved_project_context = project_context.strip()
    if not resolved_project_context:
        if resolved_domain in {"roads", "plant_transport"}:
            resolved_project_context = "roads"
        elif resolved_domain == "structures":
            resolved_project_context = "structures"
        elif resolved_domain == "civil_works":
            resolved_project_context = "civil_general"
        elif resolved_domain == "drainage_utilities":
            resolved_project_context = "water_utilities"
        elif resolved_domain == "electrical_lighting":
            resolved_project_context = "electrical_airfield"
        elif resolved_domain in {"survey", "lab_testing"}:
            resolved_project_context = "engineering_support"
        elif resolved_domain == "accommodation_furnishings":
            resolved_project_context = "temporary_facilities"
        elif resolved_domain == "preliminaries":
            resolved_project_context = "general_requirements"

    return {
        "domain": resolved_domain,
        "work_family": resolved_work_family,
        "item_kind": resolved_item_kind,
        "project_context": resolved_project_context,
    }
