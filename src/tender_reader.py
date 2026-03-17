"""Tender document reader for local text and structured offline sources."""

from __future__ import annotations

import csv
import logging
from pathlib import Path

from openpyxl import load_workbook

from .tender_models import TenderDocument, TenderSourceLine


SUPPORTED_TENDER_SUFFIXES = {".txt", ".md", ".csv", ".xlsx", ".xlsm"}


def _clean_text(value: object) -> str:
    return " ".join(str(value).replace("\r", " ").replace("\n", " ").split()).strip()


def _first_non_empty(lines: list[TenderSourceLine]) -> str:
    for line in lines:
        if line.text:
            return line.text
    return ""


def _read_text_file(path: Path) -> list[TenderSourceLine]:
    raw_lines = path.read_text(encoding="utf-8").splitlines()
    lines: list[TenderSourceLine] = []
    for index, raw in enumerate(raw_lines, start=1):
        text = _clean_text(raw)
        if not text:
            continue
        lines.append(TenderSourceLine(source_reference=f"L{index}", text=text, line_number=index))
    return lines


def _read_csv_file(path: Path) -> list[TenderSourceLine]:
    lines: list[TenderSourceLine] = []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        for row_index, row in enumerate(reader, start=1):
            values = [_clean_text(value) for value in row if _clean_text(value)]
            if not values:
                continue
            text = " | ".join(values)
            lines.append(TenderSourceLine(source_reference=f"R{row_index}", text=text, line_number=row_index))
    return lines


def _read_excel_file(path: Path) -> list[TenderSourceLine]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    lines: list[TenderSourceLine] = []
    for sheet in workbook.worksheets:
        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            values = [_clean_text(value) for value in row if _clean_text(value)]
            if not values:
                continue
            text = " | ".join(values)
            reference = f"{sheet.title}!R{row_index}"
            lines.append(
                TenderSourceLine(
                    source_reference=reference,
                    text=text,
                    line_number=row_index,
                    sheet_name=sheet.title,
                )
            )
    return lines


def read_tender_document(path: str | Path, logger: logging.Logger | None = None) -> TenderDocument:
    """Read a local tender input source into normalized document lines."""

    source_path = Path(path)
    if not source_path.exists():
        raise FileNotFoundError(f"Tender input not found: {source_path}")

    suffix = source_path.suffix.lower()
    if suffix not in SUPPORTED_TENDER_SUFFIXES:
        raise ValueError(f"Unsupported tender input type: {suffix}")

    if suffix in {".txt", ".md"}:
        lines = _read_text_file(source_path)
        document_type = "text"
    elif suffix == ".csv":
        lines = _read_csv_file(source_path)
        document_type = "csv"
    else:
        lines = _read_excel_file(source_path)
        document_type = "excel"

    title = _first_non_empty(lines) or source_path.stem
    text = "\n".join(line.text for line in lines)
    document = TenderDocument(
        source_path=source_path,
        document_name=source_path.name,
        document_type=document_type,
        title=title,
        text=text,
        lines=lines,
    )
    if logger:
        logger.info("Read tender document %s with %s normalized lines.", source_path, len(lines))
    return document
