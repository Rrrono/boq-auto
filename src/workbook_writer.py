"""Excel workbook writing and formatting for priced BOQs and review sheets."""

from __future__ import annotations

from pathlib import Path

from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from .excel_loader import load_workbook_safe
from .models import CommercialTerms, MatchResult, QuotationSummary, SheetData, TextListEntry
from .utils import ensure_parent


THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
HEADER_FONT = Font(bold=True)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
TOP_WRAP_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
TOP_ALIGNMENT = Alignment(vertical="top")
RIGHT_ALIGNMENT = Alignment(horizontal="right", vertical="top")
RIGHT_WRAP_ALIGNMENT = Alignment(horizontal="right", vertical="top", wrap_text=True)

FIXED_WIDTHS = {
    "description": 60,
    "unit": 12,
    "quantity": 14,
    "qty": 14,
    "rate": 16,
    "amount": 18,
    "value": 18,
    "confidence": 12,
}
LONG_TEXT_HEADERS = {
    "description",
    "boq_description",
    "matched_description",
    "notes",
    "review_notes",
    "action_needed",
    "commercial_review_flags",
    "alternate_options",
    "rationale",
    "basis_of_rate",
    "source_basis",
    "source_excerpt",
    "extracted_text",
    "matched_keywords",
    "source_references",
    "commercial_clues",
}
NUMERIC_HEADERS = {
    "qty",
    "quantity",
    "rate",
    "amount",
    "value",
    "subtotal",
    "base_rate",
    "confidence",
    "confidence_score",
    "regional_factor",
}
SHEET_LONG_TEXT_MAP = {
    "pricing_handoff": {"description", "source_basis", "notes", "source_reference"},
    "gap_report": {"description", "notes", "source_reference"},
    "draft_boq": {"description", "notes", "source_basis", "source_reference"},
    "tender_analysis": {"value"},
    "table": set(),
    "boq": {"description"},
}


def _normalize_header(value: object) -> str:
    return str(value or "").strip().lower().replace(" ", "_")


def _looks_numeric_format(number_format: str) -> bool:
    markers = ("0", "#", "?")
    return any(marker in str(number_format or "") for marker in markers)


def _estimate_row_height(values: list[object], widths: list[float], wrapped: list[bool]) -> float:
    lines_needed = 1
    for value, width, should_wrap in zip(values, widths, wrapped):
        if value in (None, ""):
            continue
        text = str(value)
        explicit_lines = max(1, text.count("\n") + 1)
        if should_wrap and width > 0:
            estimated = max(explicit_lines, (len(text) // max(int(width) - 4, 8)) + 1)
        else:
            estimated = explicit_lines
        lines_needed = max(lines_needed, estimated)
    return min(max(15 * lines_needed, 18), 90)


def _guess_boq_header_row(ws) -> int:
    header_candidates = {"description", "particulars", "unit", "uom", "qty", "quantity", "rate", "amount", "total"}
    best_row = 1
    best_score = -1
    scan_limit = min(ws.max_row, 20)
    for row_index in range(1, scan_limit + 1):
        headers = [_normalize_header(ws.cell(row_index, column_index).value) for column_index in range(1, ws.max_column + 1)]
        score = sum(1 for header in headers if header in header_candidates)
        if score > best_score:
            best_score = score
            best_row = row_index
    return best_row


def format_worksheet(ws, sheet_type: str) -> None:
    """Apply consistent client-ready formatting without changing workbook values or formulas."""

    if ws.max_row == 0 or ws.max_column == 0:
        return

    resolved_sheet_type = sheet_type.lower().strip()
    header_row = _guess_boq_header_row(ws) if resolved_sheet_type == "boq" else 1
    headers = [_normalize_header(ws.cell(header_row, column_index).value) for column_index in range(1, ws.max_column + 1)]

    long_text_headers = set(LONG_TEXT_HEADERS) | set(SHEET_LONG_TEXT_MAP.get(resolved_sheet_type, set()))

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = THIN_BORDER

    for column_index, header in enumerate(headers, start=1):
        letter = get_column_letter(column_index)
        width = FIXED_WIDTHS.get(header)
        if width is None:
            max_length = max(
                len(str(ws.cell(row_index, column_index).value or ""))
                for row_index in range(1, ws.max_row + 1)
            )
            width = min(max(max_length + 2, 12), 32)
        ws.column_dimensions[letter].width = width

        is_long_text = header in long_text_headers
        is_numeric = header in NUMERIC_HEADERS

        for row_index in range(1, ws.max_row + 1):
            cell = ws.cell(row_index, column_index)
            if row_index == header_row:
                cell.font = HEADER_FONT
                cell.alignment = HEADER_ALIGNMENT
                continue
            if is_numeric or (cell.data_type == "f" and header in {"rate", "amount", "qty", "quantity", "value"}):
                cell.alignment = RIGHT_WRAP_ALIGNMENT if is_long_text else RIGHT_ALIGNMENT
                if cell.value not in (None, "") and not _looks_numeric_format(cell.number_format):
                    cell.number_format = "#,##0.00"
            else:
                cell.alignment = TOP_WRAP_ALIGNMENT if is_long_text else TOP_ALIGNMENT

    ws.freeze_panes = "A2"
    ws.row_dimensions[header_row].height = 24

    widths = [ws.column_dimensions[get_column_letter(index)].width or 12 for index in range(1, ws.max_column + 1)]
    wrapped = [
        bool(ws.cell(header_row + 1 if ws.max_row > header_row else header_row, index).alignment.wrap_text)
        or headers[index - 1] in long_text_headers
        for index in range(1, ws.max_column + 1)
    ]
    for row_index in range(header_row + 1, ws.max_row + 1):
        row_values = [ws.cell(row_index, column_index).value for column_index in range(1, ws.max_column + 1)]
        ws.row_dimensions[row_index].height = _estimate_row_height(row_values, widths, wrapped)


class WorkbookWriter:
    """Write priced BOQ output without disturbing the original layout more than necessary."""

    def write(
        self,
        source_workbook: str,
        output_workbook: str,
        sheets: list[SheetData],
        results: list[MatchResult],
        quotation_summary: QuotationSummary,
        assumptions: list[TextListEntry],
        exclusions: list[TextListEntry],
        commercial_terms: CommercialTerms,
        apply_rates: bool,
        write_amount_formulas: bool,
    ) -> Path:
        """Write rates and suggestions into a copied workbook."""
        ensure_parent(Path(output_workbook))
        workbook = load_workbook_safe(source_workbook)
        result_map = {(result.boq_line.sheet_name, result.boq_line.row_number): result for result in results}

        self._apply_pricing_to_boq(workbook, sheets, result_map, apply_rates, write_amount_formulas)
        self._reset_output_sheets(workbook)
        self._write_match_suggestions(workbook, results)
        self._write_quotation_summary(workbook, quotation_summary, commercial_terms)
        self._write_text_list_sheet(workbook, "Assumptions", "Assumption", assumptions)
        self._write_text_list_sheet(workbook, "Exclusions", "Exclusion", exclusions)
        self._write_basis_of_rates(workbook, results)
        self._write_commercial_review(workbook, results)
        self._format_boq_sheets(workbook, sheets)
        self._format_generated_sheets(workbook)

        workbook.save(output_workbook)
        return Path(output_workbook)

    def _apply_pricing_to_boq(self, workbook, sheets, result_map, apply_rates: bool, write_amount_formulas: bool) -> None:
        for sheet in sheets:
            worksheet = workbook[sheet.sheet_name]
            for row in sheet.rows:
                result = result_map.get((sheet.sheet_name, row.row_number))
                if not result or result.decision != "matched" or row.is_summary_row:
                    continue
                if apply_rates and sheet.columns.rate_col and result.rate is not None:
                    worksheet.cell(row.row_number, sheet.columns.rate_col).value = result.rate
                if (
                    apply_rates
                    and write_amount_formulas
                    and sheet.columns.amount_col
                    and sheet.columns.quantity_col
                    and sheet.columns.rate_col
                ):
                    qty_ref = worksheet.cell(row.row_number, sheet.columns.quantity_col).coordinate
                    rate_ref = worksheet.cell(row.row_number, sheet.columns.rate_col).coordinate
                    amount_cell = worksheet.cell(row.row_number, sheet.columns.amount_col)
                    if amount_cell.data_type != "f" or not amount_cell.value:
                        amount_cell.value = f"={qty_ref}*{rate_ref}"

    @staticmethod
    def _reset_output_sheets(workbook) -> None:
        for extra_sheet in [
            "Match Suggestions",
            "Quotation Summary",
            "Assumptions",
            "Exclusions",
            "Basis of Rates",
            "Commercial Review",
        ]:
            if extra_sheet in workbook.sheetnames:
                del workbook[extra_sheet]

    @staticmethod
    def _write_match_suggestions(workbook, results: list[MatchResult]) -> None:
        suggestions = workbook.create_sheet("Match Suggestions")
        suggestions.append(
            [
                "sheet_name", "row_number", "boq_description", "boq_unit", "boq_quantity", "decision",
                "matched_item_code", "matched_description", "matched_unit", "base_rate", "rate",
                "confidence_score", "confidence_band", "review_flag", "section_used", "source", "region_used", "built_up",
                "basis_of_rate", "approval_status", "flag_reasons", "generic_match_flag", "category_mismatch_flag",
                "section_mismatch_flag", "commercial_review_flags", "regional_factor", "alternate_options", "rationale",
            ]
        )
        for result in results:
            line = result.boq_line
            suggestions.append(
                [
                    line.sheet_name, line.row_number, line.description, line.unit, line.quantity, result.decision,
                    result.matched_item_code, result.matched_description, result.matched_unit, result.base_rate,
                    result.rate, result.confidence_score, result.confidence_band, "YES" if result.review_flag else "NO",
                    result.section_used, result.source, result.region_used, "YES" if result.built_up else "NO",
                    result.basis_of_rate, result.approval_status, "; ".join(result.flag_reasons),
                    "YES" if result.generic_match_flag else "NO", "YES" if result.category_mismatch_flag else "NO",
                    "YES" if result.section_mismatch_flag else "NO", "; ".join(result.commercial_review_flags),
                    result.regional_factor, " || ".join(result.alternate_options), "; ".join(result.rationale),
                ]
            )

    @staticmethod
    def _write_quotation_summary(workbook, quotation_summary: QuotationSummary, commercial_terms: CommercialTerms) -> None:
        summary = workbook.create_sheet("Quotation Summary")
        summary.append(["Field", "Value"])
        summary.append(["Currency", quotation_summary.currency])
        summary.append(["Bid Ready Status", "YES" if quotation_summary.bid_ready else "NO"])
        summary.append(["Bid Ready Note", quotation_summary.bid_ready_reason])
        summary.append(["Matched Items", quotation_summary.matched_items])
        summary.append(["Flagged Items", quotation_summary.flagged_items])
        summary.append(["Overheads %", commercial_terms.overhead_pct])
        summary.append(["Profit %", commercial_terms.profit_pct])
        summary.append(["Risk %", commercial_terms.risk_pct])
        summary.append(["VAT %", commercial_terms.vat_pct])
        summary.append(["", ""])
        summary.append(["Section", "Subtotal"])
        for section_total in quotation_summary.section_totals:
            summary.append([section_total.section, section_total.subtotal])
        summary.append(["Subtotal", quotation_summary.subtotal])
        summary.append(["Overheads", quotation_summary.overhead_amount])
        summary.append(["Profit", quotation_summary.profit_amount])
        summary.append(["Risk", quotation_summary.risk_amount])
        summary.append(["Pre-VAT Total", quotation_summary.pre_vat_total])
        summary.append(["VAT", quotation_summary.vat_amount])
        summary.append(["Grand Total", quotation_summary.grand_total])

    @staticmethod
    def _write_text_list_sheet(workbook, title: str, label: str, entries: list[TextListEntry]) -> None:
        sheet = workbook.create_sheet(title)
        sheet.append(["Category", label])
        for entry in entries:
            sheet.append([entry.category, entry.text])

    @staticmethod
    def _write_basis_of_rates(workbook, results: list[MatchResult]) -> None:
        basis_sheet = workbook.create_sheet("Basis of Rates")
        basis_sheet.append(
            [
                "sheet_name", "row_number", "boq_description", "matched_item_code", "matched_description",
                "source", "basis_of_rate", "region_used", "regional_factor",
            ]
        )
        for result in results:
            basis_sheet.append(
                [
                    result.boq_line.sheet_name, result.boq_line.row_number, result.boq_line.description,
                    result.matched_item_code, result.matched_description, result.source, result.basis_of_rate,
                    result.region_used, result.regional_factor,
                ]
            )

    @staticmethod
    def _write_commercial_review(workbook, results: list[MatchResult]) -> None:
        review_sheet = workbook.create_sheet("Commercial Review")
        review_sheet.append(
            [
                "sheet_name", "row_number", "boq_description", "decision", "approval_status",
                "confidence_band", "flag_reasons", "commercial_review_flags", "alternate_options",
            ]
        )
        for result in results:
            if not result.commercial_review_flags and not result.review_flag:
                continue
            review_sheet.append(
                [
                    result.boq_line.sheet_name, result.boq_line.row_number, result.boq_line.description,
                    result.decision, result.approval_status, result.confidence_band, "; ".join(result.flag_reasons),
                    "; ".join(result.commercial_review_flags), " || ".join(result.alternate_options),
                ]
            )

    @staticmethod
    def _format_boq_sheets(workbook, sheets: list[SheetData]) -> None:
        for sheet in sheets:
            if sheet.sheet_name in workbook.sheetnames:
                format_worksheet(workbook[sheet.sheet_name], "boq")

    @staticmethod
    def _format_generated_sheets(workbook) -> None:
        for sheet_name, sheet_type in {
            "Match Suggestions": "table",
            "Quotation Summary": "tender_analysis",
            "Assumptions": "table",
            "Exclusions": "table",
            "Basis of Rates": "table",
            "Commercial Review": "table",
        }.items():
            if sheet_name in workbook.sheetnames:
                format_worksheet(workbook[sheet_name], sheet_type)
