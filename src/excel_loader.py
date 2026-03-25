"""Resilient Excel workbook loading helpers."""

from __future__ import annotations

import io
import logging
import zipfile
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

from openpyxl import load_workbook

LOGGER = logging.getLogger("boq_auto.excel_loader")
WORKBOOK_XML_PATH = "xl/workbook.xml"
WORKBOOK_RELATIONSHIP_TAG = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}definedNames"


def load_workbook_safe(source: str | Path | bytes | io.BytesIO, /, **kwargs: Any):
    """Load a workbook, retrying with sanitized workbook metadata for malformed files."""
    try:
        return load_workbook(_coerce_source(source), **kwargs)
    except Exception as exc:
        if not _should_retry_with_sanitized_metadata(exc):
            raise
        LOGGER.warning("Retrying workbook load with sanitized metadata: %s", exc)
        sanitized = sanitize_workbook_archive(source)
        return load_workbook(sanitized, **kwargs)


def sanitize_workbook_archive(source: str | Path | bytes | io.BytesIO) -> io.BytesIO:
    """Strip workbook metadata that frequently breaks openpyxl on field spreadsheets."""
    workbook_bytes = _read_source_bytes(source)
    input_buffer = io.BytesIO(workbook_bytes)
    output_buffer = io.BytesIO()

    with zipfile.ZipFile(input_buffer, "r") as input_zip, zipfile.ZipFile(output_buffer, "w", zipfile.ZIP_DEFLATED) as output_zip:
        for member in input_zip.infolist():
            payload = input_zip.read(member.filename)
            if member.filename == WORKBOOK_XML_PATH:
                payload = _remove_defined_names(payload)
            output_zip.writestr(member, payload)

    output_buffer.seek(0)
    return output_buffer


def _remove_defined_names(workbook_xml: bytes) -> bytes:
    try:
        root = ElementTree.fromstring(workbook_xml)
    except ElementTree.ParseError:
        return workbook_xml

    removed = False
    for child in list(root):
        if child.tag == WORKBOOK_RELATIONSHIP_TAG:
            root.remove(child)
            removed = True

    if not removed:
        return workbook_xml
    return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)


def _read_source_bytes(source: str | Path | bytes | io.BytesIO) -> bytes:
    if isinstance(source, bytes):
        return source
    if isinstance(source, io.BytesIO):
        return source.getvalue()
    return Path(source).read_bytes()


def _coerce_source(source: str | Path | bytes | io.BytesIO):
    if isinstance(source, bytes):
        return io.BytesIO(source)
    return source


def _should_retry_with_sanitized_metadata(exc: Exception) -> bool:
    message = str(exc).lower()
    return "could not assign names" in message or "invalid xml" in message
