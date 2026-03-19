"""Orchestration for Tender Analysis v1."""

from __future__ import annotations

from dataclasses import asdict
import logging
from pathlib import Path

from openpyxl import Workbook, load_workbook

from .boq_drafter import BOQDrafter
from .boq_gap_checker import BOQGapChecker
from .checklist_builder import build_submission_checklist
from .clarification_log import ClarificationLogBuilder
from .models import AppConfig
from .requirement_extractor import RequirementExtractor
from .scope_parser import ScopeParser
from .tender_models import TenderAnalysisResult, TenderAnalysisSummary
from .tender_reader import read_tender_document
from .utils import dump_json, ensure_parent
from .workbook_writer import format_worksheet


def _append_table(worksheet, headers: list[str], rows: list[list[object]]) -> None:
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)


class TenderWorkflow:
    """Run the tender analysis pipeline and write review-friendly outputs."""

    def __init__(self, config: AppConfig, logger: logging.Logger | None = None) -> None:
        self.config = config
        self.logger = logger
        self.extractor = RequirementExtractor(config, logger)
        self.scope_parser = ScopeParser(config, logger)
        self.drafter = BOQDrafter(config, logger)
        self.gap_checker = BOQGapChecker(config, logger)
        self.clarification_builder = ClarificationLogBuilder(config, logger)

    def analyze(
        self,
        input_path: str,
        output_path: str,
        json_path: str | None = None,
        title_override: str | None = None,
    ) -> TenderAnalysisResult:
        result = self.prepare_result(input_path=input_path, title_override=title_override, boq_path=None, include_gap_check=False)
        workbook_path = self._write_workbook(Path(output_path), result)
        result.output_workbook = workbook_path
        if json_path:
            result.output_json = Path(json_path)
            self._write_json(result.output_json, result)
        return result

    def generate_checklist(
        self,
        input_path: str,
        output_path: str,
        json_path: str | None = None,
        title_override: str | None = None,
    ) -> TenderAnalysisResult:
        return self.analyze(input_path=input_path, output_path=output_path, json_path=json_path, title_override=title_override)

    def prepare_result(
        self,
        input_path: str,
        title_override: str | None = None,
        boq_path: str | None = None,
        include_gap_check: bool = True,
    ) -> TenderAnalysisResult:
        """Build tender analysis results in memory for downstream workflows."""

        document = read_tender_document(input_path, self.logger, self.config)
        if title_override:
            document.title = title_override
        requirements = self.extractor.extract(document)
        checklist_items = build_submission_checklist(requirements)
        scope_sections = self.scope_parser.parse(document)
        draft_suggestions = self.drafter.build_suggestions(document, scope_sections)
        clarifications = self.clarification_builder.build(document, requirements, scope_sections)
        gap_items = self.gap_checker.check(scope_sections, draft_suggestions, boq_path) if include_gap_check else []
        summary = self._build_summary(
            document,
            requirements,
            checklist_items,
            scope_sections,
            draft_suggestions,
            gap_items,
            clarifications,
        )
        return TenderAnalysisResult(
            document=document,
            requirements=requirements,
            checklist_items=checklist_items,
            scope_sections=scope_sections,
            draft_suggestions=draft_suggestions,
            gap_items=gap_items,
            clarifications=clarifications,
            summary=summary,
        )

    def gap_check(
        self,
        input_path: str,
        output_path: str,
        boq_path: str | None = None,
        json_path: str | None = None,
        title_override: str | None = None,
    ) -> TenderAnalysisResult:
        result = self.prepare_result(input_path=input_path, title_override=title_override, boq_path=boq_path, include_gap_check=True)
        workbook_path = self._write_workbook(Path(output_path), result)
        result.output_workbook = workbook_path
        if json_path:
            result.output_json = Path(json_path)
            self._write_json(result.output_json, result)
        return result

    def draft_boq(
        self,
        input_path: str,
        output_path: str,
        json_path: str | None = None,
        title_override: str | None = None,
    ) -> TenderAnalysisResult:
        result = self.prepare_result(input_path=input_path, title_override=title_override, boq_path=None, include_gap_check=False)
        workbook_path = self._write_workbook(Path(output_path), result)
        result.output_workbook = workbook_path
        if json_path:
            result.output_json = Path(json_path)
            self._write_json(result.output_json, result)
        return result

    def _build_summary(self, document, requirements, checklist_items, scope_sections, draft_suggestions, gap_items, clarifications) -> TenderAnalysisSummary:
        commercial_clues = [
            requirement.description
            for requirement in requirements
            if requirement.category in {"Securities", "Periods", "Pricing Instructions"}
        ]
        review_notes = [
            f"{requirement.requirement_id}: {requirement.notes or requirement.description}"
            for requirement in requirements
            if requirement.review_flag
        ]
        return TenderAnalysisSummary(
            document_name=document.document_name,
            title=document.title,
            total_requirements=len(requirements),
            mandatory_requirements=sum(1 for requirement in requirements if requirement.mandatory),
            flagged_requirements=sum(1 for requirement in requirements if requirement.review_flag),
            checklist_items=len(checklist_items),
            scope_sections=len(scope_sections),
            detected_scope=[section.section_name for section in scope_sections],
            commercial_clues=commercial_clues,
            review_notes=review_notes,
            gap_items=len(gap_items),
            draft_suggestions=len(draft_suggestions),
            clarifications=len(clarifications),
        )

    def append_result_sheets(self, workbook_path: str | Path, result: TenderAnalysisResult) -> Path:
        """Append or refresh tender-analysis sheets on an existing workbook."""

        workbook = load_workbook(workbook_path)
        self._reset_analysis_sheets(workbook)
        self._write_result_sheets(workbook, result)
        workbook.save(workbook_path)
        return Path(workbook_path)

    def _write_workbook(self, output_path: Path, result: TenderAnalysisResult) -> Path:
        ensure_parent(output_path)
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        self._write_result_sheets(workbook, result)
        workbook.save(output_path)
        if self.logger:
            self.logger.info("Tender analysis workbook written to %s", output_path)
        return output_path

    @staticmethod
    def _reset_analysis_sheets(workbook) -> None:
        for sheet_name in [
            "Tender Analysis Summary",
            "Tender Checklist",
            "Scope Summary",
            "Draft BOQ Suggestions",
            "BOQ Gap Report",
            "Clarification Log",
            "Extracted Requirements",
        ]:
            if sheet_name in workbook.sheetnames:
                del workbook[sheet_name]

    def _write_result_sheets(self, workbook, result: TenderAnalysisResult) -> None:
        document = result.document
        requirements = result.requirements
        checklist_items = result.checklist_items
        scope_sections = result.scope_sections
        draft_suggestions = result.draft_suggestions
        gap_items = result.gap_items
        clarifications = result.clarifications
        summary = result.summary
        if summary is None:
            return

        summary_sheet = workbook.create_sheet("Tender Analysis Summary")
        summary_sheet.append(["Field", "Value"])
        for row in [
            ["Document Name", summary.document_name],
            ["Title", summary.title],
            ["Source Path", str(document.source_path)],
            ["Document Type", document.document_type],
            ["Total Requirements", summary.total_requirements],
            ["Mandatory Requirements", summary.mandatory_requirements],
            ["Flagged For Review", summary.flagged_requirements],
            ["Checklist Items", summary.checklist_items],
            ["Detected Scope Sections", ", ".join(summary.detected_scope)],
            ["Draft BOQ Suggestions", summary.draft_suggestions],
            ["BOQ Gap Findings", summary.gap_items],
            ["Clarification Items", summary.clarifications],
            ["Commercial Clues", " | ".join(summary.commercial_clues)],
            ["Review Notes", " | ".join(summary.review_notes)],
        ]:
            summary_sheet.append(row)
        format_worksheet(summary_sheet, "tender_analysis")

        checklist_sheet = workbook.create_sheet("Tender Checklist")
        _append_table(
            checklist_sheet,
            [
                "requirement_id",
                "category",
                "description",
                "mandatory",
                "source_reference",
                "confidence",
                "action_needed",
                "owner",
                "status",
                "notes",
            ],
            [
                [
                    item.requirement_id,
                    item.category,
                    item.description,
                    "Yes" if item.mandatory else "No",
                    item.source_reference,
                    round(item.confidence, 1),
                    item.action_needed,
                    item.owner,
                    item.status,
                    item.notes,
                ]
                for item in checklist_items
            ],
        )
        format_worksheet(checklist_sheet, "table")

        scope_sheet = workbook.create_sheet("Scope Summary")
        _append_table(
            scope_sheet,
            ["section_name", "confidence", "source_references", "matched_keywords", "notes"],
            [
                [
                    section.section_name,
                    round(section.confidence, 1),
                    ", ".join(section.source_references),
                    ", ".join(section.matched_keywords),
                    section.notes,
                ]
                for section in scope_sections
            ],
        )
        format_worksheet(scope_sheet, "table")

        draft_sheet = workbook.create_sheet("Draft BOQ Suggestions")
        _append_table(
            draft_sheet,
            [
                "section",
                "description",
                "unit",
                "quantity_placeholder",
                "rate_placeholder",
                "amount_placeholder",
                "source_basis",
                "source_reference",
                "source_excerpt",
                "confidence",
                "review_required",
                "notes",
            ],
            [
                [
                    item.section,
                    item.description,
                    item.unit,
                    item.quantity_placeholder,
                    item.rate_placeholder,
                    item.amount_placeholder,
                    item.source_basis,
                    item.source_reference,
                    item.source_excerpt,
                    round(item.confidence, 1),
                    "Yes" if item.review_flag else "No",
                    item.notes,
                ]
                for item in draft_suggestions
            ],
        )
        format_worksheet(draft_sheet, "draft_boq")

        gap_sheet = workbook.create_sheet("BOQ Gap Report")
        _append_table(
            gap_sheet,
            [
                "gap_type",
                "section",
                "description",
                "source_reference",
                "existing_boq_reference",
                "confidence",
                "review_flag",
                "notes",
            ],
            [
                [
                    item.gap_type,
                    item.section,
                    item.description,
                    item.source_reference,
                    item.existing_boq_reference,
                    round(item.confidence, 1),
                    "Yes" if item.review_flag else "No",
                    item.notes,
                ]
                for item in gap_items
            ],
        )
        format_worksheet(gap_sheet, "gap_report")

        clarification_sheet = workbook.create_sheet("Clarification Log")
        _append_table(
            clarification_sheet,
            [
                "clarification_id",
                "category",
                "description",
                "source_reference",
                "confidence",
                "action_needed",
                "review_flag",
                "notes",
            ],
            [
                [
                    item.clarification_id,
                    item.category,
                    item.description,
                    item.source_reference,
                    round(item.confidence, 1),
                    item.action_needed,
                    "Yes" if item.review_flag else "No",
                    item.notes,
                ]
                for item in clarifications
            ],
        )
        format_worksheet(clarification_sheet, "table")

        requirements_sheet = workbook.create_sheet("Extracted Requirements")
        _append_table(
            requirements_sheet,
            [
                "requirement_id",
                "category",
                "description",
                "mandatory",
                "source_reference",
                "confidence",
                "review_flag",
                "matched_phrase",
                "extracted_text",
                "notes",
            ],
            [
                [
                    requirement.requirement_id,
                    requirement.category,
                    requirement.description,
                    "Yes" if requirement.mandatory else "No",
                    requirement.source_reference,
                    round(requirement.confidence, 1),
                    "Yes" if requirement.review_flag else "No",
                    requirement.matched_phrase,
                    requirement.extracted_text,
                    requirement.notes,
                ]
                for requirement in requirements
            ],
        )
        format_worksheet(requirements_sheet, "table")

    def _write_json(self, output_path: Path, result: TenderAnalysisResult) -> None:
        payload = {
            "document": {
                "source_path": str(result.document.source_path),
                "document_name": result.document.document_name,
                "document_type": result.document.document_type,
                "title": result.document.title,
                "line_count": len(result.document.lines),
            },
            "requirements": [asdict(requirement) for requirement in result.requirements],
            "checklist_items": [asdict(item) for item in result.checklist_items],
            "scope_sections": [asdict(section) for section in result.scope_sections],
            "draft_suggestions": [asdict(item) for item in result.draft_suggestions],
            "gap_items": [asdict(item) for item in result.gap_items],
            "clarifications": [asdict(item) for item in result.clarifications],
            "summary": asdict(result.summary) if result.summary else {},
        }
        dump_json(output_path, payload)
        if self.logger:
            self.logger.info("Tender analysis JSON written to %s", output_path)
