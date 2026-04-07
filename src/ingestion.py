"""Manual ingestion helpers for growing the Excel rate database."""

from __future__ import annotations

import csv
import hashlib
import json
import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
import tempfile
from typing import Any

from openpyxl import Workbook, load_workbook

from .pdf_ingestion import extract_text_from_pdf
from .normalizer import normalize_text, normalize_unit
from .section_inference import classify_sheet_name
from .workbook_reader import WorkbookReader
from .config import load_config
from .utils import safe_float, truthy

RATE_LIBRARY_HEADERS = [
    "item_code", "description", "normalized_description", "section", "subsection", "unit", "rate",
    "currency", "region", "source", "source_sheet", "source_page", "basis", "crew_type", "plant_type",
    "material_type", "keywords", "alias_group", "build_up_recipe_id", "confidence_hint", "notes", "active",
]

BUILDUP_INPUT_HEADERS = [
    "input_code", "input_type", "description", "unit", "rate", "region", "source", "active",
]

CANDIDATE_MATCH_HEADERS = [
    "timestamp", "import_batch_id", "source_file", "source_sheet", "target_sheet", "item_code",
    "description", "normalized_description", "section", "subsection", "unit", "rate", "currency", "region",
    "source", "source_page", "basis", "crew_type", "plant_type", "material_type", "keywords", "alias_group",
    "build_up_recipe_id", "confidence_hint", "notes", "active", "duplicate_reason", "matched_item_code",
    "reviewer_status", "reviewer_name", "reviewed_at", "review_decision", "promote_target",
    "approved_item_code", "approved_description", "approved_rate", "approved_canonical_term",
    "approved_section_bias", "confidence_override", "reviewer_note", "promotion_status", "promoted_at",
]

REVIEW_LOG_HEADERS = [
    "timestamp", "boq_file", "sheet_name", "row_number", "boq_description", "decision",
    "matched_item_code", "matched_description", "confidence_score", "reviewer_note",
]

CANDIDATE_REVIEW_HEADERS = [
    "candidate_row_number", "reviewer_status", "reviewer_name", "reviewed_at", "review_decision",
    "promote_target", "approved_item_code", "approved_description", "approved_rate", "approved_canonical_term",
    "approved_section_bias", "confidence_override", "reviewer_note", "source_file", "target_sheet", "item_code",
    "matched_item_code", "description", "section", "subsection", "unit", "rate", "currency", "region",
    "basis", "duplicate_reason", "promotion_status",
]

ALIASES_HEADERS = ["alias", "canonical_term", "section_bias", "notes"]

SOURCE_FIELD_ALIASES = {
    "item_code": ["item_code", "code", "item", "rate_code"],
    "input_code": ["input_code", "code", "item_code", "resource_code"],
    "description": ["description", "item_description", "particulars", "resource_description", "name"],
    "section": ["section", "trade", "bill_section", "category"],
    "subsection": ["subsection", "sub_section", "subtrade"],
    "unit": ["unit", "uom", "measure"],
    "rate": ["rate", "price", "unit_rate", "cost"],
    "currency": ["currency"],
    "region": ["region", "county", "location", "market"],
    "source": ["source", "manual", "library", "origin"],
    "source_sheet": ["source_sheet", "sheet", "sheet_name"],
    "source_page": ["source_page", "page", "page_no"],
    "basis": ["basis", "basis_of_rate"],
    "crew_type": ["crew_type", "labour_type"],
    "plant_type": ["plant_type", "equipment_type"],
    "material_type": ["material_type", "material_group"],
    "keywords": ["keywords", "tags"],
    "alias_group": ["alias_group"],
    "build_up_recipe_id": ["build_up_recipe_id", "recipe_id"],
    "confidence_hint": ["confidence_hint", "confidence"],
    "notes": ["notes", "remark", "remarks"],
    "active": ["active", "status"],
    "input_type": ["input_type", "resource_type", "type"],
}

REGION_ALIASES = {
    "nrb": "Nairobi",
    "nai": "Nairobi",
    "nairobi county": "Nairobi",
    "nyanza region": "Nyanza",
    "kisumu": "Nyanza",
    "western kenya": "Western",
}
PDF_UNIT_TOKENS = {"m2", "m3", "m", "nr", "sum", "item", "kg", "km", "day", "hr", "ton", "sm", "cm", "no", "nos"}
PRICE_TOKEN_RE = re.compile(r"^\d[\d,]*(?:\.\d+)?$")


@dataclass(slots=True)
class ImportSummary:
    target_sheet: str
    source_file: str
    total_rows: int = 0
    appended: int = 0
    skipped_duplicates: int = 0
    candidates_created: int = 0
    merged: int = 0
    normalized_rows: int = 0
    reviewed: int = 0
    rejected: int = 0
    promoted: int = 0
    report_rows: int = 0
    training_records: int = 0
    notes: list[str] = field(default_factory=list)
    appended_records: list[dict[str, Any]] = field(default_factory=list)


@dataclass(slots=True)
class BoqImportPreview:
    source_file: str
    region: str
    total_extracted: int = 0
    skipped_missing_rate: int = 0
    skipped_missing_unit: int = 0
    append_count: int = 0
    candidate_count: int = 0
    duplicate_count: int = 0
    extracted_rows: list[dict[str, Any]] = field(default_factory=list)
    append_preview: list[dict[str, Any]] = field(default_factory=list)
    candidate_preview: list[dict[str, Any]] = field(default_factory=list)
    duplicate_preview: list[dict[str, Any]] = field(default_factory=list)
    alias_preview: list[dict[str, Any]] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)


def normalize_region_name(region: str) -> str:
    """Normalize common region and county labels into a consistent market name."""
    text = normalize_text(region)
    if not text:
        return ""
    return REGION_ALIASES.get(text, text.title())


def timestamp_now() -> str:
    """Return a stable import timestamp."""
    return datetime.now().isoformat(timespec="seconds")


def ensure_sheet_headers(workbook: Workbook, sheet_name: str, headers: list[str]):
    """Create an optional sheet if it does not yet exist and validate its header row."""
    if sheet_name not in workbook.sheetnames:
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(headers)
        return sheet
    sheet = workbook[sheet_name]
    existing_headers = [str(cell.value or "").strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    if not existing_headers or all(not header for header in existing_headers):
        sheet.delete_rows(1, 1)
        sheet.append(headers)
        return sheet
    missing = [header for header in headers if header not in existing_headers]
    if missing:
        for header in missing:
            sheet.cell(1, sheet.max_column + 1).value = header
    return sheet


def workbook_headers(sheet) -> list[str]:
    """Return normalized headers for a worksheet."""
    return [str(cell.value or "").strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]


def read_structured_table(input_path: str, sheet_name: str | None = None) -> list[dict[str, Any]]:
    """Read a CSV or Excel table into a list of row dictionaries."""
    path = Path(input_path)
    if path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            return list(csv.DictReader(handle))

    workbook = load_workbook(path, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    headers = workbook_headers(worksheet)
    rows: list[dict[str, Any]] = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))
    return rows


def lookup_value(row: dict[str, Any], field: str, default: Any = "") -> Any:
    """Find a source field using the known alias list."""
    normalized = {normalize_text(str(key)): value for key, value in row.items()}
    for alias in SOURCE_FIELD_ALIASES.get(field, [field]):
        key = normalize_text(alias)
        if key in normalized and normalized[key] not in (None, ""):
            return normalized[key]
    return default


def append_row(sheet, headers: list[str], values: dict[str, Any]) -> None:
    """Append a row to a sheet using header order."""
    sheet_headers = workbook_headers(sheet)
    sheet.append([values.get(header, "") for header in sheet_headers or headers])


def build_rate_library_row(
    row: dict[str, Any],
    source_file: str,
    source_sheet: str,
    defaults: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Map a source row into the RateLibrary structure."""
    defaults = defaults or {}
    description = str(lookup_value(row, "description", defaults.get("description", ""))).strip()
    unit = normalize_unit(str(lookup_value(row, "unit", defaults.get("unit", ""))))
    region = normalize_region_name(str(lookup_value(row, "region", defaults.get("region", ""))))
    return {
        "item_code": str(lookup_value(row, "item_code", defaults.get("item_code", ""))).strip(),
        "description": description,
        "normalized_description": normalize_text(description),
        "section": str(lookup_value(row, "section", defaults.get("section", ""))).strip(),
        "subsection": str(lookup_value(row, "subsection", defaults.get("subsection", ""))).strip(),
        "unit": unit,
        "rate": safe_float(lookup_value(row, "rate", defaults.get("rate", 0.0))) or 0.0,
        "currency": str(lookup_value(row, "currency", defaults.get("currency", "KES"))).strip() or "KES",
        "region": region,
        "source": str(lookup_value(row, "source", defaults.get("source", source_file))).strip() or source_file,
        "source_sheet": str(lookup_value(row, "source_sheet", defaults.get("source_sheet", source_sheet))).strip() or source_sheet,
        "source_page": str(lookup_value(row, "source_page", defaults.get("source_page", ""))).strip(),
        "basis": str(lookup_value(row, "basis", defaults.get("basis", "Imported source table"))).strip(),
        "crew_type": str(lookup_value(row, "crew_type", defaults.get("crew_type", ""))).strip(),
        "plant_type": str(lookup_value(row, "plant_type", defaults.get("plant_type", ""))).strip(),
        "material_type": str(lookup_value(row, "material_type", defaults.get("material_type", ""))).strip(),
        "keywords": str(lookup_value(row, "keywords", defaults.get("keywords", ""))).strip(),
        "alias_group": str(lookup_value(row, "alias_group", defaults.get("alias_group", ""))).strip(),
        "build_up_recipe_id": str(lookup_value(row, "build_up_recipe_id", defaults.get("build_up_recipe_id", ""))).strip(),
        "confidence_hint": safe_float(lookup_value(row, "confidence_hint", defaults.get("confidence_hint", 0.0))) or 0.0,
        "notes": str(lookup_value(row, "notes", defaults.get("notes", ""))).strip(),
        "active": defaults.get("active", True if lookup_value(row, "active", True) in (None, "") else truthy(lookup_value(row, "active", True))),
    }


def build_buildup_input_row(
    row: dict[str, Any],
    source_file: str,
    source_sheet: str = "",
    defaults: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Map a source row into the BuildUpInputs structure."""
    defaults = defaults or {}
    description = str(lookup_value(row, "description", defaults.get("description", ""))).strip()
    return {
        "input_code": str(lookup_value(row, "input_code", defaults.get("input_code", ""))).strip(),
        "input_type": str(lookup_value(row, "input_type", defaults.get("input_type", ""))).strip(),
        "description": description,
        "unit": normalize_unit(str(lookup_value(row, "unit", defaults.get("unit", "")))),
        "rate": safe_float(lookup_value(row, "rate", defaults.get("rate", 0.0))) or 0.0,
        "region": normalize_region_name(str(lookup_value(row, "region", defaults.get("region", "")))),
        "source": str(lookup_value(row, "source", defaults.get("source", source_file))).strip() or source_file,
        "active": defaults.get("active", True if lookup_value(row, "active", True) in (None, "") else truthy(lookup_value(row, "active", True))),
    }


def existing_rate_indexes(sheet) -> tuple[dict[str, int], dict[tuple[str, str, str], int]]:
    """Build duplicate-detection indexes for RateLibrary or BuildUpInputs sheets."""
    headers = workbook_headers(sheet)
    code_field = "item_code" if "item_code" in headers else "input_code"
    description_field = "normalized_description" if "normalized_description" in headers else "description"
    code_index: dict[str, int] = {}
    desc_unit_region_index: dict[tuple[str, str, str], int] = {}
    header_pos = {header: idx for idx, header in enumerate(headers)}
    for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        item_code = str(values[header_pos[code_field]] or "").strip()
        description = str(values[header_pos[description_field]] or "").strip()
        if description_field == "description":
            description = normalize_text(description)
        unit = normalize_unit(str(values[header_pos["unit"]] or ""))
        region = normalize_region_name(str(values[header_pos["region"]] or "")) if "region" in header_pos else ""
        if item_code:
            code_index[item_code] = row_number
        if description and unit:
            desc_unit_region_index[(description, unit, region)] = row_number
    return code_index, desc_unit_region_index


def write_candidate_match(
    workbook: Workbook,
    candidate_row: dict[str, Any],
) -> None:
    """Append an uncertain import to CandidateMatches."""
    sheet = ensure_sheet_headers(workbook, "CandidateMatches", CANDIDATE_MATCH_HEADERS)
    append_row(sheet, CANDIDATE_MATCH_HEADERS, candidate_row)


def write_review_log(workbook: Workbook, note: dict[str, Any]) -> None:
    """Append an automated review or merge action to ReviewLog."""
    sheet = ensure_sheet_headers(workbook, "ReviewLog", REVIEW_LOG_HEADERS)
    append_row(sheet, REVIEW_LOG_HEADERS, note)


def candidate_positions(sheet) -> dict[str, int]:
    """Return header positions for CandidateMatches-like sheets."""
    return {header: index + 1 for index, header in enumerate(workbook_headers(sheet))}


def set_candidate_defaults(sheet) -> None:
    """Ensure review workflow defaults exist on candidate rows."""
    positions = candidate_positions(sheet)
    for row_number in range(2, sheet.max_row + 1):
        if "reviewer_status" in positions and not sheet.cell(row_number, positions["reviewer_status"]).value:
            sheet.cell(row_number, positions["reviewer_status"]).value = "pending"
        if "review_decision" in positions and not sheet.cell(row_number, positions["review_decision"]).value:
            sheet.cell(row_number, positions["review_decision"]).value = "hold"
        if "promote_target" in positions and not sheet.cell(row_number, positions["promote_target"]).value:
            sheet.cell(row_number, positions["promote_target"]).value = "candidatematches"
        if "promotion_status" in positions and not sheet.cell(row_number, positions["promotion_status"]).value:
            sheet.cell(row_number, positions["promotion_status"]).value = "not_promoted"


def _review_artifact_marker(metadata: dict[str, Any], fallback: str) -> str:
    task_id = str(metadata.get("task_id") or "").strip()
    return f"schema-task:{task_id or fallback}"


def _artifact_candidate_status(record_type: str, record_status: str) -> str:
    status = str(record_status or "").strip().lower()
    if record_type in {"rate_observation", "alias_suggestion"}:
        return "approved" if status in {"approved", "logged"} else "pending"
    if status in {"approved", "logged"}:
        return "pending"
    return "pending"


def sync_review_artifacts_to_candidate_matches(db_path: str, schema_path: str | None = None) -> ImportSummary:
    """Materialize normalized reviewer artifacts back into CandidateMatches for the existing review flow."""
    from .cost_schema import CostDatabase, schema_database_path

    workbook = load_workbook(db_path)
    candidate_sheet = ensure_sheet_headers(workbook, "CandidateMatches", CANDIDATE_MATCH_HEADERS)
    set_candidate_defaults(candidate_sheet)
    ensure_sheet_headers(workbook, "ReviewLog", REVIEW_LOG_HEADERS)

    repository = CostDatabase(schema_path or schema_database_path(db_path))
    summary = ImportSummary(target_sheet="CandidateMatches", source_file=str(schema_path or db_path))
    existing_markers = {
        str(values[2] or "").strip()
        for values in candidate_sheet.iter_rows(min_row=2, values_only=True)
        if str(values[2] or "").strip().startswith("schema-task:")
    }

    def append_candidate(record_type: str, payload: dict[str, Any]) -> None:
        metadata = payload.pop("metadata")
        marker = _review_artifact_marker(metadata, payload["record_id"])
        if marker in existing_markers:
            summary.skipped_duplicates += 1
            return
        payload["source_file"] = marker
        append_row(candidate_sheet, CANDIDATE_MATCH_HEADERS, payload)
        existing_markers.add(marker)
        summary.appended += 1
        write_review_log(
            workbook,
            {
                "timestamp": payload.get("reviewed_at") or timestamp_now(),
                "boq_file": db_path,
                "sheet_name": "CandidateMatches",
                "row_number": candidate_sheet.max_row,
                "boq_description": str(payload.get("description") or ""),
                "decision": f"schema-{record_type}-synced",
                "matched_item_code": str(payload.get("matched_item_code") or payload.get("approved_item_code") or ""),
                "matched_description": str(payload.get("approved_description") or payload.get("approved_canonical_term") or payload.get("description") or ""),
                "confidence_score": safe_float(payload.get("confidence_hint")) or 0.0,
                "reviewer_note": str(payload.get("reviewer_note") or "Synced from normalized reviewer artifact"),
            },
        )

    for record in repository.fetch_rate_observations():
        metadata = json.loads(record.metadata_json or "{}")
        append_candidate(
            "rate-observation",
            {
                "timestamp": record.created_at,
                "import_batch_id": f"schema-sync-{record.created_at[:10]}",
                "record_id": record.id,
                "source_sheet": "rate_observations",
                "target_sheet": "RateLibrary",
                "item_code": str(metadata.get("approved_item_code") or metadata.get("item_code") or ""),
                "description": record.description,
                "normalized_description": normalize_text(record.description),
                "section": str(metadata.get("section") or metadata.get("category") or ""),
                "subsection": str(metadata.get("subsection") or ""),
                "unit": normalize_unit(record.unit),
                "rate": record.rate,
                "currency": str(metadata.get("currency") or "KES"),
                "region": normalize_region_name(str(metadata.get("region") or "")),
                "source": record.source or "review_task",
                "source_page": record.id,
                "basis": "Approved reviewer rate observation",
                "crew_type": str(metadata.get("crew_type") or ""),
                "plant_type": str(metadata.get("plant_type") or ""),
                "material_type": str(metadata.get("material_type") or ""),
                "keywords": str(metadata.get("keywords") or ""),
                "alias_group": str(metadata.get("alias_group") or ""),
                "build_up_recipe_id": str(metadata.get("build_up_recipe_id") or ""),
                "confidence_hint": safe_float(metadata.get("confidence_override")) or 90.0,
                "notes": "Synced from normalized rate observation",
                "active": True,
                "duplicate_reason": "schema reviewer artifact",
                "matched_item_code": str(metadata.get("matched_item_code") or ""),
                "reviewer_status": _artifact_candidate_status("rate_observation", record.status),
                "reviewer_name": record.reviewer,
                "reviewed_at": record.created_at,
                "review_decision": "manual_rate",
                "promote_target": "ratelibrary",
                "approved_item_code": str(metadata.get("approved_item_code") or metadata.get("item_code") or ""),
                "approved_description": record.canonical_description,
                "approved_rate": record.rate,
                "approved_canonical_term": record.canonical_description,
                "approved_section_bias": str(metadata.get("section") or ""),
                "confidence_override": safe_float(metadata.get("confidence_override")) or 90.0,
                "reviewer_note": str(metadata.get("reviewer_note") or "Approved reviewer rate observation"),
                "promotion_status": "not_promoted",
                "promoted_at": "",
                "metadata": metadata,
            },
        )

    for record in repository.fetch_alias_suggestions():
        metadata = json.loads(record.metadata_json or "{}")
        append_candidate(
            "alias-suggestion",
            {
                "timestamp": record.created_at,
                "import_batch_id": f"schema-sync-{record.created_at[:10]}",
                "record_id": record.id,
                "source_sheet": "alias_suggestions",
                "target_sheet": "Aliases",
                "item_code": str(metadata.get("approved_item_code") or ""),
                "description": record.alias,
                "normalized_description": normalize_text(record.alias),
                "section": record.section_bias,
                "subsection": "",
                "unit": "",
                "rate": "",
                "currency": "KES",
                "region": normalize_region_name(str(metadata.get("region") or "")),
                "source": "review_task",
                "source_page": record.id,
                "basis": "Approved reviewer alias suggestion",
                "crew_type": "",
                "plant_type": "",
                "material_type": "",
                "keywords": str(metadata.get("keywords") or ""),
                "alias_group": "",
                "build_up_recipe_id": "",
                "confidence_hint": safe_float(metadata.get("confidence_override")) or 85.0,
                "notes": "Synced from normalized alias suggestion",
                "active": True,
                "duplicate_reason": "schema reviewer artifact",
                "matched_item_code": str(metadata.get("matched_item_code") or ""),
                "reviewer_status": _artifact_candidate_status("alias_suggestion", record.status),
                "reviewer_name": record.reviewer,
                "reviewed_at": record.created_at,
                "review_decision": "alias_suggestion",
                "promote_target": "aliases",
                "approved_item_code": str(metadata.get("approved_item_code") or ""),
                "approved_description": record.canonical_term,
                "approved_rate": "",
                "approved_canonical_term": record.canonical_term,
                "approved_section_bias": record.section_bias,
                "confidence_override": safe_float(metadata.get("confidence_override")) or 85.0,
                "reviewer_note": str(metadata.get("reviewer_note") or "Approved reviewer alias suggestion"),
                "promotion_status": "not_promoted",
                "promoted_at": "",
                "metadata": metadata,
            },
        )

    for record in repository.fetch_candidate_reviews():
        metadata = json.loads(record.metadata_json or "{}")
        append_candidate(
            "candidate-review",
            {
                "timestamp": record.created_at,
                "import_batch_id": f"schema-sync-{record.created_at[:10]}",
                "record_id": record.id,
                "source_sheet": "candidate_review_records",
                "target_sheet": "CandidateMatches",
                "item_code": str(metadata.get("approved_item_code") or metadata.get("item_code") or ""),
                "description": record.description,
                "normalized_description": normalize_text(record.description),
                "section": str(metadata.get("section") or metadata.get("category") or ""),
                "subsection": str(metadata.get("subsection") or ""),
                "unit": normalize_unit(record.unit),
                "rate": "",
                "currency": str(metadata.get("currency") or "KES"),
                "region": normalize_region_name(str(metadata.get("region") or "")),
                "source": "review_task",
                "source_page": record.id,
                "basis": "Reviewer escalation / no-good-match record",
                "crew_type": str(metadata.get("crew_type") or ""),
                "plant_type": str(metadata.get("plant_type") or ""),
                "material_type": str(metadata.get("material_type") or ""),
                "keywords": str(metadata.get("keywords") or ""),
                "alias_group": "",
                "build_up_recipe_id": "",
                "confidence_hint": safe_float(metadata.get("confidence_override")) or 40.0,
                "notes": f"Synced from normalized candidate review: {record.reason}",
                "active": True,
                "duplicate_reason": "schema reviewer artifact",
                "matched_item_code": str(metadata.get("matched_item_code") or ""),
                "reviewer_status": _artifact_candidate_status("candidate_review", record.status),
                "reviewer_name": record.reviewer,
                "reviewed_at": record.created_at,
                "review_decision": record.reason or "hold",
                "promote_target": "candidatematches",
                "approved_item_code": "",
                "approved_description": record.suggested_description,
                "approved_rate": "",
                "approved_canonical_term": record.suggested_description,
                "approved_section_bias": str(metadata.get("section") or ""),
                "confidence_override": safe_float(metadata.get("confidence_override")) or "",
                "reviewer_note": str(metadata.get("reviewer_note") or record.reason or "Needs workbook review"),
                "promotion_status": "not_promoted",
                "promoted_at": "",
                "metadata": metadata,
            },
        )

    workbook.save(db_path)
    return summary


def import_structured_rows(
    db_path: str,
    input_path: str,
    target_sheet: str,
    mapper,
    defaults: dict[str, Any] | None = None,
    source_sheet: str | None = None,
) -> ImportSummary:
    """Import structured rows into a target database sheet with duplicate handling."""
    workbook = load_workbook(db_path)
    target_headers = RATE_LIBRARY_HEADERS if target_sheet == "RateLibrary" else BUILDUP_INPUT_HEADERS
    target = ensure_sheet_headers(workbook, target_sheet, target_headers)
    candidate_sheet = ensure_sheet_headers(workbook, "CandidateMatches", CANDIDATE_MATCH_HEADERS)
    set_candidate_defaults(candidate_sheet)
    ensure_sheet_headers(workbook, "ReviewLog", REVIEW_LOG_HEADERS)
    rows = read_structured_table(input_path, source_sheet)
    code_index, desc_unit_region_index = existing_rate_indexes(target)

    batch_id = f"import-{timestamp_now().replace(':', '').replace('-', '')}"
    summary = ImportSummary(target_sheet=target_sheet, source_file=input_path, total_rows=len(rows))

    for raw_row in rows:
        mapped = mapper(raw_row, Path(input_path).name, source_sheet or "")
        if defaults:
            mapped = {**mapped, **{key: value for key, value in defaults.items() if value not in (None, "")}}
        if not mapped.get("description"):
            summary.notes.append("Skipped row with blank description.")
            continue

        item_code_key = "item_code" if target_sheet == "RateLibrary" else "input_code"
        item_code = str(mapped.get(item_code_key, "")).strip()
        normalized_description = str(mapped.get("normalized_description") or normalize_text(mapped.get("description", ""))).strip()
        unit = normalize_unit(str(mapped.get("unit", "")))
        region = normalize_region_name(str(mapped.get("region", "")))
        duplicate_reason = ""
        matched_item_code = ""

        if item_code and item_code in code_index:
            duplicate_reason = f"duplicate {item_code_key}"
            matched_item_code = item_code
        elif normalized_description and unit and (normalized_description, unit, region) in desc_unit_region_index:
            duplicate_reason = "duplicate normalized description + unit + region"
            matched_item_code = str(mapped.get(item_code_key, "")) or str(desc_unit_region_index[(normalized_description, unit, region)])

        if duplicate_reason:
            existing_row_number = code_index.get(item_code) or desc_unit_region_index.get((normalized_description, unit, region))
            existing_rate = safe_float(target.cell(existing_row_number, workbook_headers(target).index("rate") + 1).value) if existing_row_number else None
            incoming_rate = safe_float(mapped.get("rate"))
            materially_different = existing_rate is None or incoming_rate is None or abs(existing_rate - incoming_rate) > 0.001
            if materially_different:
                write_candidate_match(
                    workbook,
                    {
                        "timestamp": timestamp_now(),
                        "import_batch_id": batch_id,
                        "source_file": Path(input_path).name,
                        "source_sheet": source_sheet or "",
                        "target_sheet": target_sheet,
                        "item_code": mapped.get("item_code") or mapped.get("input_code", ""),
                        "description": mapped.get("description", ""),
                        "normalized_description": normalized_description,
                        "section": mapped.get("section", ""),
                        "subsection": mapped.get("subsection", ""),
                        "unit": unit,
                        "rate": mapped.get("rate", ""),
                        "currency": mapped.get("currency", ""),
                        "region": mapped.get("region", ""),
                        "source": mapped.get("source", ""),
                        "source_page": mapped.get("source_page", ""),
                        "basis": mapped.get("basis", ""),
                        "crew_type": mapped.get("crew_type", ""),
                        "plant_type": mapped.get("plant_type", ""),
                        "material_type": mapped.get("material_type", ""),
                        "keywords": mapped.get("keywords", ""),
                        "alias_group": mapped.get("alias_group", ""),
                        "build_up_recipe_id": mapped.get("build_up_recipe_id", ""),
                        "confidence_hint": mapped.get("confidence_hint", 0.0),
                        "notes": mapped.get("notes", ""),
                        "active": mapped.get("active", True),
                        "duplicate_reason": duplicate_reason,
                        "matched_item_code": matched_item_code,
                        "reviewer_status": "pending",
                        "reviewer_name": "",
                        "reviewed_at": "",
                        "review_decision": "hold",
                        "promote_target": "candidatematches",
                        "approved_item_code": "",
                        "approved_description": "",
                        "approved_rate": "",
                        "approved_canonical_term": "",
                        "approved_section_bias": "",
                        "confidence_override": "",
                        "reviewer_note": "",
                        "promotion_status": "not_promoted",
                        "promoted_at": "",
                    },
                )
                summary.candidates_created += 1
            else:
                summary.skipped_duplicates += 1
            continue

        append_row(target, target_headers, mapped)
        if item_code:
            code_index[item_code] = target.max_row
        desc_unit_region_index[(normalized_description, unit, region)] = target.max_row
        summary.appended += 1
        summary.appended_records.append(dict(mapped))

    workbook.save(db_path)
    return summary


def _classify_structured_import_rows(
    db_path: str,
    target_sheet: str,
    mapped_rows: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    workbook = load_workbook(db_path, data_only=True)
    target_headers = RATE_LIBRARY_HEADERS if target_sheet == "RateLibrary" else BUILDUP_INPUT_HEADERS
    target = ensure_sheet_headers(workbook, target_sheet, target_headers)
    code_index, desc_unit_region_index = existing_rate_indexes(target)
    target_header_positions = {header: index + 1 for index, header in enumerate(workbook_headers(target))}

    append_rows: list[dict[str, Any]] = []
    candidate_rows: list[dict[str, Any]] = []
    duplicate_rows: list[dict[str, Any]] = []

    for mapped in mapped_rows:
        item_code_key = "item_code" if target_sheet == "RateLibrary" else "input_code"
        item_code = str(mapped.get(item_code_key, "")).strip()
        normalized_description = str(mapped.get("normalized_description") or normalize_text(mapped.get("description", ""))).strip()
        unit = normalize_unit(str(mapped.get("unit", "")))
        region = normalize_region_name(str(mapped.get("region", "")))
        duplicate_reason = ""
        matched_item_code = ""

        if item_code and item_code in code_index:
            duplicate_reason = f"duplicate {item_code_key}"
            matched_item_code = item_code
        elif normalized_description and unit and (normalized_description, unit, region) in desc_unit_region_index:
            duplicate_reason = "duplicate normalized description + unit + region"
            matched_item_code = str(mapped.get(item_code_key, "")) or str(desc_unit_region_index[(normalized_description, unit, region)])

        if not duplicate_reason:
            append_rows.append(
                {
                    "decision": "append",
                    "item_code": item_code,
                    "description": mapped.get("description", ""),
                    "section": mapped.get("section", ""),
                    "unit": unit,
                    "rate": mapped.get("rate", ""),
                    "region": region,
                    "source_sheet": mapped.get("source_sheet", ""),
                    "source_page": mapped.get("source_page", ""),
                }
            )
            if item_code:
                code_index[item_code] = -1
            if normalized_description and unit:
                desc_unit_region_index[(normalized_description, unit, region)] = -1
            continue

        existing_row_number = code_index.get(item_code) or desc_unit_region_index.get((normalized_description, unit, region))
        existing_rate = safe_float(target.cell(existing_row_number, target_header_positions["rate"]).value) if existing_row_number and existing_row_number > 0 else None
        incoming_rate = safe_float(mapped.get("rate"))
        materially_different = existing_rate is None or incoming_rate is None or abs(existing_rate - incoming_rate) > 0.001
        preview_row = {
            "decision": "candidate" if materially_different else "duplicate",
            "item_code": item_code,
            "description": mapped.get("description", ""),
            "section": mapped.get("section", ""),
            "unit": unit,
            "incoming_rate": incoming_rate,
            "existing_rate": existing_rate,
            "region": region,
            "duplicate_reason": duplicate_reason,
            "matched_item_code": matched_item_code,
            "source_sheet": mapped.get("source_sheet", ""),
            "source_page": mapped.get("source_page", ""),
        }
        if materially_different:
            candidate_rows.append(preview_row)
        else:
            duplicate_rows.append(preview_row)

    return append_rows, candidate_rows, duplicate_rows


def _boq_import_code(region: str, section: str, description: str, unit: str) -> str:
    region_code = normalize_text(region).replace(" ", "")[:3].upper() or "GEN"
    digest = hashlib.sha1(f"{normalize_text(section)}|{normalize_text(description)}|{normalize_unit(unit)}|{normalize_text(region)}".encode("utf-8")).hexdigest()[:8].upper()
    return f"BOQ-{region_code}-{digest}"


def _boq_keywords(description: str) -> str:
    tokens = [token for token in normalize_text(description).split() if len(token) > 3]
    return ", ".join(dict.fromkeys(tokens[:8]))


def _style_aliases(description: str, section: str = "") -> list[str]:
    raw = " ".join(str(description or "").split()).strip(" -.,;:")
    if not raw:
        return []
    aliases: list[str] = []
    if ";" in raw:
        aliases.append(raw.split(";", 1)[0].strip(" -.,;:"))
    without_parens = re.sub(r"\([^)]*\)", "", raw).strip(" -.,;:")
    if without_parens and without_parens != raw:
        aliases.append(without_parens)
    without_dimensions = re.sub(r"\b\d+(?:\.\d+)?(?:mm|cm|m|kg|nr|m2|m3)?\b", "", without_parens, flags=re.IGNORECASE)
    without_dimensions = re.sub(r"\b(?:x|dia|depth|thick|thickness|not exceeding)\b", "", without_dimensions, flags=re.IGNORECASE)
    without_dimensions = " ".join(without_dimensions.split()).strip(" -.,;:")
    if without_dimensions and without_dimensions != raw:
        aliases.append(without_dimensions)
    normalized_section = normalize_text(section)
    normalized_raw = normalize_text(raw)
    if normalized_section and normalized_section in normalized_raw:
        removed_section = re.sub(rf"\b{re.escape(normalized_section)}\b", "", normalized_raw).strip()
        if removed_section and removed_section != normalized_raw:
            aliases.append(removed_section.title())
    cleaned: list[str] = []
    seen: set[str] = set()
    canonical = normalize_text(raw)
    for alias in aliases:
        normalized = normalize_text(alias)
        if not normalized or normalized == canonical or len(normalized.split()) < 2:
            continue
        if normalized in seen:
            continue
        seen.add(normalized)
        cleaned.append(" ".join(alias.split()))
    return cleaned[:5]


def _sync_aliases_to_excel(db_path: str, aliases_by_description: dict[str, tuple[str, str, list[str]]]) -> int:
    workbook = load_workbook(db_path)
    alias_sheet = ensure_sheet_headers(workbook, "Aliases", ALIASES_HEADERS)
    existing = {
        (
            str(values[0] or "").strip().lower(),
            str(values[1] or "").strip().lower(),
        )
        for values in alias_sheet.iter_rows(min_row=2, values_only=True)
    }
    added = 0
    for canonical_description, (section, note, aliases) in aliases_by_description.items():
        for alias in aliases:
            alias_text = str(alias or "").strip()
            if not alias_text:
                continue
            key = (alias_text.lower(), canonical_description.lower())
            if key in existing:
                continue
            append_row(
                alias_sheet,
                ALIASES_HEADERS,
                {
                    "alias": alias_text,
                    "canonical_term": canonical_description,
                    "section_bias": section,
                    "notes": note,
                },
            )
            existing.add(key)
            added += 1
    if added:
        workbook.save(db_path)
    return added


def _sync_priced_boq_records_to_schema(
    db_path: str,
    boq_path: str,
    appended_records: list[dict[str, Any]],
    aliases_by_description: dict[str, tuple[str, str, list[str]]],
    ai_assist: bool = False,
    config=None,
) -> int:
    from .ai.embedding_provider import get_embedding_provider
    from .cost_schema import CostDatabase, build_cost_item, composed_embedding_text, schema_database_path

    if not appended_records:
        return 0
    runtime_config = config or load_config()
    provider = get_embedding_provider(runtime_config) if ai_assist else None
    repository = CostDatabase(schema_database_path(db_path))
    source = repository.register_source(Path(boq_path).stem, "priced-boq-import", boq_path)
    cost_items = []
    aliases_by_item: dict[str, list[str]] = {}
    for record in appended_records:
        description = str(record.get("description") or "").strip()
        section = str(record.get("section") or "").strip()
        keywords = [value.strip() for value in str(record.get("keywords") or "").split(",") if value.strip()]
        item = build_cost_item(
            code=str(record.get("item_code") or "").strip(),
            description=description,
            unit=str(record.get("unit") or ""),
            category=section,
            subcategory=str(record.get("subsection") or ""),
            material=str(record.get("material_type") or ""),
            keywords=keywords,
            rate=safe_float(record.get("rate")) or 0.0,
            source_id=source.id,
        )
        rule_aliases = list(aliases_by_description.get(description, ("", "", []))[2])
        if ai_assist and provider is not None and hasattr(provider, "suggest_aliases"):
            try:
                suggested = [alias for alias in provider.suggest_aliases(description) if alias]
                rule_aliases.extend(suggested)
            except Exception:
                pass
        aliases_by_item[item.id] = list(dict.fromkeys(alias for alias in rule_aliases if normalize_text(alias) != normalize_text(description)))
        cost_items.append(item)
    repository.insert_items(cost_items, aliases_by_item=aliases_by_item)
    repository.log_ingestion(source.id, "completed", f"Imported {len(cost_items)} priced BOQ item(s).")
    if provider is not None:
        for item in cost_items:
            embedding = provider.embed(composed_embedding_text(item))
            if embedding:
                repository.save_embedding(item.id, embedding, getattr(provider, "model_name", "unknown"))
    return len(cost_items)


def _collect_priced_boq_aliases(
    records: list[dict[str, Any]],
    ai_assist: bool = False,
    config=None,
) -> tuple[dict[str, tuple[str, str, list[str]]], list[dict[str, Any]]]:
    runtime_config = config or load_config()
    provider = None
    if ai_assist:
        try:
            from .ai.embedding_provider import get_embedding_provider

            provider = get_embedding_provider(runtime_config)
        except Exception:
            provider = None

    aliases_by_description: dict[str, tuple[str, str, list[str]]] = {}
    alias_preview: list[dict[str, Any]] = []
    for record in records:
        description = str(record.get("description") or "").strip()
        if not description:
            continue
        section = str(record.get("section") or "").strip()
        rule_aliases = _style_aliases(description, section)
        ai_aliases: list[str] = []
        if ai_assist and provider is not None and hasattr(provider, "suggest_aliases"):
            try:
                ai_aliases = [alias for alias in provider.suggest_aliases(description) if alias]
            except Exception:
                ai_aliases = []
        combined_aliases = list(
            dict.fromkeys(
                alias
                for alias in [*rule_aliases, *ai_aliases]
                if normalize_text(alias) != normalize_text(description)
            )
        )
        aliases_by_description[description] = (
            section,
            "Derived from priced BOQ import",
            rule_aliases,
        )
        alias_preview.append(
            {
                "description": description,
                "section": section,
                "rule_aliases": ", ".join(rule_aliases),
                "ai_aliases": ", ".join(ai_aliases),
                "combined_aliases": ", ".join(combined_aliases),
            }
        )
    return aliases_by_description, alias_preview


def _is_probable_pdf_heading(line: str) -> bool:
    stripped = " ".join(str(line or "").split()).strip()
    if not stripped:
        return False
    words = stripped.split()
    digit_count = sum(char.isdigit() for char in stripped)
    return digit_count <= 2 and len(words) <= 8 and (stripped.isupper() or len(words) <= 4)


def _extract_priced_boq_pdf_rows(
    boq_path: Path,
    normalized_region: str,
    source_label: str | None,
    runtime_config,
) -> tuple[list[dict[str, Any]], int, int]:
    text = extract_text_from_pdf(str(boq_path), config=runtime_config)
    extracted_rows: list[dict[str, Any]] = []
    skipped_missing_rate = 0
    skipped_missing_unit = 0
    current_section = ""
    pending_description = ""

    for line_number, raw_line in enumerate(text.splitlines(), start=1):
        line = " ".join(str(raw_line).split()).strip()
        if not line:
            continue
        if _is_probable_pdf_heading(line):
            current_section = line.title()
            pending_description = ""
            continue

        parsed = _parse_priced_boq_pdf_line(line)
        if parsed is None and pending_description:
            parsed = _parse_priced_boq_pdf_line(f"{pending_description} {line}")
            if parsed is not None:
                pending_description = ""
        if parsed is None:
            if not any(normalize_unit(token) in {"m2", "m3", "m", "nr", "sum", "item", "kg", "km", "day", "hr", "ton"} for token in line.split()):
                pending_description = line if len(line.split()) >= 3 else ""
            continue

        description, unit, rate = parsed
        if not unit:
            skipped_missing_unit += 1
            continue
        if rate is None or rate <= 0:
            skipped_missing_rate += 1
            continue

        section = current_section or "Imported BOQ"
        extracted_rows.append(
            {
                "item_code": _boq_import_code(normalized_region, section, description, unit),
                "description": description,
                "normalized_description": normalize_text(description),
                "section": section,
                "subsection": "",
                "unit": unit,
                "rate": rate,
                "currency": "KES",
                "region": normalized_region,
                "source": source_label.strip() if source_label else boq_path.stem,
                "source_sheet": "PDF",
                "source_page": str(line_number),
                "basis": "Imported from priced BOQ PDF review",
                "crew_type": "",
                "plant_type": "",
                "material_type": "",
                "keywords": _boq_keywords(description),
                "alias_group": "",
                "build_up_recipe_id": "",
                "confidence_hint": 0.0,
                "notes": f"Priced BOQ PDF import line {line_number}",
                "active": True,
            }
        )
    return extracted_rows, skipped_missing_rate, skipped_missing_unit


def _parse_priced_boq_pdf_line(line: str) -> tuple[str, str, float | None] | None:
    tokens = [token.strip() for token in str(line or "").split() if token.strip()]
    if len(tokens) < 5:
        return None

    unit_index = -1
    unit_value = ""
    for index, token in enumerate(tokens):
        normalized = normalize_unit(token)
        if normalized in {"m2", "m3", "m", "nr", "sum", "item", "kg", "km", "day", "hr", "ton"}:
            unit_index = index
            unit_value = normalized
            break
    if unit_index <= 1 or unit_index >= len(tokens) - 1:
        return None

    trailing_tokens = tokens[unit_index + 1 :]
    numeric_values = [safe_float(token.replace(",", "")) for token in trailing_tokens if PRICE_TOKEN_RE.match(token)]
    numeric_values = [value for value in numeric_values if value is not None]
    if len(numeric_values) < 2:
        return None

    description = " ".join(tokens[:unit_index]).strip(" -.,;:")
    if len(description.split()) < 3:
        return None
    rate = numeric_values[1]
    return description, unit_value, rate


def _extract_priced_boq_rows(
    boq_path: str,
    region: str,
    source_label: str | None = None,
    config=None,
    column_overrides: dict[str, int] | None = None,
) -> tuple[list[dict[str, Any]], int, int]:
    normalized_region = normalize_region_name(region)
    if not normalized_region:
        raise ValueError("Region is required for priced BOQ imports.")

    boq_source = Path(boq_path)
    runtime_config = config or load_config()
    extracted_rows: list[dict[str, Any]] = []
    skipped_missing_rate = 0
    skipped_missing_unit = 0

    if boq_source.suffix.lower() in {".xlsx", ".xlsm"}:
        reader = WorkbookReader(runtime_config)
        sheets = reader.read(str(boq_source), column_overrides or {})

        for sheet in sheets:
            default_section = classify_sheet_name(sheet.sheet_name) or sheet.sheet_name
            for row in sheet.rows:
                if row.is_heading or row.is_summary_row or not row.description:
                    continue
                if safe_float(row.rate) is None or safe_float(row.rate) <= 0:
                    skipped_missing_rate += 1
                    continue
                if not normalize_unit(row.unit):
                    skipped_missing_unit += 1
                    continue

                section = (row.inferred_section or default_section or sheet.sheet_name).strip()
                description = str(row.description).strip()
                unit = normalize_unit(str(row.unit))
                extracted_rows.append(
                    {
                        "item_code": _boq_import_code(normalized_region, section, description, unit),
                        "description": description,
                        "normalized_description": normalize_text(description),
                        "section": section,
                        "subsection": "",
                        "unit": unit,
                        "rate": safe_float(row.rate) or 0.0,
                        "currency": "KES",
                        "region": normalized_region,
                        "source": source_label.strip() if source_label else boq_source.stem,
                        "source_sheet": sheet.sheet_name,
                        "source_page": str(row.row_number),
                        "basis": "Imported from priced BOQ review",
                        "crew_type": "",
                        "plant_type": "",
                        "material_type": "",
                        "keywords": _boq_keywords(description),
                        "alias_group": "",
                        "build_up_recipe_id": "",
                        "confidence_hint": 0.0,
                        "notes": f"Priced BOQ import from {sheet.sheet_name}!R{row.row_number}",
                        "active": True,
                    }
                )
    elif boq_source.suffix.lower() == ".pdf":
        extracted_rows, skipped_missing_rate, skipped_missing_unit = _extract_priced_boq_pdf_rows(
            boq_source,
            normalized_region,
            source_label,
            runtime_config,
        )
    else:
        raise ValueError("Priced BOQ import currently supports Excel workbooks and PDFs only.")

    return extracted_rows, skipped_missing_rate, skipped_missing_unit


def preview_priced_boq_import(
    boq_path: str,
    region: str,
    source_label: str | None = None,
    config=None,
    column_overrides: dict[str, int] | None = None,
    ai_assist: bool = False,
    db_path: str | None = None,
) -> BoqImportPreview:
    """Preview priced BOQ extraction and style-learning output without writing to the database."""

    extracted_rows, skipped_missing_rate, skipped_missing_unit = _extract_priced_boq_rows(
        boq_path=boq_path,
        region=region,
        source_label=source_label,
        config=config,
        column_overrides=column_overrides,
    )
    _, alias_preview = _collect_priced_boq_aliases(
        extracted_rows,
        ai_assist=ai_assist,
        config=config,
    )
    preview = BoqImportPreview(
        source_file=boq_path,
        region=normalize_region_name(region),
        total_extracted=len(extracted_rows),
        skipped_missing_rate=skipped_missing_rate,
        skipped_missing_unit=skipped_missing_unit,
        extracted_rows=extracted_rows,
        alias_preview=alias_preview,
    )
    if db_path:
        append_preview, candidate_preview, duplicate_preview = _classify_structured_import_rows(
            db_path=db_path,
            target_sheet="RateLibrary",
            mapped_rows=extracted_rows,
        )
        preview.append_preview = append_preview
        preview.candidate_preview = candidate_preview
        preview.duplicate_preview = duplicate_preview
        preview.append_count = len(append_preview)
        preview.candidate_count = len(candidate_preview)
        preview.duplicate_count = len(duplicate_preview)
        preview.notes.append(
            f"Dry run decision: {preview.append_count} row(s) would append, "
            f"{preview.candidate_count} would go to CandidateMatches, and "
            f"{preview.duplicate_count} would be skipped as same-rate duplicates."
        )
    if skipped_missing_rate:
        preview.notes.append(f"Skipped {skipped_missing_rate} BOQ row(s) without a usable rate.")
    if skipped_missing_unit:
        preview.notes.append(f"Skipped {skipped_missing_unit} BOQ row(s) without a usable unit.")
    if ai_assist:
        preview.notes.append("AI-assisted alias suggestions are shown when an embedding provider is available.")
    return preview


def import_priced_boq(
    db_path: str,
    boq_path: str,
    region: str,
    source_label: str | None = None,
    config=None,
    column_overrides: dict[str, int] | None = None,
    ai_assist: bool = False,
) -> ImportSummary:
    """Import priced BOQ rows into the review-first rate library workflow."""
    normalized_region = normalize_region_name(region)
    runtime_config = config or load_config()
    boq_source = Path(boq_path)
    extracted_rows, skipped_missing_rate, skipped_missing_unit = _extract_priced_boq_rows(
        boq_path=boq_path,
        region=region,
        source_label=source_label,
        config=runtime_config,
        column_overrides=column_overrides,
    )

    with tempfile.NamedTemporaryFile("w", encoding="utf-8", newline="", suffix=".csv", delete=False) as handle:
        fieldnames = [
            "item_code", "description", "normalized_description", "section", "subsection", "unit", "rate", "currency",
            "region", "source", "source_sheet", "source_page", "basis", "crew_type", "plant_type", "material_type",
            "keywords", "alias_group", "build_up_recipe_id", "confidence_hint", "notes", "active",
        ]
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in extracted_rows:
            writer.writerow(row)
        temp_csv = Path(handle.name)

    try:
        summary = import_structured_rows(
            db_path=db_path,
            input_path=str(temp_csv),
            target_sheet="RateLibrary",
            mapper=build_rate_library_row,
            defaults={"region": normalized_region, "source": source_label.strip() if source_label else boq_source.stem},
        )
    finally:
        temp_csv.unlink(missing_ok=True)

    aliases_by_description, _ = _collect_priced_boq_aliases(
        summary.appended_records,
        ai_assist=ai_assist,
        config=runtime_config,
    )

    alias_count = _sync_aliases_to_excel(db_path, aliases_by_description)
    if alias_count:
        summary.notes.append(f"Added {alias_count} BOQ-derived alias row(s) to Aliases.")

    schema_count = _sync_priced_boq_records_to_schema(
        db_path=db_path,
        boq_path=boq_path,
        appended_records=summary.appended_records,
        aliases_by_description=aliases_by_description,
        ai_assist=ai_assist,
        config=runtime_config,
    )
    if schema_count:
        summary.notes.append(f"Synced {schema_count} imported BOQ row(s) into the normalized schema.")

    summary.total_rows = len(extracted_rows)
    if skipped_missing_rate:
        summary.notes.append(f"Skipped {skipped_missing_rate} BOQ row(s) without a usable rate.")
    if skipped_missing_unit:
        summary.notes.append(f"Skipped {skipped_missing_unit} BOQ row(s) without a usable unit.")
    return summary


def normalize_database_units(db_path: str, sheets: list[str] | None = None) -> ImportSummary:
    """Normalize units, regions, and descriptions across database sheets."""
    workbook = load_workbook(db_path)
    target_sheets = sheets or ["RateLibrary", "BuildUpInputs", "CandidateMatches"]
    summary = ImportSummary(target_sheet="multiple", source_file=db_path)
    for sheet_name in target_sheets:
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        headers = workbook_headers(sheet)
        positions = {header: index + 1 for index, header in enumerate(headers)}
        for row_number in range(2, sheet.max_row + 1):
            if "unit" in positions:
                cell = sheet.cell(row_number, positions["unit"])
                normalized = normalize_unit(str(cell.value or ""))
                if normalized and normalized != str(cell.value or ""):
                    cell.value = normalized
                    summary.normalized_rows += 1
            if "region" in positions:
                cell = sheet.cell(row_number, positions["region"])
                normalized_region = normalize_region_name(str(cell.value or ""))
                if normalized_region and normalized_region != str(cell.value or ""):
                    cell.value = normalized_region
                    summary.normalized_rows += 1
            if "description" in positions and "normalized_description" in positions:
                description = str(sheet.cell(row_number, positions["description"]).value or "")
                sheet.cell(row_number, positions["normalized_description"]).value = normalize_text(description)
    workbook.save(db_path)
    return summary


def deduplicate_database(db_path: str, sheet_name: str = "RateLibrary") -> ImportSummary:
    """Deactivate duplicate rows based on item code and normalized description plus unit."""
    workbook = load_workbook(db_path)
    sheet = ensure_sheet_headers(workbook, sheet_name, RATE_LIBRARY_HEADERS if sheet_name == "RateLibrary" else BUILDUP_INPUT_HEADERS)
    ensure_sheet_headers(workbook, "ReviewLog", REVIEW_LOG_HEADERS)
    headers = workbook_headers(sheet)
    positions = {header: index + 1 for index, header in enumerate(headers)}
    seen_codes: set[str] = set()
    seen_desc_units: set[tuple[str, str]] = set()
    summary = ImportSummary(target_sheet=sheet_name, source_file=db_path)

    for row_number in range(2, sheet.max_row + 1):
        code_field = "item_code" if "item_code" in positions else "input_code"
        item_code = str(sheet.cell(row_number, positions[code_field]).value or "").strip()
        normalized_description = str(sheet.cell(row_number, positions.get("normalized_description", positions["description"])).value or "").strip()
        if "normalized_description" not in positions:
            normalized_description = normalize_text(normalized_description)
        unit = normalize_unit(str(sheet.cell(row_number, positions["unit"]).value or ""))
        duplicate = False
        if item_code and item_code in seen_codes:
            duplicate = True
        elif normalized_description and unit and (normalized_description, unit) in seen_desc_units:
            duplicate = True

        if duplicate and "active" in positions:
            sheet.cell(row_number, positions["active"]).value = False
            summary.skipped_duplicates += 1
            write_review_log(
                workbook,
                {
                    "timestamp": timestamp_now(),
                    "boq_file": db_path,
                    "sheet_name": sheet_name,
                    "row_number": row_number,
                    "boq_description": normalized_description,
                    "decision": "duplicate-deactivated",
                    "matched_item_code": item_code,
                    "matched_description": normalized_description,
                    "confidence_score": 100,
                    "reviewer_note": "Automated deduplication pass",
                },
            )
            continue

        if item_code:
            seen_codes.add(item_code)
        if normalized_description and unit:
            seen_desc_units.add((normalized_description, unit))

    workbook.save(db_path)
    return summary


def generate_review_report(db_path: str, output_json: str | None = None) -> ImportSummary:
    """Generate or refresh a Candidate Review sheet and optional training-style JSON."""
    workbook = load_workbook(db_path)
    candidate_sheet = ensure_sheet_headers(workbook, "CandidateMatches", CANDIDATE_MATCH_HEADERS)
    set_candidate_defaults(candidate_sheet)
    if "Candidate Review" in workbook.sheetnames:
        del workbook["Candidate Review"]
    review_sheet = workbook.create_sheet("Candidate Review")
    review_sheet.append(CANDIDATE_REVIEW_HEADERS)

    cpos = candidate_positions(candidate_sheet)
    summary = ImportSummary(target_sheet="Candidate Review", source_file=db_path)
    training_rows: list[dict[str, Any]] = []

    for row_number in range(2, candidate_sheet.max_row + 1):
        row_data = {header: candidate_sheet.cell(row_number, cpos[header]).value for header in cpos}
        review_sheet.append(
            [
                row_number,
                row_data.get("reviewer_status", "pending"),
                row_data.get("reviewer_name", ""),
                row_data.get("reviewed_at", ""),
                row_data.get("review_decision", "hold"),
                row_data.get("promote_target", "candidatematches"),
                row_data.get("approved_item_code", ""),
                row_data.get("approved_description", ""),
                row_data.get("approved_rate", ""),
                row_data.get("approved_canonical_term", ""),
                row_data.get("approved_section_bias", ""),
                row_data.get("confidence_override", ""),
                row_data.get("reviewer_note", ""),
                row_data.get("source_file", ""),
                row_data.get("target_sheet", ""),
                row_data.get("item_code", ""),
                row_data.get("matched_item_code", ""),
                row_data.get("description", ""),
                row_data.get("section", ""),
                row_data.get("subsection", ""),
                row_data.get("unit", ""),
                row_data.get("rate", ""),
                row_data.get("currency", ""),
                row_data.get("region", ""),
                row_data.get("basis", ""),
                row_data.get("duplicate_reason", ""),
                row_data.get("promotion_status", "not_promoted"),
            ]
        )
        summary.report_rows += 1
        status = str(row_data.get("reviewer_status") or "").lower()
        if status in {"approved", "merge", "approved_for_merge"}:
            summary.reviewed += 1
        if status in {"rejected", "reject"}:
            summary.rejected += 1
        if status and status != "pending":
            training_rows.append(
                {
                    "candidate_row_number": row_number,
                    "reviewer_status": row_data.get("reviewer_status", ""),
                    "reviewer_name": row_data.get("reviewer_name", ""),
                    "reviewed_at": row_data.get("reviewed_at", ""),
                    "review_decision": row_data.get("review_decision", ""),
                    "promote_target": row_data.get("promote_target", ""),
                    "description": row_data.get("description", ""),
                    "normalized_description": row_data.get("normalized_description", ""),
                    "unit": row_data.get("unit", ""),
                    "rate": row_data.get("rate", ""),
                    "region": row_data.get("region", ""),
                    "matched_item_code": row_data.get("matched_item_code", ""),
                    "approved_item_code": row_data.get("approved_item_code", ""),
                    "approved_description": row_data.get("approved_description", ""),
                    "approved_canonical_term": row_data.get("approved_canonical_term", ""),
                    "confidence_hint": row_data.get("confidence_hint", ""),
                    "confidence_override": row_data.get("confidence_override", ""),
                    "reviewer_note": row_data.get("reviewer_note", ""),
                }
            )

    workbook.save(db_path)
    if output_json:
        Path(output_json).parent.mkdir(parents=True, exist_ok=True)
        Path(output_json).write_text(json.dumps(training_rows, indent=2, ensure_ascii=True), encoding="utf-8")
        summary.training_records = len(training_rows)
    return summary


def merge_reviewed_candidates(db_path: str, reviewer_name: str | None = None) -> ImportSummary:
    """Merge decisions from Candidate Review back into CandidateMatches and ReviewLog."""
    workbook = load_workbook(db_path)
    candidate_sheet = ensure_sheet_headers(workbook, "CandidateMatches", CANDIDATE_MATCH_HEADERS)
    set_candidate_defaults(candidate_sheet)
    if "Candidate Review" not in workbook.sheetnames:
        raise ValueError("Candidate Review sheet does not exist. Run review-report first.")
    review_sheet = workbook["Candidate Review"]
    ensure_sheet_headers(workbook, "ReviewLog", REVIEW_LOG_HEADERS)

    cpos = candidate_positions(candidate_sheet)
    rpos = {header: index + 1 for index, header in enumerate(workbook_headers(review_sheet))}
    summary = ImportSummary(target_sheet="CandidateMatches", source_file=db_path)

    for row_number in range(2, review_sheet.max_row + 1):
        candidate_row_number = int(safe_float(review_sheet.cell(row_number, rpos["candidate_row_number"]).value) or 0)
        if candidate_row_number < 2 or candidate_row_number > candidate_sheet.max_row:
            continue
        reviewer_status = str(review_sheet.cell(row_number, rpos["reviewer_status"]).value or "pending").strip()
        if not reviewer_status or reviewer_status.lower() == "pending":
            continue

        reviewed_at = str(review_sheet.cell(row_number, rpos["reviewed_at"]).value or "").strip() or timestamp_now()
        reviewer = str(review_sheet.cell(row_number, rpos["reviewer_name"]).value or reviewer_name or "").strip()
        review_decision = str(review_sheet.cell(row_number, rpos["review_decision"]).value or "hold").strip()
        promote_target = str(review_sheet.cell(row_number, rpos["promote_target"]).value or "candidatematches").strip().lower()
        approved_item_code = str(review_sheet.cell(row_number, rpos["approved_item_code"]).value or "").strip()
        approved_description = str(review_sheet.cell(row_number, rpos["approved_description"]).value or "").strip()
        approved_rate = review_sheet.cell(row_number, rpos["approved_rate"]).value
        approved_canonical_term = str(review_sheet.cell(row_number, rpos["approved_canonical_term"]).value or "").strip()
        approved_section_bias = str(review_sheet.cell(row_number, rpos["approved_section_bias"]).value or "").strip()
        confidence_override = review_sheet.cell(row_number, rpos["confidence_override"]).value
        reviewer_note = str(review_sheet.cell(row_number, rpos["reviewer_note"]).value or "").strip()

        updates = {
            "reviewer_status": reviewer_status.lower(),
            "reviewer_name": reviewer,
            "reviewed_at": reviewed_at,
            "review_decision": review_decision.lower(),
            "promote_target": promote_target,
            "approved_item_code": approved_item_code,
            "approved_description": approved_description,
            "approved_rate": approved_rate,
            "approved_canonical_term": approved_canonical_term,
            "approved_section_bias": approved_section_bias,
            "confidence_override": confidence_override,
            "reviewer_note": reviewer_note,
        }
        for field, value in updates.items():
            if field in cpos:
                candidate_sheet.cell(candidate_row_number, cpos[field]).value = value

        decision_label = "review-approved" if reviewer_status.lower() in {"approved", "merge", "approved_for_merge"} else "review-rejected"
        if reviewer_status.lower() in {"rejected", "reject"}:
            summary.rejected += 1
        else:
            summary.reviewed += 1
        write_review_log(
            workbook,
            {
                "timestamp": reviewed_at,
                "boq_file": db_path,
                "sheet_name": "CandidateMatches",
                "row_number": candidate_row_number,
                "boq_description": str(candidate_sheet.cell(candidate_row_number, cpos["description"]).value or ""),
                "decision": decision_label,
                "matched_item_code": str(candidate_sheet.cell(candidate_row_number, cpos["matched_item_code"]).value or ""),
                "matched_description": approved_description or str(candidate_sheet.cell(candidate_row_number, cpos["description"]).value or ""),
                "confidence_score": safe_float(confidence_override) or safe_float(candidate_sheet.cell(candidate_row_number, cpos["confidence_hint"]).value) or 0.0,
                "reviewer_note": reviewer_note or review_decision,
            },
        )

    workbook.save(db_path)
    return summary


def _promote_to_aliases(workbook: Workbook, row: dict[str, Any]) -> bool:
    """Promote an approved review row into Aliases."""
    sheet = ensure_sheet_headers(workbook, "Aliases", ALIASES_HEADERS)
    alias = str(row.get("description") or "").strip()
    canonical_term = str(row.get("approved_canonical_term") or row.get("approved_description") or row.get("description") or "").strip()
    if not alias or not canonical_term:
        return False
    existing = {(str(values[0] or "").strip().lower(), str(values[1] or "").strip().lower()) for values in sheet.iter_rows(min_row=2, values_only=True)}
    key = (alias.lower(), canonical_term.lower())
    if key in existing:
        return False
    sheet.append([alias, canonical_term, str(row.get("approved_section_bias") or row.get("section") or ""), str(row.get("reviewer_note") or "Promoted from review")])
    return True


def _promote_to_rate_library(workbook: Workbook, row: dict[str, Any]) -> bool:
    """Promote an approved review row into RateLibrary."""
    sheet = ensure_sheet_headers(workbook, "RateLibrary", RATE_LIBRARY_HEADERS)
    description = str(row.get("approved_description") or row.get("description") or "").strip()
    if not description:
        return False
    item_code = str(row.get("approved_item_code") or row.get("item_code") or "").strip()
    normalized_description = normalize_text(description)
    unit = normalize_unit(str(row.get("unit") or ""))
    existing_keys = {
        (
            str(values[0] or "").strip(),
            normalize_text(str(values[2] or values[1] or "")),
            normalize_unit(str(values[5] or "")),
        )
        for values in sheet.iter_rows(min_row=2, values_only=True)
    }
    key = (item_code, normalized_description, unit)
    if key in existing_keys:
        return False
    sheet.append(
        [
            item_code,
            description,
            normalized_description,
            str(row.get("section") or ""),
            str(row.get("subsection") or ""),
            unit,
            safe_float(row.get("approved_rate")) or safe_float(row.get("rate")) or 0.0,
            str(row.get("currency") or "KES"),
            normalize_region_name(str(row.get("region") or "")),
            str(row.get("source") or "Reviewed Match"),
            str(row.get("source_sheet") or "Candidate Review"),
            str(row.get("source_page") or ""),
            str(row.get("basis") or "Estimator-reviewed promotion"),
            str(row.get("crew_type") or ""),
            str(row.get("plant_type") or ""),
            str(row.get("material_type") or ""),
            str(row.get("keywords") or ""),
            str(row.get("alias_group") or ""),
            str(row.get("build_up_recipe_id") or ""),
            safe_float(row.get("confidence_override")) or safe_float(row.get("confidence_hint")) or 0.0,
            str(row.get("reviewer_note") or ""),
            True,
        ]
    )
    return True


def _retain_in_candidates(sheet, row_number: int) -> bool:
    """Mark an approved record as intentionally retained in CandidateMatches."""
    positions = candidate_positions(sheet)
    if "promotion_status" in positions:
        sheet.cell(row_number, positions["promotion_status"]).value = "retained_in_candidates"
    if "promoted_at" in positions:
        sheet.cell(row_number, positions["promoted_at"]).value = timestamp_now()
    return True


def promote_approved_candidates(db_path: str, output_json: str | None = None) -> ImportSummary:
    """Promote approved reviewed records into the selected live destination and export learning JSON."""
    workbook = load_workbook(db_path)
    candidate_sheet = ensure_sheet_headers(workbook, "CandidateMatches", CANDIDATE_MATCH_HEADERS)
    set_candidate_defaults(candidate_sheet)
    ensure_sheet_headers(workbook, "ReviewLog", REVIEW_LOG_HEADERS)
    summary = ImportSummary(target_sheet="CandidateMatches", source_file=db_path)
    cpos = candidate_positions(candidate_sheet)
    training_rows: list[dict[str, Any]] = []

    for row_number in range(2, candidate_sheet.max_row + 1):
        status = str(candidate_sheet.cell(row_number, cpos["reviewer_status"]).value or "").strip().lower()
        if status not in {"approved", "merge", "approved_for_merge"}:
            continue
        promotion_status = str(candidate_sheet.cell(row_number, cpos["promotion_status"]).value or "").strip().lower()
        if promotion_status in {"promoted", "retained_in_candidates"}:
            continue

        row = {header: candidate_sheet.cell(row_number, cpos[header]).value for header in cpos}
        target = str(row.get("promote_target") or "candidatematches").strip().lower()
        promoted = False
        if target == "aliases":
            promoted = _promote_to_aliases(workbook, row)
        elif target == "ratelibrary":
            promoted = _promote_to_rate_library(workbook, row)
        else:
            promoted = _retain_in_candidates(candidate_sheet, row_number)

        promoted_at = timestamp_now()
        if target == "candidatematches":
            promotion_status = "retained_in_candidates"
        else:
            promotion_status = "promoted" if promoted else "already_exists"
        candidate_sheet.cell(row_number, cpos["promotion_status"]).value = promotion_status
        candidate_sheet.cell(row_number, cpos["promoted_at"]).value = promoted_at
        summary.promoted += 1 if promoted else 0
        if not promoted and target != "candidatematches":
            summary.notes.append(f"Skipped duplicate promotion for candidate row {row_number} to {target}.")
        write_review_log(
            workbook,
            {
                "timestamp": promoted_at,
                "boq_file": db_path,
                "sheet_name": target if target != "candidatematches" else "CandidateMatches",
                "row_number": row_number,
                "boq_description": str(row.get("description") or ""),
                "decision": f"promoted-to-{target}",
                "matched_item_code": str(row.get("approved_item_code") or row.get("matched_item_code") or ""),
                "matched_description": str(row.get("approved_description") or row.get("description") or ""),
                "confidence_score": safe_float(row.get("confidence_override")) or safe_float(row.get("confidence_hint")) or 0.0,
                "reviewer_note": str(row.get("reviewer_note") or "Promoted approved review"),
            },
        )
        training_rows.append(
            {
                "timestamp": promoted_at,
                "candidate_row_number": row_number,
                "reviewer_status": row.get("reviewer_status", ""),
                "reviewer_name": row.get("reviewer_name", ""),
                "review_decision": row.get("review_decision", ""),
                "promote_target": target,
                "promotion_status": candidate_sheet.cell(row_number, cpos["promotion_status"]).value,
                "description": row.get("description", ""),
                "normalized_description": row.get("normalized_description", ""),
                "unit": row.get("unit", ""),
                "rate": row.get("rate", ""),
                "approved_rate": row.get("approved_rate", ""),
                "matched_item_code": row.get("matched_item_code", ""),
                "approved_item_code": row.get("approved_item_code", ""),
                "approved_description": row.get("approved_description", ""),
                "approved_canonical_term": row.get("approved_canonical_term", ""),
                "confidence_hint": row.get("confidence_hint", ""),
                "confidence_override": row.get("confidence_override", ""),
                "reviewer_note": row.get("reviewer_note", ""),
            }
        )

    workbook.save(db_path)
    if output_json:
        Path(output_json).parent.mkdir(parents=True, exist_ok=True)
        Path(output_json).write_text(json.dumps(training_rows, indent=2, ensure_ascii=True), encoding="utf-8")
        summary.training_records = len(training_rows)
    return summary


def merge_candidate_matches(db_path: str) -> ImportSummary:
    """Merge approved CandidateMatches entries into their target sheets and log the action."""
    workbook = load_workbook(db_path)
    candidate_sheet = ensure_sheet_headers(workbook, "CandidateMatches", CANDIDATE_MATCH_HEADERS)
    ensure_sheet_headers(workbook, "ReviewLog", REVIEW_LOG_HEADERS)
    summary = ImportSummary(target_sheet="CandidateMatches", source_file=db_path)
    headers = workbook_headers(candidate_sheet)
    positions = {header: index + 1 for index, header in enumerate(headers)}

    for row_number in range(2, candidate_sheet.max_row + 1):
        status = str(candidate_sheet.cell(row_number, positions["reviewer_status"]).value or "").strip().lower()
        if status not in {"approved", "approve_append", "approve_update"}:
            continue

        target_sheet_name = str(candidate_sheet.cell(row_number, positions["target_sheet"]).value or "RateLibrary")
        target_headers = RATE_LIBRARY_HEADERS if target_sheet_name == "RateLibrary" else BUILDUP_INPUT_HEADERS
        target_sheet = ensure_sheet_headers(workbook, target_sheet_name, target_headers)
        values = {header: candidate_sheet.cell(row_number, positions[header]).value for header in headers if header in positions}

        if target_sheet_name == "RateLibrary":
            row_values = {
                "item_code": values.get("item_code", ""),
                "description": values.get("description", ""),
                "normalized_description": values.get("normalized_description", ""),
                "section": values.get("section", ""),
                "subsection": values.get("subsection", ""),
                "unit": values.get("unit", ""),
                "rate": values.get("rate", ""),
                "currency": values.get("currency", "KES"),
                "region": values.get("region", ""),
                "source": values.get("source", ""),
                "source_sheet": values.get("source_sheet", ""),
                "source_page": values.get("source_page", ""),
                "basis": values.get("basis", ""),
                "crew_type": values.get("crew_type", ""),
                "plant_type": values.get("plant_type", ""),
                "material_type": values.get("material_type", ""),
                "keywords": values.get("keywords", ""),
                "alias_group": values.get("alias_group", ""),
                "build_up_recipe_id": values.get("build_up_recipe_id", ""),
                "confidence_hint": values.get("confidence_hint", 0.0),
                "notes": values.get("notes", ""),
                "active": values.get("active", True),
            }
        else:
            row_values = {
                "input_code": values.get("item_code", ""),
                "input_type": values.get("material_type") or values.get("plant_type") or values.get("crew_type") or values.get("notes", ""),
                "description": values.get("description", ""),
                "unit": values.get("unit", ""),
                "rate": values.get("rate", ""),
                "region": values.get("region", ""),
                "source": values.get("source", ""),
                "active": values.get("active", True),
            }

        append_row(target_sheet, target_headers, row_values)
        candidate_sheet.cell(row_number, positions["reviewer_status"]).value = "merged"
        write_review_log(
            workbook,
            {
                "timestamp": timestamp_now(),
                "boq_file": db_path,
                "sheet_name": target_sheet_name,
                "row_number": target_sheet.max_row,
                "boq_description": str(values.get("description") or ""),
                "decision": "candidate-merged",
                "matched_item_code": str(values.get("matched_item_code") or values.get("item_code") or ""),
                "matched_description": str(values.get("description") or ""),
                "confidence_score": safe_float(values.get("confidence_hint")) or 0.0,
                "reviewer_note": str(values.get("reviewer_note") or "Merged from CandidateMatches"),
            },
        )
        summary.merged += 1

    workbook.save(db_path)
    return summary
