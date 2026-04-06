"""Audit and unmatched export helpers."""

from __future__ import annotations

import csv
from dataclasses import asdict, is_dataclass
from pathlib import Path

from openpyxl import load_workbook

from .models import MatchResult
from .utils import dump_json, ensure_parent


def _json_safe(value):
    """Convert dataclasses recursively for JSON output."""
    if is_dataclass(value):
        return asdict(value)
    return value


def write_audit_json(path: Path, results: list[MatchResult], metadata: dict) -> Path:
    """Write a JSON audit trail for a pricing run."""
    payload = {
        "metadata": {key: _json_safe(value) for key, value in metadata.items()},
        "results": [
            {
                "sheet_name": item.boq_line.sheet_name,
                "row_number": item.boq_line.row_number,
                "description": item.boq_line.description,
                "unit": item.boq_line.unit,
                "quantity": item.boq_line.quantity,
                "decision": item.decision,
                "matched_item_code": item.matched_item_code,
                "matched_description": item.matched_description,
                "base_rate": item.base_rate,
                "rate": item.rate,
                "confidence_score": item.confidence_score,
                "confidence_band": item.confidence_band,
                "review_flag": item.review_flag,
                "flag_reasons": item.flag_reasons,
                "generic_match_flag": item.generic_match_flag,
                "category_mismatch_flag": item.category_mismatch_flag,
                "section_mismatch_flag": item.section_mismatch_flag,
                "built_up": item.built_up,
                "basis_of_rate": item.basis_of_rate,
                "approval_status": item.approval_status,
                "commercial_review_flags": item.commercial_review_flags,
                "alternate_options": item.alternate_options,
                "regional_factor": item.regional_factor,
                "rationale": item.rationale,
            }
            for item in results
        ],
    }
    dump_json(path, payload)
    return path


def export_unmatched_csv(path: Path, results: list[MatchResult]) -> Path:
    """Export unmatched or review rows into CSV."""
    ensure_parent(path)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            [
                "sheet_name",
                "row_number",
                "boq_description",
                "boq_unit",
                "decision",
                "matched_item_code",
                "matched_description",
                "confidence_score",
                "confidence_band",
                "review_flag",
                "approval_status",
                "flag_reasons",
                "commercial_review_flags",
            ]
        )
        for result in results:
            if result.decision == "matched" and not result.review_flag:
                continue
            writer.writerow(
                [
                    result.boq_line.sheet_name,
                    result.boq_line.row_number,
                    result.boq_line.description,
                    result.boq_line.unit,
                    result.decision,
                    result.matched_item_code,
                    result.matched_description,
                    result.confidence_score,
                    result.confidence_band,
                    result.review_flag,
                    result.approval_status,
                    "; ".join(result.flag_reasons),
                    "; ".join(result.commercial_review_flags),
                ]
            )
    return path


def export_unmatched_from_workbook(input_workbook: str, output_csv: str) -> Path:
    """Export unmatched items from a priced workbook's Match Suggestions sheet."""
    workbook = load_workbook(input_workbook, data_only=True)
    if "Match Suggestions" not in workbook.sheetnames:
        raise ValueError("The input workbook does not contain a 'Match Suggestions' sheet.")

    sheet = workbook["Match Suggestions"]
    output = Path(output_csv)
    ensure_parent(output)
    with output.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        headers = [cell.value for cell in sheet[1]]
        writer.writerow(headers)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            decision = str(row[5] or "")
            review_flag = str(row[12] or "")
            if decision != "matched" or review_flag.upper() == "YES":
                writer.writerow(row)
    return output
