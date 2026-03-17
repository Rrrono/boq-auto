"""Excel BOQ workbook reading and column detection."""

from __future__ import annotations

import logging
from collections import Counter

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .models import BOQLine, ColumnMap, SheetData
from .normalizer import normalize_text
from .utils import safe_float

SHEET_SKIP_PATTERNS = {
    "summary",
    "abstract",
    "collection",
    "recap",
    "quotation summary",
    "match suggestions",
    "candidate review",
    "commercial review",
    "basis of rates",
    "assumptions",
    "exclusions",
}

HEADING_HINTS = {
    "preliminaries",
    "dayworks",
    "earthworks",
    "concrete",
    "finishes",
    "general items",
    "contractor equipment",
    "contractors equipment",
}

TOTAL_PATTERNS = {"subtotal", "sub total", "total", "carried forward", "brought forward", "sum"}


class WorkbookReader:
    """Read BOQ rows from Excel while preserving workbook structure."""

    def __init__(self, config, logger: logging.Logger | None = None) -> None:
        self.config = config
        self.logger = logger or logging.getLogger("boq_auto")

    def read(self, workbook_path: str, column_overrides: dict[str, int] | None = None) -> list[SheetData]:
        """Read a BOQ workbook into structured sheet data."""
        workbook = load_workbook(workbook_path)
        sheets: list[SheetData] = []
        for worksheet in workbook.worksheets:
            classification = self.classify_sheet(worksheet)
            if classification in {"skip", "summary"}:
                self.logger.info("Skipping sheet '%s' classified as %s", worksheet.title, classification)
                continue
            columns = self.detect_columns(worksheet, column_overrides or {})
            rows = self.read_rows(worksheet, columns)
            if not rows and classification != "boq":
                self.logger.info("Skipping sheet '%s' because no BOQ rows were detected", worksheet.title)
                continue
            sheets.append(SheetData(sheet_name=worksheet.title, columns=columns, classification=classification, rows=rows))
        return sheets

    def classify_sheet(self, worksheet: Worksheet) -> str:
        """Classify a sheet as BOQ, summary, or skip."""
        title = normalize_text(worksheet.title)
        if any(pattern in title for pattern in SHEET_SKIP_PATTERNS):
            return "summary"
        if any(pattern in title for pattern in {"boq", "bill", "dayworks", "preliminaries", "earthworks", "general items", "contractor equipment", "contractors equipment"}):
            return "boq"

        sample_text = normalize_text(" ".join(filter(None, [self._row_text(worksheet, row) for row in range(1, min(worksheet.max_row, 8) + 1)])))
        if any(pattern in sample_text for pattern in SHEET_SKIP_PATTERNS):
            return "summary"
        return "boq"

    def detect_columns(self, worksheet: Worksheet, overrides: dict[str, int]) -> ColumnMap:
        """Detect column positions from a likely header row."""
        max_scan = int(self.config.get("processing.max_header_scan_rows", 20))
        keywords = {
            "description_col": self.config.get("columns.description_keywords", []),
            "unit_col": self.config.get("columns.unit_keywords", []),
            "quantity_col": self.config.get("columns.quantity_keywords", []),
            "rate_col": self.config.get("columns.rate_keywords", []),
            "amount_col": self.config.get("columns.amount_keywords", []),
        }

        best_map = ColumnMap(header_row=1)
        best_score = -1.0
        for row_number in range(1, min(worksheet.max_row, max_scan) + 1):
            candidate = ColumnMap(header_row=row_number)
            score = 0.0
            row_values = [self._cell_text(worksheet, row_number, column_number) for column_number in range(1, worksheet.max_column + 1)]
            normalized_row = [normalize_text(value) for value in row_values]
            for column_number, value in enumerate(normalized_row, start=1):
                for field_name, terms in keywords.items():
                    if getattr(candidate, field_name) is not None:
                        continue
                    if any(normalize_text(term) == value or normalize_text(term) in value for term in terms):
                        setattr(candidate, field_name, column_number)
                        score += 3
                if value in {"item description", "description", "particulars"}:
                    score += 2
                if value in {"qty", "quantity"}:
                    score += 2
                if value in {"rate", "amount"}:
                    score += 2

            next_rows_numeric = self._numeric_density(worksheet, row_number + 1, row_number + 4)
            score += next_rows_numeric
            if candidate.description_col is not None:
                score += 2
            if score > best_score:
                best_score = score
                best_map = candidate

        for key, value in overrides.items():
            if hasattr(best_map, key) and value:
                setattr(best_map, key, value)
        inferred = self._infer_missing_columns(worksheet, best_map)
        return inferred

    def read_rows(self, worksheet: Worksheet, columns: ColumnMap) -> list[BOQLine]:
        """Read BOQ item rows below the detected header row."""
        results: list[BOQLine] = []
        blank_streak = 0
        for row_number in range(columns.header_row + 1, worksheet.max_row + 1):
            description = self._cell_text(worksheet, row_number, columns.description_col)
            unit = self._cell_text(worksheet, row_number, columns.unit_col)
            quantity = self._cell_float(worksheet, row_number, columns.quantity_col)
            rate = self._cell_float(worksheet, row_number, columns.rate_col)
            amount = self._cell_float(worksheet, row_number, columns.amount_col)
            row_text = normalize_text(" ".join(part for part in [description, unit] if part))

            if not any([description, unit, quantity, rate, amount]):
                blank_streak += 1
                if blank_streak >= 5 and results:
                    break
                continue
            blank_streak = 0

            is_subtotal = any(pattern in row_text for pattern in {"subtotal", "sub total", "carried forward", "brought forward"})
            is_total = any(pattern in row_text for pattern in {"grand total", "total", "contract sum"})
            is_summary_row = is_subtotal or is_total
            if self._is_noise_row(row_text):
                continue

            is_heading = self._is_heading_row(description, unit, quantity, rate, amount, row_text)
            if is_summary_row and amount is None and quantity is None and rate is None:
                continue

            results.append(
                BOQLine(
                    sheet_name=worksheet.title,
                    row_number=row_number,
                    description=description,
                    unit=unit,
                    quantity=quantity,
                    rate=rate,
                    amount=amount,
                    is_heading=is_heading,
                    is_subtotal=is_subtotal,
                    is_total=is_total,
                    is_summary_row=is_summary_row,
                )
            )
        return results

    def _infer_missing_columns(self, worksheet: Worksheet, columns: ColumnMap) -> ColumnMap:
        """Infer missing columns from data patterns in non-standard layouts."""
        sample_rows = list(range(columns.header_row + 1, min(worksheet.max_row, columns.header_row + 8) + 1))
        numeric_counts: Counter[int] = Counter()
        text_counts: Counter[int] = Counter()
        for row_number in sample_rows:
            for column_number in range(1, worksheet.max_column + 1):
                text = self._cell_text(worksheet, row_number, column_number)
                if text:
                    text_counts[column_number] += 1
                if self._cell_float(worksheet, row_number, column_number) is not None:
                    numeric_counts[column_number] += 1

        if columns.description_col is None:
            columns.description_col = text_counts.most_common(1)[0][0] if text_counts else 1
        if columns.quantity_col is None and numeric_counts:
            columns.quantity_col = numeric_counts.most_common(1)[0][0]
        if columns.amount_col is None and numeric_counts:
            columns.amount_col = numeric_counts.most_common(1)[-1][0]
        if columns.rate_col is None and numeric_counts:
            numeric_columns = [column for column, _ in numeric_counts.most_common()]
            if len(numeric_columns) >= 2:
                columns.rate_col = numeric_columns[1]
        if columns.unit_col is None:
            for column_number in range(1, worksheet.max_column + 1):
                values = [normalize_text(self._cell_text(worksheet, row_number, column_number)) for row_number in sample_rows]
                if any(value in {"m2", "m3", "nr", "sum", "day", "item", "ton"} for value in values):
                    columns.unit_col = column_number
                    break
        return columns

    def _numeric_density(self, worksheet: Worksheet, start_row: int, end_row: int) -> float:
        """Estimate how data-like the rows beneath a header candidate are."""
        score = 0.0
        for row_number in range(start_row, min(end_row, worksheet.max_row) + 1):
            numeric_cells = sum(
                1 for column_number in range(1, worksheet.max_column + 1)
                if self._cell_float(worksheet, row_number, column_number) is not None
            )
            score += min(numeric_cells, 3) * 0.5
        return score

    def _row_text(self, worksheet: Worksheet, row_number: int) -> str:
        return " ".join(self._cell_text(worksheet, row_number, column_number) for column_number in range(1, worksheet.max_column + 1)).strip()

    @staticmethod
    def _is_heading_row(description: str, unit: str, quantity: float | None, rate: float | None, amount: float | None, row_text: str) -> bool:
        if any(pattern in row_text for pattern in HEADING_HINTS) and quantity is None and rate is None and amount is None:
            return True
        return bool(description and not unit and quantity is None and rate is None and amount is None)

    @staticmethod
    def _is_noise_row(row_text: str) -> bool:
        return not row_text or row_text in {"page", "continued"} or any(pattern == row_text for pattern in TOTAL_PATTERNS)

    @staticmethod
    def _merged_anchor_value(worksheet: Worksheet, row: int, column: int) -> str:
        for merged_range in worksheet.merged_cells.ranges:
            if merged_range.min_row <= row <= merged_range.max_row and merged_range.min_col <= column <= merged_range.max_col:
                return str(worksheet.cell(merged_range.min_row, merged_range.min_col).value or "").strip()
        return ""

    def _cell_text(self, worksheet: Worksheet, row: int, column: int | None) -> str:
        if not column:
            return ""
        value = worksheet.cell(row, column).value
        if value in (None, ""):
            merged_value = self._merged_anchor_value(worksheet, row, column)
            if merged_value:
                return merged_value
        return str(value or "").strip()

    def _cell_float(self, worksheet: Worksheet, row: int, column: int | None) -> float | None:
        if not column:
            return None
        value = worksheet.cell(row, column).value
        if value in (None, ""):
            merged_value = self._merged_anchor_value(worksheet, row, column)
            if merged_value:
                return safe_float(merged_value)
        return safe_float(value)
