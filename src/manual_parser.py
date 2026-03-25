"""Cost manual ingestion helpers for PDF and Excel sources."""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
import csv
import re
import tempfile
from typing import Any

from openpyxl import load_workbook

from src.cost_schema import CostDatabase, build_cost_item, schema_database_path
from src.ingestion import ALIASES_HEADERS, ImportSummary, append_row, ensure_sheet_headers, import_structured_rows
from src.normalizer import normalize_text, normalize_unit
from src.pdf_ingestion import extract_text_from_pdf
from src.utils import safe_float


UNIT_PATTERN = r"(m2|m3|m|nr|sum|item|kg|km|day|hr|ton|sm|cm|no|nos)"
RATE_TOKEN_PATTERN = re.compile(r"(?P<value>\d[\d,]*(?:\.\d+)?)")
MATERIAL_KEYWORDS = {
    "concrete": ["concrete", "blinding", "screed"],
    "steel": ["reinforcement", "rebar", "steel"],
    "pipe": ["pipe", "hdpe", "pvc", "culvert"],
    "soil": ["fill", "backfill", "excavation", "earth"],
    "paint": ["paint", "coating"],
}


@dataclass(slots=True)
class ManualItem:
    code: str
    item_name: str
    unit: str
    description: str
    rate: float = 0.0
    category: str = ""
    material: str = ""
    keywords: list[str] = field(default_factory=list)
    aliases: list[str] = field(default_factory=list)


def load_manual_pdf(path: str) -> str:
    return extract_text_from_pdf(path)


def parse_manual_text(text: str) -> list[ManualItem]:
    """Extract likely manual rows from PDF text."""

    items: list[ManualItem] = []
    pending_line = ""
    for raw_line in text.splitlines():
        line = " ".join(str(raw_line).split())
        if not line:
            continue
        match = _match_manual_item_line(line)
        if not match and pending_line:
            match = _match_manual_item_line(f"{pending_line} {line}")
            if match:
                line = f"{pending_line} {line}"
                pending_line = ""
        if match:
            description = _clean_description(str(match.group("description") or "").strip())
            rates_text = str(match.group("rates") or "")
            rate = _extract_rate(rates_text)
            code = str(match.group("code") or "").strip()
            if not description or (not code and rate is None) or (not code and not _looks_like_rate_block(rates_text)) or not _is_plausible_manual_item(description, code):
                pending_line = line
                continue
            items.append(
                ManualItem(
                    code=code,
                    item_name=description,
                    unit=normalize_unit(str(match.group("unit") or "")),
                    description=description,
                    rate=rate or 0.0,
                    category=_infer_category(description),
                    material=_infer_material(description),
                    keywords=_extract_keywords(description),
                )
            )
            continue
        if items and _looks_like_continuation(line):
            previous = items[-1]
            previous.description = f"{previous.description} {line}".strip()
            previous.item_name = previous.description
            pending_line = ""
            continue
        if _looks_like_item_prefix(line):
            pending_line = line
        else:
            pending_line = ""
    return items


def _match_manual_item_line(line: str):
    return re.match(
        rf"^(?:(?P<code>\d+(?:\.\d+){{0,4}})\s+)?(?P<description>.+?)\s+(?P<unit>{UNIT_PATTERN})\b(?:\s+(?P<rates>.+))?$",
        line,
        flags=re.IGNORECASE,
    )


def _extract_rate(rates_text: str) -> float | None:
    for match in RATE_TOKEN_PATTERN.finditer(str(rates_text or "")):
        value = safe_float(match.group("value").replace(",", ""))
        if value is not None:
            return value
    return None


def _clean_description(description: str) -> str:
    cleaned = re.sub(r"^[^\w]+", "", description or "").strip()
    return cleaned.rstrip(" .,:;")


def _looks_like_continuation(line: str) -> bool:
    stripped = str(line or "").strip()
    if not stripped:
        return False
    if stripped[:1].islower():
        return True
    return bool(re.match(r"^[\(\[\-.,;:/]", stripped))


def _looks_like_item_prefix(line: str) -> bool:
    stripped = str(line or "").strip()
    if not stripped:
        return False
    if re.match(r"^\d+(?:\.\d+){0,4}\s+", stripped):
        return True
    words = stripped.split()
    return len(words) >= 3 and any(char.isdigit() for char in stripped)


def _is_plausible_manual_item(description: str, code: str) -> bool:
    if code:
        return True
    normalized = " ".join(str(description or "").split())
    if len(normalized) > 180:
        return False
    if normalized.count(".") > 2:
        return False
    return len(normalized.split()) <= 24


def _looks_like_rate_block(rates_text: str) -> bool:
    text = str(rates_text or "").strip()
    if not text:
        return False
    numeric_tokens = RATE_TOKEN_PATTERN.findall(text)
    alpha_chars = sum(1 for char in text if char.isalpha())
    return len(numeric_tokens) >= 1 and alpha_chars <= 12


def load_manual_excel(path: str, sheet_name: str | None = None) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    headers = [str(cell.value or "").strip() for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    rows: list[dict[str, Any]] = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))
    return rows


def parse_manual_rows(rows: list[dict[str, Any]]) -> list[ManualItem]:
    items: list[ManualItem] = []
    for row in rows:
        description = str(
            row.get("description")
            or row.get("item_name")
            or row.get("item")
            or row.get("name")
            or ""
        ).strip()
        if not description:
            continue
        items.append(
            ManualItem(
                code=str(row.get("code") or row.get("item_code") or ""),
                item_name=description,
                unit=normalize_unit(str(row.get("unit") or row.get("uom") or "")),
                description=description,
                rate=safe_float(row.get("rate")) or 0.0,
                category=str(row.get("category") or row.get("section") or _infer_category(description)),
                material=str(row.get("material") or _infer_material(description)),
                keywords=_extract_keywords(description),
            )
        )
    return items


def extract_manual_items(input_path: str, sheet_name: str | None = None) -> list[ManualItem]:
    path = Path(input_path)
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return parse_manual_text(load_manual_pdf(str(path)))
    if suffix in {".xlsx", ".xlsm"}:
        return parse_manual_rows(load_manual_excel(str(path), sheet_name=sheet_name))
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            return parse_manual_rows(list(csv.DictReader(handle)))
    raise ValueError(f"Unsupported manual source type: {path.suffix}")


def ingest_manual_source(input_path: str, master_database_path: str, source_name: str | None = None, source_version: str = "v1") -> ImportSummary:
    items = extract_manual_items(input_path)
    return ingest_manual_to_database(items, master_database_path, source_name=source_name or Path(input_path).stem, source_file=input_path, source_version=source_version)


def ingest_manual_to_database(
    items: list[ManualItem],
    master_database_path: str,
    source_name: str = "Manual Review",
    source_file: str = "",
    source_version: str = "v1",
    ai_enhance: bool = False,
    embedding_provider=None,
) -> ImportSummary:
    """Append reviewed manual items to the Excel master DB and normalized schema sidecar."""

    db_path = Path(master_database_path)
    if not db_path.exists():
        raise FileNotFoundError(f"Master database not found: {db_path}")

    schema_db = CostDatabase(schema_database_path(db_path))
    source = schema_db.register_source(source_name, source_version, source_file or source_name)
    prepared_items = [enhance_manual_item(item, enable_ai=ai_enhance, provider=embedding_provider) for item in items]
    cost_items = [
        build_cost_item(
            code=item.code,
            description=item.description or item.item_name,
            unit=item.unit,
            category=item.category,
            subcategory="",
            material=item.material,
            keywords=item.keywords,
            rate=item.rate,
            source_id=source.id,
        )
        for item in prepared_items
    ]
    aliases_by_item = {cost_item.id: list(prepared_items[index].aliases) for index, cost_item in enumerate(cost_items)}
    schema_db.insert_items(cost_items, aliases_by_item=aliases_by_item)
    schema_db.log_ingestion(source.id, "completed", f"Ingested {len(cost_items)} reviewed item(s).")

    with tempfile.NamedTemporaryFile("w", encoding="utf-8", newline="", suffix=".csv", delete=False) as handle:
        writer = csv.DictWriter(handle, fieldnames=["item_code", "description", "unit", "rate", "section", "material", "keywords"])
        writer.writeheader()
        for item in prepared_items:
            writer.writerow(
                {
                    "item_code": item.code,
                    "description": item.description or item.item_name,
                    "unit": normalize_unit(item.unit),
                    "rate": item.rate,
                    "section": item.category,
                    "material": item.material,
                    "keywords": ", ".join(item.keywords),
                }
            )
        temp_csv = Path(handle.name)

    try:
        summary = import_structured_rows(
            db_path=str(db_path),
            input_path=str(temp_csv),
            target_sheet="RateLibrary",
            mapper=lambda row, source_file_name, source_sheet_name: {
                "item_code": str(row.get("item_code") or ""),
                "description": str(row.get("description") or ""),
                "normalized_description": normalize_text(str(row.get("description") or "")),
                "section": str(row.get("section") or ""),
                "subsection": "",
                "unit": normalize_unit(str(row.get("unit") or "")),
                "rate": safe_float(row.get("rate")) or 0.0,
                "currency": "KES",
                "region": "",
                "source": source_name,
                "source_sheet": source_sheet_name,
                "source_page": "",
                "basis": "Manual ingestion review",
                "crew_type": "",
                "plant_type": "",
                "material_type": str(row.get("material") or ""),
                "keywords": str(row.get("keywords") or ""),
                "alias_group": "",
                "build_up_recipe_id": "",
                "confidence_hint": 0.0,
                "notes": "",
                "active": True,
            },
        )
        alias_count = _sync_aliases_to_excel(db_path, prepared_items)
        if alias_count:
            summary.notes.append(f"Added {alias_count} alias row(s) to Aliases.")
        schema_db.log_ingestion(source.id, "completed", f"Excel append summary: appended={summary.appended} candidates={summary.candidates_created}")
        if embedding_provider is not None:
            for index, cost_item in enumerate(cost_items):
                embedding = embedding_provider.embed(_embedding_text_from_manual_item(prepared_items[index]))
                if embedding:
                    schema_db.save_embedding(cost_item.id, embedding, getattr(embedding_provider, "model_name", "unknown"))
        return summary
    finally:
        temp_csv.unlink(missing_ok=True)


def enhance_manual_item(item: ManualItem, enable_ai: bool = False, provider=None) -> ManualItem:
    """Apply rule-based parsing first, then optional AI-style enrichment."""

    updated = ManualItem(
        code=item.code,
        item_name=item.item_name,
        unit=normalize_unit(item.unit),
        description=item.description,
        rate=item.rate,
        category=item.category or _infer_category(item.description),
        material=item.material or _infer_material(item.description),
        keywords=item.keywords or _extract_keywords(item.description),
        aliases=list(item.aliases),
    )
    if not updated.aliases:
        updated.aliases = _rule_aliases(updated)
    if enable_ai:
        updated.category = updated.category or _suggest_category(updated.description)
        if provider is not None and hasattr(provider, "suggest_ingestion_attributes"):
            try:
                suggestion = provider.suggest_ingestion_attributes(
                    updated.description,
                    unit=updated.unit,
                    section=updated.category,
                )
            except Exception:
                suggestion = {}
            if isinstance(suggestion, dict):
                suggested_category = str(suggestion.get("category", "") or "").strip()
                suggested_material = str(suggestion.get("material", "") or "").strip()
                suggested_keywords = [str(value).strip() for value in suggestion.get("keywords", []) if str(value).strip()]
                suggested_aliases = [str(value).strip() for value in suggestion.get("aliases", []) if str(value).strip()]
                if suggested_category:
                    updated.category = suggested_category
                if suggested_material and not updated.material:
                    updated.material = suggested_material
                if suggested_keywords:
                    updated.keywords = list(dict.fromkeys([*updated.keywords, *suggested_keywords]))
                if suggested_aliases:
                    updated.aliases.extend([alias for alias in suggested_aliases if alias and alias not in updated.aliases])
        elif provider is not None and hasattr(provider, "suggest_aliases"):
            try:
                suggested = provider.suggest_aliases(updated.description)
                updated.aliases.extend([alias for alias in suggested if alias and alias not in updated.aliases])
            except Exception:
                pass
        else:
            updated.aliases.extend(_ai_style_aliases(updated))
        updated.aliases = list(dict.fromkeys(updated.aliases))
    return updated


def _infer_category(description: str) -> str:
    text = normalize_text(description)
    if "concrete" in text or "reinforcement" in text or "formwork" in text:
        return "concrete"
    if "excavat" in text or "fill" in text or "backfill" in text:
        return "earthworks"
    if "pipe" in text or "drain" in text or "manhole" in text:
        return "drainage"
    if "paint" in text or "tile" in text or "plaster" in text:
        return "finishes"
    return "general"


def _infer_material(description: str) -> str:
    text = normalize_text(description)
    for material, keywords in MATERIAL_KEYWORDS.items():
        if any(keyword in text for keyword in keywords):
            return material
    return ""


def _extract_keywords(description: str) -> list[str]:
    text = normalize_text(description)
    terms = [term for term in text.split() if len(term) > 3]
    return list(dict.fromkeys(terms[:8]))


def _rule_aliases(item: ManualItem) -> list[str]:
    aliases: list[str] = []
    text = normalize_text(item.description)
    if "backfill" in text:
        aliases.append("backfilling")
    if "reinforcement" in text:
        aliases.append("rebar")
    if item.material:
        aliases.append(item.material)
    return list(dict.fromkeys(aliases))


def _suggest_category(description: str) -> str:
    return _infer_category(description)


def _ai_style_aliases(item: ManualItem) -> list[str]:
    aliases = list(item.aliases)
    if item.material and item.material not in aliases:
        aliases.append(item.material)
    if item.category and item.category not in aliases:
        aliases.append(item.category)
    return aliases


def _embedding_text_from_manual_item(item: ManualItem) -> str:
    return " | ".join([item.category.strip(), item.material.strip(), item.description.strip(), item.unit.strip()])


def _sync_aliases_to_excel(db_path: Path, items: list[ManualItem]) -> int:
    workbook = load_workbook(db_path)
    alias_sheet = ensure_sheet_headers(workbook, "Aliases", ALIASES_HEADERS)
    existing = {
        (
            str(values[0] or "").strip().lower(),
            str(values[1] or "").strip().lower(),
        )
        for values in alias_sheet.iter_rows(min_row=2, values_only=True)
    }
    added = 0
    for item in items:
        canonical_term = (item.description or item.item_name or "").strip()
        if not canonical_term:
            continue
        for alias in item.aliases:
            alias_text = str(alias or "").strip()
            if not alias_text:
                continue
            key = (alias_text.lower(), canonical_term.lower())
            if key in existing:
                continue
            append_row(
                alias_sheet,
                ALIASES_HEADERS,
                {
                    "alias": alias_text,
                    "canonical_term": canonical_term,
                    "section_bias": item.category,
                    "notes": "Manual ingestion review",
                },
            )
            existing.add(key)
            added += 1
    if added:
        workbook.save(db_path)
    return added
