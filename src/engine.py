"""Core pricing engine orchestration."""

from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import load_workbook

from .audit import export_unmatched_csv, write_audit_json
from .buildup import price_build_up_recipe
from .commercial import resolve_commercial_terms, resolve_regional_factor, summarize_quote
from .matcher import Matcher, MatchingWeights
from .models import (
    AliasEntry,
    BuildUpInput,
    BuildUpRecipeLine,
    CommercialTerms,
    DatabaseBundle,
    MatchResult,
    QuotationSummary,
    RateItem,
    RegionalAdjustment,
    RunArtifacts,
    SectionRule,
    TextListEntry,
)
from .normalizer import normalize_text
from .section_inference import infer_section
from .utils import safe_float, truthy
from .workbook_reader import WorkbookReader
from .workbook_writer import WorkbookWriter


REQUIRED_SHEETS = {
    "RateLibrary",
    "Aliases",
    "SectionMap",
    "BuildUpInputs",
    "BuildUpRecipes",
    "Controls",
    "Rules",
    "ReviewLog",
}


class PricingEngine:
    """Run workbook pricing against the Excel database."""

    def __init__(self, config, logger: logging.Logger) -> None:
        self.config = config
        self.logger = logger

    def load_database(self, path: str) -> DatabaseBundle:
        """Load the Excel database workbook into in-memory models."""
        workbook = load_workbook(path, data_only=True)
        missing = REQUIRED_SHEETS - set(workbook.sheetnames)
        if missing:
            raise ValueError(f"Database is missing required sheets: {', '.join(sorted(missing))}")

        bundle = DatabaseBundle()
        bundle.rate_items = self._load_rate_library(workbook["RateLibrary"])
        bundle.aliases = self._load_aliases(workbook["Aliases"])
        bundle.section_rules = self._load_section_rules(workbook["SectionMap"])
        bundle.build_inputs = self._load_build_inputs(workbook["BuildUpInputs"])
        bundle.build_recipes = self._load_build_recipes(workbook["BuildUpRecipes"])
        if "RegionalAdjustments" in workbook.sheetnames:
            bundle.regional_adjustments = self._load_regional_adjustments(workbook["RegionalAdjustments"])
        if "Assumptions" in workbook.sheetnames:
            bundle.assumptions = self._load_text_entries(workbook["Assumptions"], "assumption")
        if "Exclusions" in workbook.sheetnames:
            bundle.exclusions = self._load_text_entries(workbook["Exclusions"], "exclusion")
        bundle.controls = self._load_key_values(workbook["Controls"])
        bundle.rules = self._load_generic_rows(workbook["Rules"])
        self.logger.info("Loaded %s active rate items", len(bundle.rate_items))
        return bundle

    def price_workbook(
        self,
        db_path: str,
        boq_path: str,
        output_path: str,
        region: str | None = None,
        threshold: float | None = None,
        apply_rates: bool | None = None,
        column_overrides: dict[str, int] | None = None,
    ) -> RunArtifacts:
        """Price a BOQ workbook and write outputs."""
        bundle = self.load_database(db_path)
        region_value = region or self.config.default_region
        commercial_terms = resolve_commercial_terms(self.config, bundle.controls)
        weights = MatchingWeights(
            threshold=threshold or float(self.config.get("matching.threshold", 78)),
            review_threshold=float(self.config.get("matching.review_threshold", 65)),
            strong_threshold=float(self.config.get("matching.strong_threshold", 88)),
            region_bonus=float(self.config.get("matching.region_bonus", 4)),
            section_bonus=float(self.config.get("matching.section_bonus", 8)),
            unit_bonus=float(self.config.get("matching.unit_bonus", 6)),
            alias_bonus=float(self.config.get("matching.alias_bonus", 5)),
            unit_penalty=float(self.config.get("matching.unit_penalty", 18)),
        )
        matcher = Matcher(bundle.rate_items, bundle.aliases, weights)
        reader = WorkbookReader(self.config, self.logger)
        writer = WorkbookWriter()

        sheets = reader.read(boq_path, column_overrides)
        results: list[MatchResult] = []
        for sheet in sheets:
            self.logger.info("Processing sheet '%s' classified as %s with %s rows", sheet.sheet_name, sheet.classification, len(sheet.rows))
            nearby_headings: list[str] = []
            for row in sheet.rows:
                if row.is_heading:
                    nearby_headings.append(row.description)
                    continue
                if row.is_subtotal or row.is_total or row.is_summary_row:
                    self.logger.debug("Skipping summary row %s on %s", row.row_number, row.sheet_name)
                    continue
                row.inferred_section = infer_section(sheet.sheet_name, row, nearby_headings, bundle.section_rules)
                match = matcher.match(row, region_value)
                direct_score = match.confidence_score
                if match.confidence_score < weights.threshold:
                    build_up = price_build_up_recipe(
                        row,
                        bundle.build_recipes,
                        bundle.build_inputs,
                        region_value,
                        threshold=weights.threshold,
                    )
                    if build_up:
                        build_up.rationale.append(f"direct-match-score={direct_score:.2f}")
                        if build_up.confidence_score >= match.confidence_score:
                            build_up.rationale.append("build-up selected over direct match")
                            if match.rate is not None:
                                build_up.rationale.append(f"direct-rate={match.rate:.2f}")
                            match = build_up
                        else:
                            match.rationale.append(f"build-up-score={build_up.confidence_score:.2f}")
                            match.rationale.append("direct match retained over build-up")
                    else:
                        match.rationale.append("no build-up candidate available")
                if match.confidence_score < weights.review_threshold:
                    match.decision = "unmatched"
                    match.review_flag = True
                    match.commercial_review_flags.append("low confidence requires manual pricing review")
                self._apply_commercial_adjustments(match, bundle.regional_adjustments, region_value, commercial_terms)
                self.logger.debug(
                    "Row %s on %s -> %s (%s) score=%s",
                    row.row_number,
                    row.sheet_name,
                    match.matched_item_code or "unmatched",
                    match.decision,
                    match.confidence_score,
                )
                results.append(match)

        quotation_summary = summarize_quote(results, commercial_terms)

        output_workbook = writer.write(
            boq_path,
            output_path,
            sheets,
            results,
            quotation_summary,
            bundle.assumptions,
            bundle.exclusions,
            commercial_terms,
            apply_rates if apply_rates is not None else bool(self.config.get("processing.apply_rates", False)),
            bool(self.config.get("processing.write_amount_formulas", True)),
        )

        unmatched_csv: Path | None = None
        if self.config.get("processing.export_unmatched_csv", True):
            unmatched_csv = Path(output_workbook.parent / f"{output_workbook.stem}_unmatched.csv")
            export_unmatched_csv(unmatched_csv, results)

        audit_json: Path | None = None
        if self.config.get("processing.export_audit_json", True):
            audit_json = Path(output_workbook.parent / f"{output_workbook.stem}_audit.json")
            write_audit_json(
                audit_json,
                results,
                metadata={
                    "db_path": db_path,
                    "boq_path": boq_path,
                    "region": region_value,
                    "commercial_terms": commercial_terms,
                    "quotation_summary": quotation_summary,
                },
            )

        matched = sum(1 for item in results if item.decision == "matched")
        flagged = sum(1 for item in results if item.review_flag)
        return RunArtifacts(
            output_workbook=output_workbook,
            unmatched_csv=unmatched_csv,
            audit_json=audit_json,
            processed=len(results),
            matched=matched,
            flagged=flagged,
            quotation_summary=quotation_summary,
        )

    def validate_database(self, db_path: str) -> list[str]:
        """Validate workbook sheets and minimum required columns."""
        workbook = load_workbook(db_path, data_only=True)
        errors: list[str] = []
        missing = REQUIRED_SHEETS - set(workbook.sheetnames)
        if missing:
            errors.append(f"Missing sheets: {', '.join(sorted(missing))}")
            return errors

        sheet_columns = {
            "RateLibrary": [
                "item_code", "description", "normalized_description", "section", "subsection",
                "unit", "rate", "currency", "region", "source", "source_sheet", "source_page",
                "basis", "crew_type", "plant_type", "material_type", "keywords", "alias_group",
                "build_up_recipe_id", "confidence_hint", "notes", "active",
            ],
            "Aliases": ["alias", "canonical_term", "section_bias", "notes"],
            "SectionMap": ["trigger_text", "inferred_section", "priority"],
            "BuildUpInputs": ["input_code", "input_type", "description", "unit", "rate", "region", "source", "active"],
            "BuildUpRecipes": ["recipe_id", "recipe_name", "output_description", "output_unit", "section", "component_code", "factor", "waste_factor", "notes"],
            "Controls": ["key", "value"],
            "ReviewLog": ["timestamp", "boq_file", "sheet_name", "row_number", "boq_description", "decision", "matched_item_code", "matched_description", "confidence_score", "reviewer_note"],
            "CandidateMatches": [
                "timestamp", "import_batch_id", "source_file", "source_sheet", "target_sheet", "item_code",
                "description", "normalized_description", "section", "subsection", "unit", "rate", "currency", "region",
                "source", "source_page", "basis", "crew_type", "plant_type", "material_type", "keywords", "alias_group",
                "build_up_recipe_id", "confidence_hint", "notes", "active", "duplicate_reason", "matched_item_code",
                "reviewer_status", "reviewer_name", "reviewed_at", "review_decision", "promote_target",
                "approved_item_code", "approved_description", "approved_rate", "approved_canonical_term",
                "approved_section_bias", "confidence_override", "reviewer_note", "promotion_status", "promoted_at",
            ],
            "RegionalAdjustments": ["region", "section", "factor", "notes", "active"],
            "Assumptions": ["assumption", "category", "active"],
            "Exclusions": ["exclusion", "category", "active"],
        }
        for sheet_name, required_columns in sheet_columns.items():
            if sheet_name not in workbook.sheetnames:
                continue
            headers = self._headers(workbook[sheet_name])
            missing_headers = [column for column in required_columns if column not in headers]
            if missing_headers:
                errors.append(f"{sheet_name} missing columns: {', '.join(missing_headers)}")
        return errors

    def _load_rate_library(self, sheet) -> list[RateItem]:
        headers = self._headers(sheet)
        items: list[RateItem] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            values = dict(zip(headers, row))
            if not values.get("description"):
                continue
            items.append(
                RateItem(
                    item_code=str(values.get("item_code") or ""),
                    description=str(values.get("description") or ""),
                    normalized_description=normalize_text(
                        str(values.get("normalized_description") or values.get("description") or "")
                    ),
                    section=str(values.get("section") or ""),
                    subsection=str(values.get("subsection") or ""),
                    unit=str(values.get("unit") or ""),
                    rate=safe_float(values.get("rate")) or 0.0,
                    currency=str(values.get("currency") or "KES"),
                    region=str(values.get("region") or ""),
                    source=str(values.get("source") or ""),
                    source_sheet=str(values.get("source_sheet") or ""),
                    source_page=str(values.get("source_page") or ""),
                    basis=str(values.get("basis") or ""),
                    crew_type=str(values.get("crew_type") or ""),
                    plant_type=str(values.get("plant_type") or ""),
                    material_type=str(values.get("material_type") or ""),
                    keywords=str(values.get("keywords") or ""),
                    alias_group=str(values.get("alias_group") or ""),
                    build_up_recipe_id=str(values.get("build_up_recipe_id") or ""),
                    confidence_hint=safe_float(values.get("confidence_hint")) or 0.0,
                    notes=str(values.get("notes") or ""),
                    active=True if values.get("active") in (None, "") else truthy(values.get("active")),
                )
            )
        return items

    def _load_aliases(self, sheet) -> list[AliasEntry]:
        headers = self._headers(sheet)
        items: list[AliasEntry] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            values = dict(zip(headers, row))
            if not values.get("alias"):
                continue
            items.append(
                AliasEntry(
                    alias=str(values.get("alias") or ""),
                    canonical_term=str(values.get("canonical_term") or ""),
                    section_bias=str(values.get("section_bias") or ""),
                    notes=str(values.get("notes") or ""),
                )
            )
        return items

    def _load_section_rules(self, sheet) -> list[SectionRule]:
        headers = self._headers(sheet)
        items: list[SectionRule] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            values = dict(zip(headers, row))
            if not values.get("trigger_text"):
                continue
            items.append(
                SectionRule(
                    trigger_text=str(values.get("trigger_text") or ""),
                    inferred_section=str(values.get("inferred_section") or ""),
                    priority=int(safe_float(values.get("priority")) or 0),
                )
            )
        return items

    def _load_build_inputs(self, sheet) -> list[BuildUpInput]:
        headers = self._headers(sheet)
        items: list[BuildUpInput] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            values = dict(zip(headers, row))
            if not values.get("input_code"):
                continue
            items.append(
                BuildUpInput(
                    input_code=str(values.get("input_code") or ""),
                    input_type=str(values.get("input_type") or ""),
                    description=str(values.get("description") or ""),
                    unit=str(values.get("unit") or ""),
                    rate=safe_float(values.get("rate")) or 0.0,
                    region=str(values.get("region") or ""),
                    source=str(values.get("source") or ""),
                    active=True if values.get("active") in (None, "") else truthy(values.get("active")),
                )
            )
        return items

    def _load_build_recipes(self, sheet) -> list[BuildUpRecipeLine]:
        headers = self._headers(sheet)
        items: list[BuildUpRecipeLine] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            values = dict(zip(headers, row))
            if not values.get("recipe_id"):
                continue
            items.append(
                BuildUpRecipeLine(
                    recipe_id=str(values.get("recipe_id") or ""),
                    recipe_name=str(values.get("recipe_name") or ""),
                    output_description=str(values.get("output_description") or ""),
                    output_unit=str(values.get("output_unit") or ""),
                    section=str(values.get("section") or ""),
                    component_code=str(values.get("component_code") or ""),
                    factor=safe_float(values.get("factor")) or 0.0,
                    waste_factor=safe_float(values.get("waste_factor")) or 0.0,
                    notes=str(values.get("notes") or ""),
                )
            )
        return items

    def _load_key_values(self, sheet) -> dict[str, str]:
        headers = self._headers(sheet)
        items: dict[str, str] = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            values = dict(zip(headers, row))
            key = str(values.get("key") or "").strip()
            if key:
                items[key] = str(values.get("value") or "")
        return items

    def _load_regional_adjustments(self, sheet) -> list[RegionalAdjustment]:
        headers = self._headers(sheet)
        items: list[RegionalAdjustment] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            values = dict(zip(headers, row))
            if not values.get("region"):
                continue
            items.append(
                RegionalAdjustment(
                    region=str(values.get("region") or ""),
                    section=str(values.get("section") or "*"),
                    factor=safe_float(values.get("factor")) or 1.0,
                    notes=str(values.get("notes") or ""),
                    active=True if values.get("active") in (None, "") else truthy(values.get("active")),
                )
            )
        return items

    def _load_text_entries(self, sheet, text_field: str) -> list[TextListEntry]:
        headers = self._headers(sheet)
        items: list[TextListEntry] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            values = dict(zip(headers, row))
            text = str(values.get(text_field) or "").strip()
            if not text:
                continue
            active = True if values.get("active") in (None, "") else truthy(values.get("active"))
            if active:
                items.append(
                    TextListEntry(
                        text=text,
                        category=str(values.get("category") or ""),
                        active=active,
                    )
                )
        return items

    def _load_generic_rows(self, sheet) -> list[dict]:
        headers = self._headers(sheet)
        return [dict(zip(headers, row)) for row in sheet.iter_rows(min_row=2, values_only=True)]

    def _apply_commercial_adjustments(
        self,
        match: MatchResult,
        adjustments: list[RegionalAdjustment],
        region: str,
        terms: CommercialTerms,
    ) -> None:
        """Apply regional factors and commercial review metadata to a match result."""
        section = match.section_used or match.boq_line.inferred_section or match.boq_line.sheet_name
        factor, notes = resolve_regional_factor(adjustments, region, section)
        if match.rate is not None:
            match.base_rate = match.rate if match.base_rate is None else match.base_rate
            match.rate = round(match.rate * factor, 2)
        match.regional_factor = factor
        match.approval_status = "Approved for Pricing" if match.decision == "matched" and not match.review_flag else terms.default_approval_status
        if factor != 1.0:
            match.rationale.append(f"regional-factor={factor:.3f}")
            if notes:
                match.rationale.append(notes)
        if not match.basis_of_rate:
            match.commercial_review_flags.append("Missing basis of rate")
        if match.alternate_options:
            match.commercial_review_flags.append("Alternate rate options available")
        if match.review_flag:
            match.commercial_review_flags.append("Confidence below strong threshold")
        if match.built_up:
            match.commercial_review_flags.append("Build-up fallback used")
        match.commercial_review_flags = list(dict.fromkeys(flag for flag in match.commercial_review_flags if flag))

    @staticmethod
    def _headers(sheet) -> list[str]:
        return [str(cell.value or "").strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
