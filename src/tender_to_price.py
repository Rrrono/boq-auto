"""Integrated tender-to-pricing orchestration."""

from __future__ import annotations

from dataclasses import asdict, dataclass, field
import logging
from pathlib import Path

from openpyxl import Workbook, load_workbook

from .engine import PricingEngine
from .models import QuotationSummary, RunArtifacts
from .normalizer import normalize_text
from .section_inference import classify_sheet_name
from .tender_models import PricingHandoffRow, TenderAnalysisResult
from .tender_workflow import TenderWorkflow
from .utils import dump_json, ensure_parent
from .workbook_reader import WorkbookReader
from .workbook_writer import format_worksheet


@dataclass(slots=True)
class TenderToPriceArtifacts:
    """Integrated artifacts for the tender-to-price command."""

    output_workbook: Path
    pricing_handoff_rows: list[PricingHandoffRow] = field(default_factory=list)
    pricing_artifacts: RunArtifacts | None = None
    tender_result: TenderAnalysisResult | None = None
    handoff_workbook: Path | None = None
    output_json: Path | None = None


class TenderToPriceRunner:
    """Bridge tender-analysis outputs into the existing pricing engine."""

    def __init__(
        self,
        config,
        logger: logging.Logger | None = None,
        tender_workflow: TenderWorkflow | None = None,
        pricing_engine: PricingEngine | None = None,
    ) -> None:
        self.config = config
        self.logger = logger or logging.getLogger("boq_auto")
        self.tender_workflow = tender_workflow or TenderWorkflow(config, self.logger)
        self.pricing_engine = pricing_engine or PricingEngine(config, self.logger)
        self.reader = WorkbookReader(config, self.logger)
        self.handoff_sheet_name = str(config.get("tender_to_price.worksheet_names.handoff", "Pricing Handoff"))
        self.include_gap_findings = bool(config.get("tender_to_price.include_gap_findings_in_handoff", True))

    def run(
        self,
        input_path: str,
        db_path: str,
        output_path: str,
        boq_path: str | None = None,
        region: str | None = None,
        threshold: float | None = None,
        apply_rates: bool = False,
        title_override: str | None = None,
        json_path: str | None = None,
        matching_mode: str = "rule",
    ) -> TenderToPriceArtifacts:
        """Run the integrated tender-to-price workflow."""

        tender_result = self.tender_workflow.prepare_result(
            input_path=input_path,
            title_override=title_override,
            boq_path=boq_path,
            include_gap_check=True,
        )
        handoff_rows = self.build_pricing_handoff_rows(tender_result, boq_path)
        tender_result.pricing_handoff = handoff_rows

        handoff_workbook: Path | None = None
        if boq_path:
            pricing_artifacts = self.pricing_engine.price_workbook(
                db_path=db_path,
                boq_path=boq_path,
                output_path=output_path,
                region=region,
                threshold=threshold,
                apply_rates=apply_rates,
                matching_mode=matching_mode,
            )
            self._write_pricing_handoff_sheet(pricing_artifacts.output_workbook, handoff_rows)
            final_workbook = pricing_artifacts.output_workbook
        else:
            handoff_workbook = Path(output_path).with_name(f"{Path(output_path).stem}_handoff.xlsx")
            pricing_ready_rows = [row for row in handoff_rows if row.source_origin != "gap_check"]
            self._create_handoff_workbook(handoff_workbook, pricing_ready_rows)
            pricing_artifacts = self.pricing_engine.price_workbook(
                db_path=db_path,
                boq_path=str(handoff_workbook),
                output_path=output_path,
                region=region,
                threshold=threshold,
                apply_rates=apply_rates,
                matching_mode=matching_mode,
            )
            final_workbook = pricing_artifacts.output_workbook

        self.tender_workflow.append_result_sheets(final_workbook, tender_result)

        artifacts = TenderToPriceArtifacts(
            output_workbook=final_workbook,
            pricing_handoff_rows=handoff_rows,
            pricing_artifacts=pricing_artifacts,
            tender_result=tender_result,
            handoff_workbook=handoff_workbook,
        )
        if json_path:
            artifacts.output_json = Path(json_path)
            self._write_json(artifacts.output_json, artifacts)

        self.logger.info(
            "Tender-to-price workflow complete: output=%s handoff_rows=%s priced_items=%s",
            final_workbook,
            len(handoff_rows),
            pricing_artifacts.processed if pricing_artifacts else 0,
        )
        return artifacts

    def build_pricing_handoff_rows(
        self,
        tender_result: TenderAnalysisResult,
        boq_path: str | None = None,
    ) -> list[PricingHandoffRow]:
        """Build pricing handoff rows from existing BOQ, tender drafts, and actionable gaps."""

        rows: list[PricingHandoffRow] = []
        seen: set[tuple[str, str, str]] = set()

        if boq_path:
            for sheet in self.reader.read(boq_path):
                section_name = classify_sheet_name(sheet.sheet_name) or sheet.sheet_name
                for row in sheet.rows:
                    if row.is_heading or row.is_summary_row or not row.description:
                        continue
                    signature = ("existing_boq", section_name, normalize_text(row.description))
                    if signature in seen:
                        continue
                    seen.add(signature)
                    rows.append(
                        PricingHandoffRow(
                            section=section_name,
                            description=row.description,
                            unit=row.unit,
                            quantity=row.quantity,
                            source_basis=f"Existing BOQ row from {sheet.sheet_name}",
                            source_origin="existing_boq",
                            inferred_from_tender=False,
                            confidence=95.0,
                            spec_attributes="",
                            review_required=row.quantity is None or not row.unit,
                            notes="Existing BOQ item carried into the integrated pricing handoff.",
                            source_reference=f"{sheet.sheet_name}!R{row.row_number}",
                        )
                    )

        for suggestion in tender_result.draft_suggestions:
            signature = ("tender_draft", suggestion.section, normalize_text(suggestion.description))
            if signature in seen:
                continue
            seen.add(signature)
            rows.append(
                PricingHandoffRow(
                    section=suggestion.section,
                    description=suggestion.description,
                    unit=suggestion.unit,
                    quantity=None,
                    source_basis=suggestion.source_basis,
                    source_origin="tender_draft",
                    inferred_from_tender=True,
                    confidence=suggestion.confidence,
                    spec_attributes=suggestion.spec_attributes,
                    review_required=True,
                    notes=suggestion.notes,
                    source_reference=suggestion.source_reference,
                )
            )

        if not self.include_gap_findings:
            return rows

        for gap in tender_result.gap_items:
            if gap.gap_type not in {"Missing Item", "Measurement Clarification"}:
                continue
            signature = ("gap_check", gap.section, normalize_text(gap.description))
            if signature in seen:
                continue
            seen.add(signature)
            rows.append(
                PricingHandoffRow(
                    section=gap.section,
                    description=gap.description,
                    unit="",
                    quantity=None,
                    source_basis=f"Gap-check finding: {gap.gap_type}",
                    source_origin="gap_check",
                    inferred_from_tender=True,
                    confidence=gap.confidence,
                    spec_attributes="",
                    review_required=True,
                    notes=gap.notes,
                    source_reference=gap.source_reference,
                )
            )
        return rows

    def _create_handoff_workbook(self, workbook_path: Path, handoff_rows: list[PricingHandoffRow]) -> Path:
        ensure_parent(workbook_path)
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = self.handoff_sheet_name
        worksheet.append(
            [
                "section",
                "description",
                "unit",
                "quantity",
                "rate",
                "amount",
                "source_basis",
                "source_origin",
                "inferred_from_tender",
                "confidence",
                "spec_attributes",
                "review_required",
                "notes",
                "source_reference",
            ]
        )
        for row in handoff_rows:
            worksheet.append(
                [
                    row.section,
                    row.description,
                    row.unit,
                    row.quantity,
                    None,
                    None,
                    row.source_basis,
                    row.source_origin,
                    "YES" if row.inferred_from_tender else "NO",
                    row.confidence,
                    row.spec_attributes,
                    "YES" if row.review_required else "NO",
                    row.notes,
                    row.source_reference,
                ]
            )
        format_worksheet(worksheet, "pricing_handoff")
        workbook.save(workbook_path)
        return workbook_path

    def _write_pricing_handoff_sheet(self, workbook_path: Path | str, handoff_rows: list[PricingHandoffRow]) -> Path:
        workbook = load_workbook(workbook_path)
        if self.handoff_sheet_name in workbook.sheetnames:
            del workbook[self.handoff_sheet_name]
        worksheet = workbook.create_sheet(self.handoff_sheet_name)
        worksheet.append(
            [
                "section",
                "description",
                "unit",
                "quantity",
                "source_basis",
                "source_origin",
                "inferred_from_tender",
                "confidence",
                "spec_attributes",
                "review_required",
                "notes",
                "source_reference",
            ]
        )
        for row in handoff_rows:
            worksheet.append(
                [
                    row.section,
                    row.description,
                    row.unit,
                    row.quantity,
                    row.source_basis,
                    row.source_origin,
                    "YES" if row.inferred_from_tender else "NO",
                    round(row.confidence, 1),
                    row.spec_attributes,
                    "YES" if row.review_required else "NO",
                    row.notes,
                    row.source_reference,
                ]
            )
        format_worksheet(worksheet, "pricing_handoff")
        workbook.save(workbook_path)
        return Path(workbook_path)

    def _write_json(self, output_path: Path, artifacts: TenderToPriceArtifacts) -> None:
        payload = {
            "output_workbook": str(artifacts.output_workbook),
            "handoff_workbook": str(artifacts.handoff_workbook) if artifacts.handoff_workbook else "",
            "pricing_handoff_rows": [asdict(row) for row in artifacts.pricing_handoff_rows],
            "pricing_artifacts": {
                "processed": artifacts.pricing_artifacts.processed if artifacts.pricing_artifacts else 0,
                "matched": artifacts.pricing_artifacts.matched if artifacts.pricing_artifacts else 0,
                "flagged": artifacts.pricing_artifacts.flagged if artifacts.pricing_artifacts else 0,
                "quotation_summary": asdict(artifacts.pricing_artifacts.quotation_summary)
                if artifacts.pricing_artifacts and artifacts.pricing_artifacts.quotation_summary
                else asdict(QuotationSummary()),
            },
            "tender_summary": asdict(artifacts.tender_result.summary)
            if artifacts.tender_result and artifacts.tender_result.summary
            else {},
        }
        dump_json(output_path, payload)
