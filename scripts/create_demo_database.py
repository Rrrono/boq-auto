"""Create a demo Excel pricing database for BOQ AUTO."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook


OUTPUT = Path("database/qs_database.xlsx")


def add_sheet(workbook: Workbook, title: str, headers: list[str], rows: list[list]) -> None:
    """Create a sheet and populate rows."""
    if title in workbook.sheetnames:
        del workbook[title]
    sheet = workbook.create_sheet(title)
    sheet.append(headers)
    for row in rows:
        sheet.append(row)


def main() -> None:
    """Generate a realistic demo database workbook."""
    workbook = Workbook()
    if "Sheet" in workbook.sheetnames:
        del workbook["Sheet"]

    rate_headers = [
        "item_code", "description", "normalized_description", "section", "subsection", "unit", "rate",
        "currency", "region", "source", "source_sheet", "source_page", "basis", "crew_type", "plant_type",
        "material_type", "keywords", "alias_group", "build_up_recipe_id", "confidence_hint", "notes", "active",
    ]
    rate_rows = [
        ["PL001", "15 tonne tipper lorry", "15 ton tipper lorry", "Dayworks", "Plant", "day", 18500, "KES", "Nyanza", "QS Library", "Plant", "12", "hire", "", "transport", "", "tipper lorry haulage", "tipper", "", 4, "regional rate", True],
        ["PL001A", "10 tonne tipper lorry", "10 ton tipper lorry", "Dayworks", "Plant", "day", 16200, "KES", "Nyanza", "QS Library", "Plant", "12", "hire", "", "transport", "", "tipper lorry haulage", "tipper", "", 2, "alternate size option", True],
        ["PL002", "2 cm/hr dewatering pump", "2 cm/hr dewatering pump", "Dayworks", "Plant", "day", 6200, "KES", "Nyanza", "QS Library", "Plant", "14", "hire", "", "pump", "", "dewatering pump", "pump", "", 4, "", True],
        ["PL003", "Concrete mixer 0.3 0.7 m3/min", "concrete mixer 0.3 0.7 m3/min", "Concrete", "Plant", "day", 8500, "KES", "Nairobi", "QS Library", "Plant", "16", "hire", "", "mixer", "", "concrete mixer", "mixer", "", 3, "", True],
        ["PL003A", "Concrete mixer 0.5 m3/min", "concrete mixer 0.5 m3/min", "Concrete", "Plant", "day", 9100, "KES", "Nyanza", "QS Library", "Plant", "16", "hire", "", "mixer", "", "concrete mixer", "mixer", "", 4, "regional alternative", True],
        ["PL004", "Sheep foot roller 15 tons", "sheep foot roller 15 ton", "Earthworks", "Plant", "day", 24000, "KES", "Nyanza", "QS Library", "Plant", "18", "hire", "", "roller", "", "compaction roller", "roller", "", 4, "", True],
        ["PL005", "Excavator with loader attachment 1.7 m3", "excavator loader attachment 1.7 m3", "Earthworks", "Plant", "day", 36000, "KES", "Nyanza", "QS Library", "Plant", "21", "hire", "", "excavator", "", "excavator loader", "excavator", "", 6, "", True],
        ["PL006", "Pick-up truck 1 1.5 tonne capacity", "pick up truck 1 1.5 ton capacity", "Dayworks", "Transport", "day", 6500, "KES", "Nyanza", "QS Library", "Transport", "5", "hire", "", "vehicle", "", "pickup truck", "pickup", "", 3, "", True],
        ["PL007", "Compressor with drill 250 cfm complete with tools hoses and steel bits", "compressor drill 250 cfm tools hoses steel bits", "Dayworks", "Plant", "day", 15500, "KES", "Nyanza", "QS Library", "Plant", "24", "hire", "", "compressor", "", "compressor drill", "compressor", "", 5, "", True],
        ["PL008", "Concrete vibrator poker type", "concrete vibrator poker", "Concrete", "Plant", "day", 2500, "KES", "Nyanza", "QS Library", "Plant", "25", "hire", "", "vibrator", "", "concrete vibrator", "vibrator", "", 4, "", True],
        ["PL009", "Self propelled water tanker 6000 20000 litres with pick up pump", "self propelled water tanker 6000 20000 litre pick up pump", "Earthworks", "Plant", "day", 28500, "KES", "Nyanza", "QS Library", "Plant", "28", "hire", "", "tanker", "", "water tanker", "tanker", "", 5, "", True],
        ["PL010", "Crawler dozer with dozer and hydraulic ripper attachments", "crawler dozer hydraulic ripper", "Earthworks", "Plant", "day", 42000, "KES", "Nyanza", "QS Library", "Plant", "29", "hire", "", "dozer", "", "crawler dozer", "dozer", "", 5, "", True],
        ["PL011", "Motor grader complete with hydraulic ripper or scarifier", "motor grader hydraulic ripper scarifier", "Earthworks", "Plant", "day", 34000, "KES", "Nyanza", "QS Library", "Plant", "31", "hire", "", "grader", "", "motor grader", "grader", "", 5, "", True],
        ["PR001", "Preliminaries and general items", "preliminaries general items", "Preliminaries", "General", "sum", 125000, "KES", "Nyanza", "QS Library", "Prelims", "1", "lump sum", "", "", "", "preliminaries", "prelims", "", 7, "", True],
        ["EW001", "Excavation in ordinary soil", "excavation ordinary soil", "Earthworks", "Excavation", "m3", 950, "KES", "Nyanza", "QS Library", "Earthworks", "8", "supply and execute", "", "", "", "excavation", "earth", "", 5, "", True],
        ["CO001", "Mass concrete class 15", "mass concrete class 15", "Concrete", "In-situ", "m3", 15200, "KES", "Nyanza", "QS Library", "Concrete", "11", "supply and place", "", "", "concrete", "mass concrete", "concrete", "RCP001", 6, "", True],
        ["FN001", "Plaster and render finish to walls", "plaster render finish walls", "Finishes", "Wall Finishes", "m2", 780, "KES", "Nyanza", "QS Library", "Finishes", "4", "supply and apply", "", "", "", "plaster finish", "finish", "", 4, "", True],
    ]
    add_sheet(workbook, "RateLibrary", rate_headers, rate_rows)

    add_sheet(
        workbook,
        "Aliases",
        ["alias", "canonical_term", "section_bias", "notes"],
        [
            ["tipper", "tipper lorry", "Dayworks", ""],
            ["pickup", "pick up", "Dayworks", ""],
            ["water bowser", "water tanker", "Earthworks", ""],
            ["vibrator", "concrete vibrator", "Concrete", ""],
            ["grader", "motor grader", "Earthworks", ""],
            ["dozer", "crawler dozer", "Earthworks", ""],
        ],
    )

    add_sheet(
        workbook,
        "SectionMap",
        ["trigger_text", "inferred_section", "priority"],
        [
            ["preliminaries", "Preliminaries", 100],
            ["dayworks", "Dayworks", 95],
            ["earthworks", "Earthworks", 90],
            ["concrete", "Concrete", 90],
            ["finishes", "Finishes", 90],
            ["excavation", "Earthworks", 80],
            ["plaster", "Finishes", 80],
        ],
    )

    add_sheet(
        workbook,
        "BuildUpInputs",
        ["input_code", "input_type", "description", "unit", "rate", "region", "source", "active"],
        [
            ["CEM001", "material", "Cement", "bag", 850, "Nyanza", "Demo Input", True],
            ["SAND01", "material", "Sand", "m3", 3200, "Nyanza", "Demo Input", True],
            ["BALL01", "material", "Ballast", "m3", 4200, "Nyanza", "Demo Input", True],
            ["LAB001", "labour", "Concrete gang", "day", 6500, "Nyanza", "Demo Input", True],
            ["WATR01", "material", "Water", "litre", 0.02, "Nyanza", "Demo Input", True],
        ],
    )

    add_sheet(
        workbook,
        "BuildUpRecipes",
        ["recipe_id", "recipe_name", "output_description", "output_unit", "section", "component_code", "factor", "waste_factor", "notes"],
        [
            ["RCP001", "Mass Concrete Class 15", "Mass concrete class 15", "m3", "Concrete", "CEM001", 6, 0.05, ""],
            ["RCP001", "Mass Concrete Class 15", "Mass concrete class 15", "m3", "Concrete", "SAND01", 0.5, 0.05, ""],
            ["RCP001", "Mass Concrete Class 15", "Mass concrete class 15", "m3", "Concrete", "BALL01", 0.8, 0.05, ""],
            ["RCP001", "Mass Concrete Class 15", "Mass concrete class 15", "m3", "Concrete", "LAB001", 0.25, 0.00, ""],
            ["RCP001", "Mass Concrete Class 15", "Mass concrete class 15", "m3", "Concrete", "WATR01", 180, 0.00, ""],
        ],
    )

    add_sheet(
        workbook,
        "Controls",
        ["key", "value"],
        [
            ["default_currency", "KES"],
            ["country", "Kenya"],
            ["default_region", "Nyanza"],
            ["overheads_pct", "8"],
            ["profit_pct", "5"],
            ["risk_pct", "3"],
            ["vat_pct", "16"],
            ["default_approval_status", "Pending Commercial Review"],
        ],
    )

    add_sheet(
        workbook,
        "Rules",
        ["rule_name", "rule_value"],
        [["unit_match_required", "false"], ["prefer_regional_rates", "true"]],
    )

    add_sheet(
        workbook,
        "ReviewLog",
        ["timestamp", "boq_file", "sheet_name", "row_number", "boq_description", "decision", "matched_item_code", "matched_description", "confidence_score", "reviewer_note"],
        [],
    )

    add_sheet(
        workbook,
        "CandidateMatches",
        [
            "timestamp", "import_batch_id", "source_file", "source_sheet", "target_sheet", "item_code",
            "description", "normalized_description", "section", "subsection", "unit", "rate", "currency", "region",
            "source", "source_page", "basis", "crew_type", "plant_type", "material_type", "keywords", "alias_group",
            "build_up_recipe_id", "confidence_hint", "notes", "active", "duplicate_reason", "matched_item_code",
            "reviewer_status", "reviewer_name", "reviewed_at", "review_decision", "promote_target",
            "approved_item_code", "approved_description", "approved_rate", "approved_canonical_term",
            "approved_section_bias", "confidence_override", "reviewer_note", "promotion_status", "promoted_at",
        ],
        [
            [
                "2026-03-17T09:00:00", "demo-import-001", "reviewed_rates.csv", "Sheet1", "RateLibrary", "PL001X",
                "Tipper truck hire", "tipper truck hire", "Dayworks", "Plant", "day", 17800, "KES", "Nyanza",
                "Estimator Review", "44", "Reviewed daywork comparison", "", "transport", "", "tipper lorry", "tipper",
                "", 72, "close duplicate requires review", True, "duplicate normalized description + unit", "PL001",
                "approved", "Senior QS", "2026-03-17T10:15:00", "accept", "aliases",
                "", "", "", "tipper lorry", "Dayworks", 88, "Prefer alias for informal wording", "not_promoted", "",
            ],
            [
                "2026-03-17T09:05:00", "demo-import-002", "plant_rates.csv", "Plant", "RateLibrary", "PL003X",
                "Concrete mixer 0.4 m3/min", "concrete mixer 0.4 m3/min", "Concrete", "Plant", "day", 8900, "KES", "Nyanza",
                "Estimator Review", "46", "Alternative mixer size", "", "mixer", "", "concrete mixer", "mixer",
                "", 69, "review size and rate before promotion", True, "duplicate normalized description + unit", "PL003A",
                "pending", "", "", "hold", "ratelibrary",
                "PL003B", "Concrete mixer 0.4 m3/min", 8900, "", "", "", "", "not_promoted", "",
            ],
        ],
    )

    add_sheet(
        workbook,
        "RegionalAdjustments",
        ["region", "section", "factor", "notes", "active"],
        [
            ["Nyanza", "*", 1.00, "Base demo region", True],
            ["Nairobi", "Concrete", 1.08, "Urban concrete premium", True],
            ["Nyanza", "Earthworks", 1.03, "Remote haulage adjustment", True],
        ],
    )

    add_sheet(
        workbook,
        "Assumptions",
        ["assumption", "category", "active"],
        [
            ["Rates are based on normal working hours and weekday operations.", "Commercial", True],
            ["Tender quantities are assumed to be measured and remeasured on site.", "Measurement", True],
            ["Quoted rates exclude abnormal weather disruption and security incidents.", "Risk", True],
        ],
    )

    add_sheet(
        workbook,
        "Exclusions",
        ["exclusion", "category", "active"],
        [
            ["Permit fees and statutory application charges.", "Statutory", True],
            ["Client-supplied design changes after award.", "Variation", True],
            ["Delays arising from restricted site access by third parties.", "Programme", True],
        ],
    )

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT)
    print(f"Demo database created at {OUTPUT}")


if __name__ == "__main__":
    main()
