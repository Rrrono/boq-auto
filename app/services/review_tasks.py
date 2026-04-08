"""Review task generation and workflow helpers."""

from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path
import os

from sqlalchemy.orm import Session

from app.models.job import ReviewTaskBridgeSummaryResponse, ReviewTaskBridgeSyncResponse, ReviewTaskResponse, ReviewTaskSyncResponse
from app.orm_models import Job, JobRun, ReviewTask
from src.cost_schema import CostDatabase
from src.ingestion import candidate_positions, generate_review_report, sync_review_artifacts_to_candidate_matches
from app.services.cost_engine import resolve_runtime_database_paths
from openpyxl import load_workbook


def _row_key(item: dict) -> str:
    sheet_name = str(item.get("sheet_name") or "").strip()
    row_number = int(item.get("row_number") or 0)
    description = str(item.get("description") or "").strip()
    return f"{sheet_name}:{row_number}:{description}".strip()


def _should_create_task(item: dict) -> bool:
    decision = str(item.get("decision") or "").strip().lower()
    review_flag = bool(item.get("review_flag") or False)
    return decision in {"review", "unmatched"} or review_flag


def _response_schema(*choices: str) -> list[str]:
    return [choice for choice in choices if choice]


def _focus_area(item: dict) -> tuple[str, bool]:
    description = str(item.get("description") or "").strip().lower()
    flag_reasons = [str(reason).strip().lower() for reason in item.get("flag_reasons") or [] if str(reason).strip()]
    if any(keyword in description for keyword in {"survey", "theodolite", "level", "chainage"}):
        return "survey", True
    if any(keyword in description for keyword in {"laboratory", "lab", "sieve", "mould", "cube test"}):
        return "lab_testing", True
    if any(keyword in description for keyword in {"chair", "desk", "cupboard", "table", "bed", "wardrobe"}):
        return "furniture_accommodation", True
    if any(keyword in description for keyword in {"cable", "conduit", "transformer", "lighting", "electrical"}):
        return "electrical_support", True
    if any(keyword in description for keyword in {"pipe", "valve", "manhole", "culvert", "sewer", "drain"}):
        return "pipes_fluids", True
    if any(keyword in description for keyword in {"grader", "roller", "tipper", "excavator", "paver", "compactor"}):
        return "plant_transport", True
    if "category_mismatch" in flag_reasons or "section_mismatch" in flag_reasons or "generic_match" in flag_reasons:
        return "general_gap", True
    return "", False


def _task_blueprint(item: dict) -> tuple[str, str, list[str], str, bool]:
    decision = str(item.get("decision") or "").strip().lower()
    description = str(item.get("description") or "").strip() or "this BOQ line"
    unit = str(item.get("unit") or "").strip()
    matched_description = str(item.get("matched_description") or "").strip()
    confidence_band = str(item.get("confidence_band") or "very_low").strip().lower()
    flag_reasons = [str(reason).strip().lower() for reason in item.get("flag_reasons") or [] if str(reason).strip()]
    category_mismatch = bool(item.get("category_mismatch_flag") or False)
    section_mismatch = bool(item.get("section_mismatch_flag") or False)
    focus_area, specialist_gap_flag = _focus_area(item)

    if specialist_gap_flag and decision == "unmatched":
        return (
            "specialist_rate_entry",
            f"The engine could not place '{description}' into a reliable {focus_area.replace('_', ' ') if focus_area else 'specialist'} bucket. Confirm the category direction first, then enter a practical rate for {unit or 'the required unit'} or keep it unmatched.",
            _response_schema("category_direction", "manual_rate", "no_good_match", "matched_description", "reviewer_note"),
            focus_area,
            True,
        )

    if specialist_gap_flag and (category_mismatch or section_mismatch or confidence_band in {"low", "very_low"}):
        return (
            "specialist_classification",
            f"'{description}' looks like a {focus_area.replace('_', ' ') if focus_area else 'specialist'} knowledge gap. Describe the right category direction and only keep the current match if it is genuinely valid.",
            _response_schema("confirm_match", "category_direction", "manual_rate", "no_good_match", "matched_description", "reviewer_note"),
            focus_area,
            True,
        )

    if decision == "unmatched":
        return (
            "manual_rate_entry",
            f"No acceptable library match was found for '{description}'. Enter a practical rate for {unit or 'the required unit'} or confirm that the item should remain unmatched.",
            _response_schema("manual_rate", "no_good_match", "reviewer_note"),
            focus_area,
            specialist_gap_flag,
        )

    if category_mismatch or any("category" in reason for reason in flag_reasons):
        return (
            "category_classification",
            f"The engine suspects a category mismatch for '{description}'. Decide whether the current match belongs to the right work family or describe the correct category direction.",
            _response_schema("confirm_match", "no_good_match", "matched_description", "reviewer_note"),
            focus_area,
            specialist_gap_flag,
        )

    if section_mismatch or any("section" in reason for reason in flag_reasons):
        return (
            "section_alignment",
            f"The row '{description}' may belong to a different section than the current suggestion. Confirm the match only if the section context is still valid.",
            _response_schema("confirm_match", "no_good_match", "matched_description", "reviewer_note"),
            focus_area,
            specialist_gap_flag,
        )

    if not matched_description or confidence_band in {"very_low", "low"}:
        return (
            "candidate_selection",
            f"Review the weak match for '{description}' and decide whether to keep the current suggestion or replace it with a better description/rate.",
            _response_schema("confirm_match", "manual_rate", "no_good_match", "matched_description", "reviewer_note"),
            focus_area,
            specialist_gap_flag,
        )

    return (
        "match_confirmation",
        f"Confirm whether '{matched_description}' is an acceptable match for '{description}' in {unit or 'the stated'} unit.",
        _response_schema("confirm_match", "manual_rate", "no_good_match", "reviewer_note"),
        focus_area,
        specialist_gap_flag,
    )


def serialize_review_task(task: ReviewTask) -> ReviewTaskResponse:
    flag_reasons = json.loads(task.flag_reasons_json or "[]")
    if not isinstance(flag_reasons, list):
        flag_reasons = []
    response_schema = json.loads(task.response_schema_json or "[]")
    if not isinstance(response_schema, list):
        response_schema = []
    return ReviewTaskResponse(
        id=task.id,
        job_id=task.job_id,
        job_run_id=task.job_run_id,
        status=task.status,
        source_row_key=task.source_row_key,
        sheet_name=task.sheet_name,
        row_number=task.row_number,
        description=task.description,
        matched_description=task.matched_description,
        matched_item_code=task.matched_item_code,
        task_type=task.task_type,
        focus_area=_focus_area(
            {
                "description": task.description,
                "flag_reasons": flag_reasons,
                "decision": task.decision,
            }
        )[0],
        specialist_gap_flag=_focus_area(
            {
                "description": task.description,
                "flag_reasons": flag_reasons,
                "decision": task.decision,
            }
        )[1],
        task_question=task.task_question,
        response_schema=[str(value) for value in response_schema],
        unit=task.unit,
        decision=task.decision,
        confidence_score=task.confidence_score,
        confidence_band=task.confidence_band,
        flag_reasons=[str(reason) for reason in flag_reasons],
        reviewer_uid=task.reviewer_uid,
        reviewer_email=task.reviewer_email,
        submitted_decision=task.submitted_decision,
        submitted_match_description=task.submitted_match_description,
        submitted_rate=task.submitted_rate,
        reviewer_note=task.reviewer_note,
        qa_status=task.qa_status,
        qa_reviewer_uid=task.qa_reviewer_uid,
        qa_reviewer_email=task.qa_reviewer_email,
        qa_note=task.qa_note,
        promotion_target=task.promotion_target,
        promotion_status=task.promotion_status,
        feedback_action=task.feedback_action,
        submitted_at=task.submitted_at,
        qa_updated_at=task.qa_updated_at,
        feedback_logged_at=task.feedback_logged_at,
        created_at=task.created_at,
        updated_at=task.updated_at,
    )


def sync_review_tasks_for_job(db: Session, job: Job) -> ReviewTaskSyncResponse:
    latest_run = next(iter(sorted(job.runs, key=lambda item: item.created_at, reverse=True)), None)
    if latest_run is None:
        return ReviewTaskSyncResponse(job_id=job.id, synced_count=0, open_count=0, tasks=[])

    payload = json.loads(latest_run.result_payload or "{}")
    items = payload.get("items", [])
    synced_count = 0
    now = datetime.now(timezone.utc)

    for item in items:
        if not isinstance(item, dict) or not _should_create_task(item):
            continue

        source_row_key = _row_key(item)
        if not source_row_key:
            continue

        task = (
            db.query(ReviewTask)
            .filter(ReviewTask.job_run_id == latest_run.id, ReviewTask.source_row_key == source_row_key)
            .first()
        )
        if task is None:
            task = ReviewTask(
                job_id=job.id,
                job_run_id=latest_run.id,
                source_row_key=source_row_key,
            )
            db.add(task)
            task.status = "open"
        task.sheet_name = str(item.get("sheet_name") or "")
        task.row_number = int(item.get("row_number") or 0)
        task.description = str(item.get("description") or "")
        task.matched_description = str(item.get("matched_description") or "")
        task.matched_item_code = str(item.get("matched_item_code") or "")
        task.task_type, task.task_question, response_schema, _focus_area_label, _specialist_gap_flag = _task_blueprint(item)
        task.response_schema_json = json.dumps(response_schema, ensure_ascii=True)
        task.unit = str(item.get("unit") or "")
        task.decision = str(item.get("decision") or "review")
        task.confidence_score = float(item.get("confidence_score") or 0.0)
        task.confidence_band = str(item.get("confidence_band") or "very_low")
        task.flag_reasons_json = json.dumps(item.get("flag_reasons") or [], ensure_ascii=True)
        task.updated_at = now
        synced_count += 1

    db.commit()
    tasks = (
        db.query(ReviewTask)
        .filter(ReviewTask.job_id == job.id, ReviewTask.job_run_id == latest_run.id)
        .order_by(ReviewTask.created_at.asc())
        .all()
    )
    open_count = sum(1 for task in tasks if task.status != "submitted")
    return ReviewTaskSyncResponse(
        job_id=job.id,
        synced_count=synced_count,
        open_count=open_count,
        tasks=[serialize_review_task(task) for task in tasks],
    )


def list_review_tasks(db: Session, *, status: str | None = None, reviewer_uid: str | None = None) -> list[ReviewTask]:
    query = db.query(ReviewTask).order_by(ReviewTask.updated_at.desc(), ReviewTask.created_at.desc())
    if status:
        query = query.filter(ReviewTask.status == status)
    if reviewer_uid:
        query = query.filter(ReviewTask.reviewer_uid == reviewer_uid)
    return list(query.all())


def get_review_task(db: Session, task_id: int) -> ReviewTask | None:
    return db.query(ReviewTask).filter(ReviewTask.id == task_id).first()


def claim_review_task(db: Session, task: ReviewTask, reviewer_uid: str, reviewer_email: str | None) -> ReviewTask:
    task.status = "claimed"
    task.reviewer_uid = reviewer_uid
    task.reviewer_email = reviewer_email or ""
    task.updated_at = datetime.now(timezone.utc)
    db.commit()
    db.refresh(task)
    return task


def submit_review_task(
    db: Session,
    task: ReviewTask,
    reviewer_uid: str,
    reviewer_email: str | None,
    *,
    decision: str,
    matched_description: str,
    rate: float | None,
    reviewer_note: str,
) -> ReviewTask:
    task.status = "submitted"
    task.reviewer_uid = reviewer_uid
    task.reviewer_email = reviewer_email or ""
    task.submitted_decision = decision.strip().lower()
    task.submitted_match_description = matched_description.strip()
    task.submitted_rate = rate
    task.reviewer_note = reviewer_note.strip()
    task.qa_status = "pending"
    task.qa_reviewer_uid = ""
    task.qa_reviewer_email = ""
    task.qa_note = ""
    task.promotion_target = ""
    task.promotion_status = "pending"
    task.feedback_action = ""
    now = datetime.now(timezone.utc)
    task.submitted_at = now
    task.qa_updated_at = None
    task.feedback_logged_at = None
    task.updated_at = now
    db.commit()
    db.refresh(task)
    return task


def _resolve_workbook_and_schema_paths() -> tuple[Path | None, Path | None]:
    workbook_path, schema_path = resolve_runtime_database_paths()
    return workbook_path, schema_path


def _count_synced_candidate_rows(workbook_path: Path) -> tuple[int, int]:
    if not workbook_path.exists():
        return 0, 0
    workbook = load_workbook(workbook_path)
    if "CandidateMatches" not in workbook.sheetnames:
        return 0, 0
    sheet = workbook["CandidateMatches"]
    positions = candidate_positions(sheet)
    synced_count = 0
    pending_count = 0
    for row_number in range(2, sheet.max_row + 1):
        source_file = str(sheet.cell(row_number, positions["source_file"]).value or "").strip() if "source_file" in positions else ""
        if source_file.startswith("schema-task:"):
            synced_count += 1
            reviewer_status = str(sheet.cell(row_number, positions["reviewer_status"]).value or "").strip().lower() if "reviewer_status" in positions else ""
            promotion_status = str(sheet.cell(row_number, positions["promotion_status"]).value or "").strip().lower() if "promotion_status" in positions else ""
            if reviewer_status == "pending" or promotion_status in {"not_promoted", "pending"}:
                pending_count += 1
    return synced_count, pending_count


def get_review_task_bridge_summary() -> ReviewTaskBridgeSummaryResponse:
    workbook_path, schema_path = _resolve_workbook_and_schema_paths()
    if workbook_path is None or schema_path is None or not schema_path.exists():
        return ReviewTaskBridgeSummaryResponse(available=False)

    repository = CostDatabase(schema_path)
    synced_candidate_rows, pending_workbook_candidates = _count_synced_candidate_rows(workbook_path)
    return ReviewTaskBridgeSummaryResponse(
        available=True,
        workbook_path=str(workbook_path),
        schema_path=str(schema_path),
        rate_observations=len(repository.fetch_rate_observations()),
        alias_suggestions=len(repository.fetch_alias_suggestions()),
        candidate_review_records=len(repository.fetch_candidate_reviews()),
        synced_candidate_rows=synced_candidate_rows,
        pending_workbook_candidates=pending_workbook_candidates,
    )


def sync_review_task_bridge(*, refresh_review_report: bool = True) -> ReviewTaskBridgeSyncResponse:
    workbook_path, schema_path = _resolve_workbook_and_schema_paths()
    if workbook_path is None or schema_path is None or not schema_path.exists():
        summary = get_review_task_bridge_summary()
        return ReviewTaskBridgeSyncResponse(available=False, bridge=summary)

    sync_summary = sync_review_artifacts_to_candidate_matches(str(workbook_path), str(schema_path))
    review_report_rows = 0
    if refresh_review_report:
        report_summary = generate_review_report(str(workbook_path))
        review_report_rows = report_summary.report_rows
    bridge = get_review_task_bridge_summary()
    return ReviewTaskBridgeSyncResponse(
        available=True,
        workbook_path=str(workbook_path),
        schema_path=str(schema_path),
        synced_count=sync_summary.appended,
        skipped_duplicates=sync_summary.skipped_duplicates,
        review_report_rows=review_report_rows,
        bridge=bridge,
    )


def _promotion_plan(task: ReviewTask, qa_status: str) -> tuple[str, str, str]:
    submitted_decision = (task.submitted_decision or "").strip().lower()
    if qa_status == "approved":
        if submitted_decision == "confirm_match" and task.matched_item_code:
            return "match_feedback", "logged", "accepted"
        if submitted_decision == "confirm_match" and task.submitted_match_description and task.submitted_match_description != task.matched_description:
            return "alias_suggestion", "ready", ""
        if submitted_decision == "manual_rate":
            return "rate_observation", "ready", "rejected"
        if submitted_decision == "no_good_match":
            return "candidate_review", "ready", "rejected"
    if qa_status == "escalated":
        return "candidate_review", "needs_attention", ""
    if qa_status == "rejected":
        return "", "closed", ""
    return "", "pending", ""


def _log_feedback_if_possible(task: ReviewTask, feedback_action: str) -> datetime | None:
    if not feedback_action or not task.matched_item_code:
        return None
    _, schema_path = _resolve_workbook_and_schema_paths()
    if schema_path is None or not Path(schema_path).exists():
        return None
    repository = CostDatabase(schema_path)
    resolved_item_id = repository.resolve_item_id(task.matched_item_code)
    if not resolved_item_id:
        return None
    repository.log_match_feedback(
        query_text=task.description,
        item_id=resolved_item_id,
        action=feedback_action,
        alternative_item_id="",
    )
    return datetime.now(timezone.utc)


def _persist_promotion_artifact(task: ReviewTask, reviewer_email: str | None) -> str:
    if task.promotion_status not in {"ready", "needs_attention"} or not task.promotion_target:
        return task.promotion_status
    _, schema_path = _resolve_workbook_and_schema_paths()
    if schema_path is None:
        return task.promotion_status

    repository = CostDatabase(schema_path)
    reviewer = (reviewer_email or task.qa_reviewer_email or task.reviewer_email or "").strip()
    metadata = {
        "job_id": task.job_id,
        "job_run_id": task.job_run_id,
        "source_row_key": task.source_row_key,
        "task_id": task.id,
        "decision": task.submitted_decision,
        "qa_status": task.qa_status,
        "matched_item_code": task.matched_item_code,
    }

    if task.promotion_target == "rate_observation" and task.submitted_rate is not None:
        repository.record_rate_observation(
            task.description,
            task.submitted_match_description or task.matched_description or task.description,
            task.unit,
            task.submitted_rate,
            source="review_task",
            reviewer=reviewer,
            status="approved",
            metadata=metadata,
        )
        return "logged"

    if task.promotion_target == "candidate_review":
        repository.record_candidate_review(
            task.description,
            task.submitted_match_description or task.matched_description,
            task.unit,
            reason=task.qa_note or task.reviewer_note or "review_required",
            reviewer=reviewer,
            status="pending",
            metadata=metadata,
        )
        return "logged"

    if task.promotion_target == "alias_suggestion":
        repository.record_alias_suggestion(
            task.description,
            task.submitted_match_description,
            section_bias="",
            reviewer=reviewer,
            status="pending",
            metadata=metadata,
        )
        return "logged"

    return task.promotion_status


def qa_review_task(
    db: Session,
    task: ReviewTask,
    reviewer_uid: str,
    reviewer_email: str | None,
    *,
    qa_status: str,
    qa_note: str,
) -> ReviewTask:
    normalized_status = qa_status.strip().lower()
    task.qa_status = normalized_status
    task.qa_reviewer_uid = reviewer_uid
    task.qa_reviewer_email = reviewer_email or ""
    task.qa_note = qa_note.strip()
    promotion_target, promotion_status, feedback_action = _promotion_plan(task, normalized_status)
    task.promotion_target = promotion_target
    task.promotion_status = promotion_status
    task.feedback_action = feedback_action
    now = datetime.now(timezone.utc)
    task.qa_updated_at = now
    task.feedback_logged_at = _log_feedback_if_possible(task, feedback_action)
    task.promotion_status = _persist_promotion_artifact(task, reviewer_email or task.qa_reviewer_email)
    task.updated_at = now
    db.commit()
    db.refresh(task)
    return task
