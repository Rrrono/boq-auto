from __future__ import annotations

from io import BytesIO
import os
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from app.db import Base, engine, init_db
from app.db import SessionLocal
from app.main import app
from app.orm_models import JobRun, ReviewTask
from src.cost_schema import CostDatabase, schema_database_path


client = TestClient(app)


def _build_workbook_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "BOQ"
    sheet.append(["Description", "Unit", "Quantity"])
    sheet.append(["Excavate foundation trench", "m3", 10])
    sheet.append(["Concrete to strip footing", "m3", 5])
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _build_unmatched_workbook_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "BOQ"
    sheet.append(["Description", "Unit", "Quantity"])
    sheet.append(["Engineer office modular container with kitchenette", "item", 1])
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _build_specialist_gap_workbook_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "BOQ"
    sheet.append(["Description", "Unit", "Quantity"])
    sheet.append(["Survey Equipment", "Lump Sum", 1])
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _build_runtime_database(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "RateLibrary"
    sheet.append(["item_code", "description", "normalized_description", "section", "subsection", "unit", "rate"])
    sheet.append(["A1", "Excavate foundation trench", "excavate foundation trench", "earthworks", "", "m3", 1200])
    alias_sheet = workbook.create_sheet("Aliases")
    alias_sheet.append(["alias", "canonical_term", "section_bias", "notes"])
    workbook.save(path)


def setup_function() -> None:
    Base.metadata.drop_all(bind=engine)
    init_db()
    os.environ.pop("BOQ_AUTO_API_DB_PATH", None)


def test_create_job_and_list_jobs() -> None:
    response = client.post("/jobs", json={"title": "KAA Demo Job", "region": "Nairobi"})
    assert response.status_code == 200
    body = response.json()
    assert body["title"] == "KAA Demo Job"
    assert body["region"] == "Nairobi"

    list_response = client.get("/jobs")
    assert list_response.status_code == 200
    jobs = list_response.json()
    assert len(jobs) == 1
    assert jobs[0]["id"] == body["id"]


def test_upload_boq_and_price_job() -> None:
    create_response = client.post("/jobs", json={"title": "Pricing Job", "region": "Nairobi"})
    job_id = create_response.json()["id"]

    upload_response = client.post(
        f"/jobs/{job_id}/files",
        data={"file_type": "boq"},
        files={"file": ("sample.xlsx", _build_workbook_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert upload_response.status_code == 200
    uploaded_job = upload_response.json()
    assert uploaded_job["files"][0]["file_type"] == "boq"

    price_response = client.post(f"/jobs/{job_id}/price-boq")
    assert price_response.status_code == 200
    priced = price_response.json()
    assert priced["job"]["status"] == "priced"
    assert priced["pricing"]["summary"]["item_count"] == 2

    results_response = client.get(f"/jobs/{job_id}/results")
    assert results_response.status_code == 200
    results = results_response.json()
    assert results["summary"]["item_count"] == 2
    assert len(results["items"]) == 2


def test_price_check_and_knowledge_queue() -> None:
    create_response = client.post("/jobs", json={"title": "Insight Job", "region": "Nairobi"})
    job_id = create_response.json()["id"]

    upload_response = client.post(
        f"/jobs/{job_id}/files",
        data={"file_type": "boq"},
        files={"file": ("sample.xlsx", _build_workbook_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert upload_response.status_code == 200

    price_response = client.post(f"/jobs/{job_id}/price-boq")
    assert price_response.status_code == 200

    price_check_response = client.get("/price-check", params={"q": "concrete"})
    assert price_check_response.status_code == 200
    price_check = price_check_response.json()
    assert price_check["scanned_jobs"] == 1
    assert price_check["filtered_rows"] >= 1
    assert any("concrete" in item["description"].lower() or "concrete" in item["matched_description"].lower() for item in price_check["observations"])
    first_observation = price_check["observations"][0]
    assert "confidence_band" in first_observation
    assert "flag_reasons" in first_observation
    assert "generic_match_flag" in first_observation
    assert "category_mismatch_flag" in first_observation
    assert "section_mismatch_flag" in first_observation
    assert "display_matched_description" in first_observation

    knowledge_response = client.get("/knowledge/candidates")
    assert knowledge_response.status_code == 200
    queue = knowledge_response.json()
    assert queue["scanned_jobs"] == 1
    assert queue["candidate_count"] >= 0
    assert "focus_areas" in queue
    assert "candidates" in queue
    if queue["candidates"]:
        first_candidate = queue["candidates"][0]
        assert "confidence_band" in first_candidate
        assert "flag_reasons" in first_candidate
        assert "generic_match_flag" in first_candidate
        assert "category_mismatch_flag" in first_candidate
        assert "section_mismatch_flag" in first_candidate
        assert "display_matched_description" in first_candidate


def test_sync_claim_and_submit_review_tasks() -> None:
    create_response = client.post("/jobs", json={"title": "Reviewer Job", "region": "Nairobi"})
    job_id = create_response.json()["id"]

    upload_response = client.post(
        f"/jobs/{job_id}/files",
        data={"file_type": "boq"},
        files={"file": ("sample.xlsx", _build_workbook_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert upload_response.status_code == 200

    price_response = client.post(f"/jobs/{job_id}/price-boq")
    assert price_response.status_code == 200

    sync_response = client.post(f"/jobs/{job_id}/review-tasks/sync")
    assert sync_response.status_code == 200
    sync_body = sync_response.json()
    assert sync_body["job_id"] == job_id
    assert "tasks" in sync_body
    assert sync_body["tasks"][0]["task_type"] in {"candidate_selection", "match_confirmation", "manual_rate_entry", "category_classification", "section_alignment", "specialist_classification", "specialist_rate_entry"}
    assert sync_body["tasks"][0]["task_question"]
    assert isinstance(sync_body["tasks"][0]["response_schema"], list)
    assert "focus_area" in sync_body["tasks"][0]
    assert "specialist_gap_flag" in sync_body["tasks"][0]

    tasks_response = client.get("/review-tasks")
    assert tasks_response.status_code == 200
    tasks = tasks_response.json()
    assert len(tasks) >= 1

    first_task = tasks[0]
    claim_response = client.post(f"/review-tasks/{first_task['id']}/claim")
    assert claim_response.status_code == 200
    claimed = claim_response.json()
    assert claimed["status"] == "claimed"

    submit_response = client.post(
        f"/review-tasks/{first_task['id']}/submit",
        json={
            "decision": "manual_rate",
            "category_direction": "",
            "matched_description": "Manual review item",
            "rate": 1250.0,
            "reviewer_note": "Reviewed for marketplace workflow smoke test.",
        },
    )
    assert submit_response.status_code == 200
    submitted = submit_response.json()
    assert submitted["status"] == "submitted"
    assert submitted["submitted_decision"] == "manual_rate"
    assert submitted["submitted_rate"] == 1250.0


def test_review_task_cannot_be_claimed_or_submitted_twice() -> None:
    create_response = client.post("/jobs", json={"title": "Reviewer Guardrail Job", "region": "Nairobi"})
    job_id = create_response.json()["id"]

    client.post(
        f"/jobs/{job_id}/files",
        data={"file_type": "boq"},
        files={"file": ("sample.xlsx", _build_unmatched_workbook_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    client.post(f"/jobs/{job_id}/price-boq")
    sync_response = client.post(f"/jobs/{job_id}/review-tasks/sync")
    task_id = sync_response.json()["tasks"][0]["id"]

    first_claim = client.post(f"/review-tasks/{task_id}/claim")
    assert first_claim.status_code == 200

    first_submit = client.post(
        f"/review-tasks/{task_id}/submit",
        json={
            "decision": "confirm_match",
            "category_direction": "",
            "matched_description": "Confirmed from review queue",
            "rate": None,
            "reviewer_note": "First submission",
        },
    )
    assert first_submit.status_code == 200

    second_claim = client.post(f"/review-tasks/{task_id}/claim")
    assert second_claim.status_code == 409

    second_submit = client.post(
        f"/review-tasks/{task_id}/submit",
        json={
            "decision": "no_good_match",
            "category_direction": "",
            "matched_description": "",
            "rate": None,
            "reviewer_note": "Second submission should fail",
        },
    )
    assert second_submit.status_code == 409


def test_bulk_claim_review_tasks_claims_open_tasks_and_skips_submitted() -> None:
    create_response = client.post("/jobs", json={"title": "Bulk Claim Job", "region": "Nairobi"})
    job_id = create_response.json()["id"]

    client.post(
        f"/jobs/{job_id}/files",
        data={"file_type": "boq"},
        files={"file": ("survey.xlsx", _build_specialist_gap_workbook_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    client.post(f"/jobs/{job_id}/price-boq")
    sync_response = client.post(f"/jobs/{job_id}/review-tasks/sync")
    task_id = sync_response.json()["tasks"][0]["id"]

    submitted_job = client.post("/jobs", json={"title": "Submitted Bulk Claim Guardrail", "region": "Nairobi"})
    submitted_job_id = submitted_job.json()["id"]
    client.post(
        f"/jobs/{submitted_job_id}/files",
        data={"file_type": "boq"},
        files={"file": ("sample.xlsx", _build_unmatched_workbook_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    client.post(f"/jobs/{submitted_job_id}/price-boq")
    submitted_sync = client.post(f"/jobs/{submitted_job_id}/review-tasks/sync")
    submitted_task_id = submitted_sync.json()["tasks"][0]["id"]
    client.post(f"/review-tasks/{submitted_task_id}/claim")
    client.post(
        f"/review-tasks/{submitted_task_id}/submit",
        json={
            "decision": "no_good_match",
            "category_direction": "survey",
            "matched_description": "",
            "rate": None,
            "reviewer_note": "Already submitted",
        },
    )

    bulk_response = client.post(
        "/review-tasks/bulk/claim",
        json={"task_ids": [task_id, submitted_task_id, 999999]},
    )
    assert bulk_response.status_code == 200
    body = bulk_response.json()
    assert body["requested_count"] == 3
    assert body["claimed_count"] == 1
    assert body["skipped_count"] == 2
    assert len(body["tasks"]) == 1
    assert body["tasks"][0]["id"] == task_id
    assert body["tasks"][0]["status"] == "claimed"


def test_bulk_qa_review_tasks_updates_submitted_tasks() -> None:
    create_response = client.post("/jobs", json={"title": "Bulk QA Job", "region": "Nairobi"})
    job_id = create_response.json()["id"]

    client.post(
        f"/jobs/{job_id}/files",
        data={"file_type": "boq"},
        files={"file": ("sample.xlsx", _build_unmatched_workbook_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    client.post(f"/jobs/{job_id}/price-boq")
    sync_response = client.post(f"/jobs/{job_id}/review-tasks/sync")
    task_id = sync_response.json()["tasks"][0]["id"]

    client.post(f"/review-tasks/{task_id}/claim")
    client.post(
        f"/review-tasks/{task_id}/submit",
        json={
            "decision": "manual_rate",
            "category_direction": "survey",
            "matched_description": "Bulk QA manual rate",
            "rate": 4100.0,
            "reviewer_note": "Ready for bulk QA",
        },
    )

    bulk_response = client.post(
        "/review-tasks/bulk/qa",
        json={
            "task_ids": [task_id, 999999],
            "qa_status": "approved",
            "qa_note": "Batch approved for reviewer milestone.",
        },
    )
    assert bulk_response.status_code == 200
    body = bulk_response.json()
    assert body["requested_count"] == 2
    assert body["updated_count"] == 1
    assert body["skipped_count"] == 1
    assert len(body["tasks"]) == 1
    assert body["tasks"][0]["id"] == task_id
    assert body["tasks"][0]["qa_status"] == "approved"


def test_review_task_can_move_into_qa_states() -> None:
    create_response = client.post("/jobs", json={"title": "Reviewer QA Job", "region": "Nairobi"})
    job_id = create_response.json()["id"]

    client.post(
        f"/jobs/{job_id}/files",
        data={"file_type": "boq"},
        files={"file": ("sample.xlsx", _build_unmatched_workbook_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    client.post(f"/jobs/{job_id}/price-boq")
    sync_response = client.post(f"/jobs/{job_id}/review-tasks/sync")
    task_id = sync_response.json()["tasks"][0]["id"]

    client.post(f"/review-tasks/{task_id}/claim")
    client.post(
        f"/review-tasks/{task_id}/submit",
        json={
            "decision": "manual_rate",
            "category_direction": "",
            "matched_description": "Manual line",
            "rate": 2250.0,
            "reviewer_note": "Ready for QA and promotion planning.",
        },
    )

    qa_response = client.post(
        f"/review-tasks/{task_id}/qa",
        json={
            "qa_status": "approved",
            "qa_note": "Good reviewer submission.",
        },
    )
    assert qa_response.status_code == 200
    qa_body = qa_response.json()
    assert qa_body["qa_status"] == "approved"
    assert qa_body["qa_note"] == "Good reviewer submission."
    assert qa_body["promotion_target"] == "rate_observation"
    assert qa_body["promotion_status"] == "logged"
    assert qa_body["feedback_action"] == "rejected"

    invalid_qa_response = client.post(
        f"/review-tasks/{task_id}/qa",
        json={
            "qa_status": "invalid_state",
            "qa_note": "",
        },
    )
    assert invalid_qa_response.status_code == 400


def test_approved_manual_rate_creates_rate_observation(tmp_path) -> None:
    schema_source = tmp_path / "runtime_master.xlsx"
    _build_runtime_database(schema_source)
    os.environ["BOQ_AUTO_API_DB_PATH"] = str(schema_source)
    repository = CostDatabase(schema_source)
    repository.initialize()

    create_response = client.post("/jobs", json={"title": "Reviewer Promotion Job", "region": "Nairobi"})
    job_id = create_response.json()["id"]

    with SessionLocal() as db:
        job_run = JobRun(
            job_id=job_id,
            run_type="price_boq",
            status="completed",
            processed=1,
            matched=0,
            flagged=1,
            total_cost=0.0,
            currency="KES",
            output_storage_uri="",
            audit_storage_uri="",
            result_payload="{}",
        )
        db.add(job_run)
        db.flush()
        task = ReviewTask(
            job_id=job_id,
            job_run_id=job_run.id,
            status="claimed",
            source_row_key="BOQ:2:Engineer office modular container with kitchenette",
            sheet_name="BOQ",
            row_number=2,
            description="Engineer office modular container with kitchenette",
            matched_description="",
            matched_item_code="",
            task_type="manual_rate_entry",
            task_question="Enter a practical rate or confirm the item should remain unmatched.",
            response_schema_json='["manual_rate","no_good_match","reviewer_note"]',
            unit="item",
            decision="unmatched",
            confidence_score=0.0,
            confidence_band="very_low",
            flag_reasons_json='["confidence_very_low"]',
            reviewer_uid="local-dev",
            reviewer_email="",
        )
        db.add(task)
        db.commit()
        task_id = task.id

    client.post(
        f"/review-tasks/{task_id}/submit",
        json={
          "decision": "manual_rate",
          "category_direction": "",
          "matched_description": "Manual reviewed trench rate",
          "rate": 3100.0,
          "reviewer_note": "Use reviewed site rate.",
        },
    )
    qa_response = client.post(
        f"/review-tasks/{task_id}/qa",
        json={
            "qa_status": "approved",
            "qa_note": "Promote reviewed manual rate.",
        },
    )

    assert qa_response.status_code == 200
    qa_body = qa_response.json()
    assert qa_body["promotion_status"] == "logged"

    observations = repository.fetch_rate_observations()
    assert observations
    assert observations[0].rate == 3100.0
    assert observations[0].canonical_description == "Manual reviewed trench rate"


def test_unmatched_rows_create_manual_rate_tasks() -> None:
    create_response = client.post("/jobs", json={"title": "Reviewer Task Type Job", "region": "Nairobi"})
    job_id = create_response.json()["id"]

    upload_response = client.post(
        f"/jobs/{job_id}/files",
        data={"file_type": "boq"},
        files={"file": ("sample.xlsx", _build_workbook_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert upload_response.status_code == 200

    price_response = client.post(f"/jobs/{job_id}/price-boq")
    assert price_response.status_code == 200

    sync_response = client.post(f"/jobs/{job_id}/review-tasks/sync")
    assert sync_response.status_code == 200
    tasks = sync_response.json()["tasks"]
    assert tasks
    assert any(task["task_question"] for task in tasks)
    assert any(task["response_schema"] for task in tasks)


def test_specialist_gap_rows_create_specialist_task_types() -> None:
    create_response = client.post("/jobs", json={"title": "Specialist Gap Job", "region": "Nairobi"})
    job_id = create_response.json()["id"]

    upload_response = client.post(
        f"/jobs/{job_id}/files",
        data={"file_type": "boq"},
        files={"file": ("survey.xlsx", _build_specialist_gap_workbook_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert upload_response.status_code == 200

    price_response = client.post(f"/jobs/{job_id}/price-boq")
    assert price_response.status_code == 200

    sync_response = client.post(f"/jobs/{job_id}/review-tasks/sync")
    assert sync_response.status_code == 200
    tasks = sync_response.json()["tasks"]
    assert tasks
    specialist_tasks = [task for task in tasks if task["specialist_gap_flag"]]
    assert specialist_tasks
    assert specialist_tasks[0]["task_type"] in {"specialist_classification", "specialist_rate_entry"}
    assert specialist_tasks[0]["focus_area"] in {"survey", "general_gap"}
    assert "category_direction" in specialist_tasks[0]["response_schema"]

    filtered_response = client.get("/review-tasks", params={"focus_area": "survey", "specialist_only": "true"})
    assert filtered_response.status_code == 200
    filtered_tasks = filtered_response.json()
    assert filtered_tasks
    assert all(task["specialist_gap_flag"] for task in filtered_tasks)
    assert all(task["focus_area"] == "survey" or task["submitted_category_direction"] == "survey" for task in filtered_tasks)


def test_category_direction_submission_is_persisted_and_promoted(tmp_path) -> None:
    schema_source = tmp_path / "runtime_master.xlsx"
    _build_runtime_database(schema_source)
    os.environ["BOQ_AUTO_API_DB_PATH"] = str(schema_source)
    repository = CostDatabase(schema_source)
    repository.initialize()

    create_response = client.post("/jobs", json={"title": "Category Direction Job", "region": "Nairobi"})
    job_id = create_response.json()["id"]

    with SessionLocal() as db:
        job_run = JobRun(
            job_id=job_id,
            run_type="price_boq",
            status="completed",
            processed=1,
            matched=0,
            flagged=1,
            total_cost=0.0,
            currency="KES",
            output_storage_uri="",
            audit_storage_uri="",
            result_payload="{}",
        )
        db.add(job_run)
        db.flush()
        task = ReviewTask(
            job_id=job_id,
            job_run_id=job_run.id,
            status="claimed",
            source_row_key="BOQ:2:Survey Equipment",
            sheet_name="BOQ",
            row_number=2,
            description="Survey Equipment",
            matched_description="",
            matched_item_code="",
            task_type="specialist_classification",
            task_question="Describe the right category direction for this specialist row.",
            response_schema_json='["category_direction","manual_rate","reviewer_note"]',
            unit="Lump Sum",
            decision="unmatched",
            confidence_score=0.0,
            confidence_band="very_low",
            flag_reasons_json='["generic_match","confidence_very_low"]',
            reviewer_uid="local-dev",
            reviewer_email="",
        )
        db.add(task)
        db.commit()
        task_id = task.id

    submit_response = client.post(
        f"/review-tasks/{task_id}/submit",
        json={
            "decision": "category_direction",
            "category_direction": "survey",
            "matched_description": "",
            "rate": None,
            "reviewer_note": "This belongs in survey equipment/services, not general preliminaries.",
        },
    )
    assert submit_response.status_code == 200
    submitted = submit_response.json()
    assert submitted["submitted_category_direction"] == "survey"

    qa_response = client.post(
        f"/review-tasks/{task_id}/qa",
        json={
            "qa_status": "approved",
            "qa_note": "Good category correction.",
        },
    )
    assert qa_response.status_code == 200
    qa_body = qa_response.json()
    assert qa_body["promotion_target"] == "candidate_review"
    assert qa_body["promotion_status"] == "logged"

    reviews = repository.fetch_candidate_reviews()
    assert reviews
    assert reviews[0].reason == "survey"

    summary_response = client.get("/review-tasks/bridge")
    assert summary_response.status_code == 200
    summary_body = summary_response.json()
    labels = {entry["label"] for entry in summary_body["taxonomy_backlog"]}
    assert "survey" in labels


def test_review_task_bridge_summary_and_sync_endpoint(tmp_path) -> None:
    schema_source = tmp_path / "runtime_master.xlsx"
    _build_runtime_database(schema_source)
    os.environ["BOQ_AUTO_API_DB_PATH"] = str(schema_source)
    repository = CostDatabase(schema_source)
    repository.record_rate_observation(
        "Crawler excavator hire",
        "Crawler excavator hire",
        "day",
        26500.0,
        source="review_task",
        reviewer="Senior QS",
        metadata={"task_id": "task-rate-bridge", "section": "Dayworks", "region": "Nyanza"},
    )
    repository.record_alias_suggestion(
        "Vibrating roller",
        "Vibratory roller",
        section_bias="Roadworks",
        reviewer="Senior QS",
        status="approved",
        metadata={"task_id": "task-alias-bridge"},
    )

    summary_response = client.get("/review-tasks/bridge")
    assert summary_response.status_code == 200
    summary_body = summary_response.json()
    assert summary_body["available"] is True
    assert summary_body["rate_observations"] >= 1
    assert summary_body["alias_suggestions"] >= 1
    assert "taxonomy_backlog" in summary_body

    sync_response = client.post("/review-tasks/bridge/sync")
    assert sync_response.status_code == 200
    sync_body = sync_response.json()
    assert sync_body["available"] is True
    assert sync_body["synced_count"] >= 2
    assert sync_body["bridge"]["synced_candidate_rows"] >= 2
    assert sync_body["review_report_rows"] >= 2

    workbook = load_workbook(schema_source)
    candidate_sheet = workbook["CandidateMatches"]
    source_markers = [str(row[2] or "") for row in candidate_sheet.iter_rows(min_row=2, values_only=True)]
    assert "schema-task:task-rate-bridge" in source_markers
    assert "schema-task:task-alias-bridge" in source_markers
