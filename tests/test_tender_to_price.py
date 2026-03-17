import logging
from pathlib import Path

from openpyxl import Workbook, load_workbook

from src.cli import build_parser
from src.config import load_config
from src.models import QuotationSummary, RunArtifacts
from src.tender_to_price import TenderToPriceRunner
from src.tender_workflow import TenderWorkflow


class FakePricingEngine:
    def price_workbook(
        self,
        db_path: str,
        boq_path: str,
        output_path: str,
        region: str | None = None,
        threshold: float | None = None,
        apply_rates: bool | None = None,
        column_overrides: dict[str, int] | None = None,
    ) -> RunArtifacts:
        workbook = load_workbook(boq_path)
        workbook.save(output_path)
        return RunArtifacts(
            output_workbook=Path(output_path),
            unmatched_csv=None,
            audit_json=None,
            processed=1,
            matched=1,
            flagged=1,
            quotation_summary=QuotationSummary(),
        )


def _create_simple_boq(path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Earthworks"
    sheet.append(["Description", "Unit", "Qty", "Rate", "Amount"])
    sheet.append(["Excavate foundations", "m3", 12, None, None])
    workbook.save(path)
    return path


def test_tender_price_cli_command_exists() -> None:
    parser = build_parser()
    args = parser.parse_args(
        [
            "tender-price",
            "--input",
            "tender/demo_tender_notice.txt",
            "--db",
            "database/qs_database.xlsx",
            "--out",
            "output/integrated.xlsx",
        ]
    )
    assert args.command == "tender-price"


def test_tender_only_handoff_preserves_missing_quantities() -> None:
    config = load_config()
    runner = TenderToPriceRunner(
        config,
        logging.getLogger("test"),
        tender_workflow=TenderWorkflow(config),
        pricing_engine=FakePricingEngine(),
    )
    result = runner.tender_workflow.prepare_result("tender/demo_tender_scope_only.txt", include_gap_check=True)
    rows = runner.build_pricing_handoff_rows(result, boq_path=None)

    assert rows
    tender_rows = [row for row in rows if row.source_origin == "tender_draft"]
    assert tender_rows
    assert all(row.quantity is None for row in tender_rows)
    assert all(row.review_required is True for row in tender_rows)


def test_tender_price_run_generates_integrated_workbook_with_boq(tmp_path) -> None:
    config = load_config()
    boq_path = _create_simple_boq(tmp_path / "demo_boq.xlsx")
    output_path = tmp_path / "integrated.xlsx"

    runner = TenderToPriceRunner(
        config,
        logging.getLogger("test"),
        tender_workflow=TenderWorkflow(config),
        pricing_engine=FakePricingEngine(),
    )
    artifacts = runner.run(
        input_path="tender/demo_tender_notice.txt",
        db_path="database/qs_database.xlsx",
        output_path=str(output_path),
        boq_path=str(boq_path),
    )

    workbook = load_workbook(artifacts.output_workbook)
    assert "Pricing Handoff" in workbook.sheetnames
    assert "Tender Analysis Summary" in workbook.sheetnames
    assert "BOQ Gap Report" in workbook.sheetnames
