from __future__ import annotations

from io import BytesIO
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook

from src.excel_loader import load_workbook_safe


def _build_workbook_with_invalid_defined_names() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "BOQ"
    sheet["A1"] = "Description"
    sheet["A2"] = "Concrete works"

    stream = BytesIO()
    workbook.save(stream)
    source_bytes = stream.getvalue()

    mutated = BytesIO()
    with ZipFile(BytesIO(source_bytes), "r") as source_zip, ZipFile(mutated, "w", ZIP_DEFLATED) as output_zip:
        for member in source_zip.infolist():
            payload = source_zip.read(member.filename)
            if member.filename == "xl/workbook.xml":
                text = payload.decode("utf-8")
                text = text.replace(
                    "</sheets>",
                    "</sheets><definedNames><definedName name=\"BrokenRange\">#REF!</definedName></definedNames>",
                )
                payload = text.encode("utf-8")
            output_zip.writestr(member, payload)

    return mutated.getvalue()


def test_load_workbook_safe_recovers_from_invalid_defined_names() -> None:
    workbook = load_workbook_safe(_build_workbook_with_invalid_defined_names())

    assert workbook.sheetnames == ["BOQ"]
    assert workbook["BOQ"]["A2"].value == "Concrete works"
