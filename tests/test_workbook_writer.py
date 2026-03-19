from openpyxl import Workbook

from src.workbook_writer import format_worksheet


def test_format_worksheet_applies_client_ready_layout_to_table() -> None:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Pricing Handoff"
    ws.append(["description", "unit", "quantity", "rate", "amount", "notes"])
    ws.append(
        [
            "Excavate and dispose surplus material from foundation trenches with long review text.",
            "m3",
            12.5,
            1450,
            "=C2*D2",
            "Review against site conditions and confirm disposal lead before issue.",
        ]
    )

    format_worksheet(ws, "pricing_handoff")

    assert ws.freeze_panes == "A2"
    assert ws.column_dimensions["A"].width == 60
    assert ws["A2"].alignment.wrap_text is True
    assert ws["C2"].alignment.horizontal == "right"
    assert ws["D2"].number_format == "#,##0.00"
    assert ws["E2"].value == "=C2*D2"
    assert ws.row_dimensions[1].height == 24
    assert ws.row_dimensions[2].height and ws.row_dimensions[2].height > 18


def test_format_worksheet_detects_boq_columns_and_preserves_formula_cells() -> None:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Earthworks"
    ws["A1"] = "Bill No. 1 Earthworks"
    ws.append(["", "", "", "", ""])
    ws.append(["Description", "Unit", "Qty", "Rate", "Amount"])
    ws.append(["Bulk excavation in ordinary soil", "m3", 25, 900, "=C4*D4"])

    format_worksheet(ws, "boq")

    assert ws.freeze_panes == "A2"
    assert ws["A3"].font.bold is True
    assert ws["A4"].alignment.wrap_text is True
    assert ws["C4"].alignment.horizontal == "right"
    assert ws["E4"].value == "=C4*D4"
