from __future__ import annotations

from io import BytesIO
import os

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from app.main import app


client = TestClient(app)


def _build_workbook_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "BOQ"
    sheet.append(["Description", "Unit", "Quantity"])
    sheet.append(["Excavate foundation trench", "m3", 10])
    sheet.append(["Concrete to strip footing", "m3", 5])
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def test_health_check() -> None:
    response = client.get("/health")
    assert response.status_code == 200
    assert response.json() == {"status": "ok"}


def test_upload_boq_returns_json_summary() -> None:
    os.environ.pop("BOQ_AUTO_GCS_BUCKET", None)
    payload = _build_workbook_bytes()
    response = client.post(
        "/upload-boq",
        files={"file": ("sample.xlsx", payload, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"region": "Nairobi", "response_format": "json"},
    )
    assert response.status_code == 200
    body = response.json()
    assert body["filename"] == "sample.xlsx"
    assert body["summary"]["item_count"] == 2
    assert body["summary"]["matched_count"] >= 0
    assert body["summary"]["total_cost"] >= 0
    assert len(body["items"]) == 2
    assert "decision" in body["items"][0]
    assert body["database_path"]
    assert body["input_storage_uri"] is None
    assert body["output_storage_uri"] is None


def test_upload_boq_returns_processed_excel() -> None:
    os.environ.pop("BOQ_AUTO_GCS_BUCKET", None)
    payload = _build_workbook_bytes()
    response = client.post(
        "/upload-boq",
        files={"file": ("sample.xlsx", payload, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"region": "Mombasa", "response_format": "excel"},
    )
    assert response.status_code == 200
    assert response.headers["content-type"].startswith("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    workbook = load_workbook(BytesIO(response.content))
    assert "Match Suggestions" in workbook.sheetnames
    assert "Quotation Summary" in workbook.sheetnames


def test_upload_boq_rejects_non_excel_file() -> None:
    response = client.post(
        "/upload-boq",
        files={"file": ("sample.txt", b"not an excel file", "text/plain")},
        data={"region": "Nairobi", "response_format": "json"},
    )
    assert response.status_code == 400
    assert "Only Excel workbooks" in response.json()["detail"]
