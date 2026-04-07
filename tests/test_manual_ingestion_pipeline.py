from openpyxl import Workbook, load_workbook

from src.cost_schema import CostDatabase, schema_database_path
from src.manual_parser import ManualItem, enhance_manual_item, ingest_manual_to_database, parse_manual_rows, parse_manual_text


def _build_master_db(path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "RateLibrary"
    sheet.append(
        [
            "item_code", "description", "normalized_description", "section", "subsection", "unit", "rate",
            "currency", "region", "source", "source_sheet", "source_page", "basis", "crew_type", "plant_type",
            "material_type", "keywords", "alias_group", "build_up_recipe_id", "confidence_hint", "notes", "active",
        ]
    )
    workbook.create_sheet("Aliases").append(["alias", "canonical_term", "section_bias", "notes"])
    workbook.create_sheet("SectionMap").append(["trigger_text", "inferred_section", "priority"])
    workbook.create_sheet("BuildUpInputs").append(["input_code", "input_type", "description", "unit", "rate", "region", "source", "active"])
    workbook.create_sheet("BuildUpRecipes").append(["recipe_id", "recipe_name", "output_description", "output_unit", "section", "component_code", "factor", "waste_factor", "notes"])
    workbook.create_sheet("Controls").append(["key", "value"])
    workbook.create_sheet("Rules").append(["key", "value"])
    workbook.create_sheet("ReviewLog").append(["timestamp", "boq_file", "sheet_name", "row_number", "boq_description", "decision", "matched_item_code", "matched_description", "confidence_score", "reviewer_note"])
    workbook.save(path)


def test_parse_manual_text_extracts_items_and_merges_continuations() -> None:
    text = "\n".join(
        [
            "10.60 Excavation in foundation trenches m3 1250",
            "10.70 Supply and place selected fill m3 980",
            "including watering and compaction",
        ]
    )

    items = parse_manual_text(text)

    assert len(items) == 2
    assert "watering and compaction" in items[1].description


def test_parse_manual_text_handles_ocr_style_rows_without_item_codes() -> None:
    text = "\n".join(
        [
            "Clear site of works of grass, shrubs, SM 500.00 500.00 500.00",
            "Excavate in normal soil; depth ne 1.5m CM 350.00 350.00 330.00",
        ]
    )

    items = parse_manual_text(text)

    assert len(items) == 2
    assert items[0].unit == "m2"
    assert items[0].rate == 500.0
    assert items[1].unit == "m3"
    assert items[1].rate == 350.0


def test_parse_manual_text_handles_split_description_and_rate_line() -> None:
    text = "\n".join(
        [
            "Excavate in normal soil depth 3.0 - 4.5m",
            "CM 500.00 500.00 500.00",
        ]
    )

    items = parse_manual_text(text)

    assert len(items) == 1
    assert items[0].description.startswith("Excavate in normal soil")
    assert items[0].unit == "m3"
    assert items[0].rate == 500.0


def test_parse_manual_rows_normalizes_excel_like_rows() -> None:
    items = parse_manual_rows([{"code": "A1", "description": "Concrete works", "unit": "M2", "rate": 100}])
    assert items[0].unit == "m2"
    assert items[0].category == "concrete"


def test_ingest_manual_to_database_updates_excel_and_schema(tmp_path) -> None:
    master_db = tmp_path / "master.xlsx"
    _build_master_db(master_db)

    summary = ingest_manual_to_database(
        [ManualItem(code="A1", item_name="Excavation", unit="m3", description="Excavation", rate=1000.0, category="earthworks", material="soil")],
        str(master_db),
        source_name="Manual Review",
        source_file="manual.pdf",
    )

    repository = CostDatabase(schema_database_path(master_db))
    workbook = load_workbook(master_db, data_only=True)
    alias_rows = list(workbook["Aliases"].iter_rows(min_row=2, values_only=True))
    assert summary.appended == 1
    assert len(repository.fetch_items()) == 1
    assert len(repository.fetch_ingestion_logs()) >= 1
    assert repository.fetch_items()[0].keywords
    assert repository.fetch_items()[0].domain
    assert repository.fetch_items()[0].work_family
    assert alias_rows


def test_enhance_manual_item_uses_structured_ai_suggestions() -> None:
    class FixedProvider:
        def suggest_ingestion_attributes(self, description: str, *, unit: str = "", section: str = "", region: str = "") -> dict:
            return {
                "aliases": ["trench excavation"],
                "category": "earthworks",
                "material": "soil",
                "keywords": ["excavation", "trench"],
            }

    item = ManualItem(code="A1", item_name="Excavation", unit="m3", description="Excavation in trench", rate=1000.0)

    enriched = enhance_manual_item(item, enable_ai=True, provider=FixedProvider())

    assert enriched.category == "earthworks"
    assert enriched.material == "soil"
    assert "trench excavation" in enriched.aliases
    assert "trench" in enriched.keywords
