from openpyxl import Workbook

from src.config import load_config
from src.workbook_reader import WorkbookReader


def test_workbook_reader_skips_summary_and_detects_nonstandard_columns(tmp_path) -> None:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Bill No. 3 Earthworks"
    ws["A1"] = "Bill No. 3 Earthworks"
    ws.merge_cells("A1:E1")
    ws.append(["", "", "", "", ""])
    ws.append(["Particulars", "UOM", "Qty", "Unit Rate", "Total"])
    ws.append(["Earthworks", "", "", "", ""])
    ws.append(["Excavation in ordinary soil", "m3", 25, "", ""])
    ws.append(["Subtotal Earthworks", "", "", "", 25000])

    summary = workbook.create_sheet("Summary")
    summary.append(["Section", "Amount"])
    summary.append(["Earthworks", 25000])

    path = tmp_path / "messy_boq.xlsx"
    workbook.save(path)

    reader = WorkbookReader(load_config())
    sheets = reader.read(str(path))
    assert len(sheets) == 1
    assert sheets[0].sheet_name == "Bill No. 3 Earthworks"
    assert sheets[0].columns.description_col == 1
    assert sheets[0].columns.unit_col == 2
    assert sheets[0].columns.quantity_col == 3
    assert sheets[0].classification == "boq"
    assert any(row.description == "Excavation in ordinary soil" for row in sheets[0].rows)


def test_workbook_reader_handles_dayworks_sheet_and_merged_header(tmp_path) -> None:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Contractor's Equipment"
    ws["A1"] = "Dayworks"
    ws.merge_cells("A1:D1")
    ws.append(["Description", "Unit", "Rate", "Amount"])
    ws.append(["15 tonne tipper lorry", "day", 18500, ""])
    path = tmp_path / "dayworks.xlsx"
    workbook.save(path)

    reader = WorkbookReader(load_config())
    sheets = reader.read(str(path))
    assert sheets[0].columns.description_col == 1
    assert sheets[0].rows[0].description == "15 tonne tipper lorry"


def test_workbook_reader_preserves_spec_attributes_column(tmp_path) -> None:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Pricing Handoff"
    ws.append(["Description", "Unit", "Qty", "Spec Attributes"])
    ws.append(["LED light fittings complete", "nr", 20, "600 x 600; LED; Ceiling mounted"])
    path = tmp_path / "handoff.xlsx"
    workbook.save(path)

    reader = WorkbookReader(load_config())
    sheets = reader.read(str(path))

    assert sheets[0].columns.spec_attributes_col == 4
    assert sheets[0].rows[0].spec_attributes == "600 x 600; LED; Ceiling mounted"
