from openpyxl import Workbook, load_workbook

from src.ingestion import (
    ALIASES_HEADERS,
    CANDIDATE_MATCH_HEADERS,
    RATE_LIBRARY_HEADERS,
    REVIEW_LOG_HEADERS,
    generate_review_report,
    merge_reviewed_candidates,
    promote_approved_candidates,
    sync_review_artifacts_to_candidate_matches,
)
from src.cost_schema import CostDatabase


def _make_review_db(path) -> None:
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)

    rate_sheet = workbook.create_sheet("RateLibrary")
    rate_sheet.append(RATE_LIBRARY_HEADERS)
    rate_sheet.append(
        ["PL001", "15 tonne tipper lorry", "15 ton tipper lorry", "Dayworks", "Plant", "day", 18500, "KES", "Nyanza", "QS", "Plant", "12", "hire", "", "transport", "", "", "", "", 4, "", True]
    )

    aliases = workbook.create_sheet("Aliases")
    aliases.append(ALIASES_HEADERS)

    review_log = workbook.create_sheet("ReviewLog")
    review_log.append(REVIEW_LOG_HEADERS)

    candidates = workbook.create_sheet("CandidateMatches")
    candidates.append(CANDIDATE_MATCH_HEADERS)
    candidates.append(
        [
            "2026-03-17T09:00:00", "batch-1", "review.csv", "Sheet1", "RateLibrary", "PL001X",
            "Tipper truck hire", "tipper truck hire", "Dayworks", "Plant", "day", 17800, "KES", "Nyanza",
            "Estimator Review", "44", "Reviewed daywork comparison", "", "transport", "", "tipper lorry", "tipper",
            "", 72, "close duplicate requires review", True, "duplicate normalized description + unit", "PL001",
            "approved", "Senior QS", "2026-03-17T10:15:00", "accept", "aliases",
            "", "", "", "tipper lorry", "Dayworks", 88, "Prefer alias", "not_promoted", "",
        ]
    )
    workbook.save(path)


def test_generate_review_report_creates_candidate_review_sheet(tmp_path) -> None:
    db_path = tmp_path / "review_db.xlsx"
    _make_review_db(db_path)
    summary = generate_review_report(str(db_path))
    workbook = load_workbook(db_path)
    assert "Candidate Review" in workbook.sheetnames
    assert summary.report_rows == 1


def test_merge_reviewed_and_promote_alias(tmp_path) -> None:
    db_path = tmp_path / "review_db.xlsx"
    json_path = tmp_path / "training.json"
    _make_review_db(db_path)
    generate_review_report(str(db_path))

    workbook = load_workbook(db_path)
    review_sheet = workbook["Candidate Review"]
    headers = [cell.value for cell in review_sheet[1]]
    positions = {header: idx + 1 for idx, header in enumerate(headers)}
    review_sheet.cell(2, positions["reviewer_status"]).value = "approved"
    review_sheet.cell(2, positions["promote_target"]).value = "aliases"
    review_sheet.cell(2, positions["approved_canonical_term"]).value = "tipper lorry"
    review_sheet.cell(2, positions["confidence_override"]).value = 91
    workbook.save(db_path)

    merge_summary = merge_reviewed_candidates(str(db_path), "Senior QS")
    promote_summary = promote_approved_candidates(str(db_path), str(json_path))

    workbook = load_workbook(db_path)
    alias_sheet = workbook["Aliases"]
    assert merge_summary.reviewed == 1
    assert promote_summary.promoted == 1
    assert alias_sheet.max_row == 2
    assert json_path.exists()


def test_sync_review_artifacts_bridges_schema_records_into_candidate_matches(tmp_path) -> None:
    db_path = tmp_path / "review_db.xlsx"
    _make_review_db(db_path)
    repository = CostDatabase(db_path)
    repository.record_rate_observation(
        "Excavator hire",
        "Excavator hire",
        "day",
        24500,
        source="review_task",
        reviewer="QS A",
        metadata={"task_id": "task-rate-1", "section": "Dayworks", "region": "Nyanza", "reviewer_note": "Approved field rate"},
    )
    repository.record_alias_suggestion(
        "Tipper truck",
        "Tipper lorry",
        section_bias="Dayworks",
        reviewer="QS B",
        status="approved",
        metadata={"task_id": "task-alias-1", "reviewer_note": "Use simpler phrasing"},
    )
    repository.record_candidate_review(
        "HDPE pressure pipe",
        "HDPE pressure pipe",
        "m",
        reason="no_good_match",
        reviewer="QS C",
        status="approved",
        metadata={"task_id": "task-review-1", "section": "Utilities", "region": "Nairobi"},
    )

    summary = sync_review_artifacts_to_candidate_matches(str(db_path))

    workbook = load_workbook(db_path)
    candidate_sheet = workbook["CandidateMatches"]
    rows = list(candidate_sheet.iter_rows(min_row=2, values_only=True))
    source_markers = {str(row[2] or "") for row in rows}
    appended_rows = [row for row in rows if str(row[2] or "").startswith("schema-task:")]

    assert summary.appended == 3
    assert "schema-task:task-rate-1" in source_markers
    assert "schema-task:task-alias-1" in source_markers
    assert "schema-task:task-review-1" in source_markers
    assert len(appended_rows) == 3

    rate_row = next(row for row in appended_rows if str(row[2]) == "schema-task:task-rate-1")
    alias_row = next(row for row in appended_rows if str(row[2]) == "schema-task:task-alias-1")
    review_row = next(row for row in appended_rows if str(row[2]) == "schema-task:task-review-1")

    assert rate_row[33] == "ratelibrary"
    assert rate_row[28] == "approved"
    assert alias_row[33] == "aliases"
    assert alias_row[28] == "approved"
    assert review_row[33] == "candidatematches"
    assert review_row[28] == "pending"

    rerun_summary = sync_review_artifacts_to_candidate_matches(str(db_path))
    assert rerun_summary.appended == 0
    assert rerun_summary.skipped_duplicates == 3


def test_synced_review_artifacts_can_flow_into_existing_promotions(tmp_path) -> None:
    db_path = tmp_path / "review_db.xlsx"
    _make_review_db(db_path)
    repository = CostDatabase(db_path)
    repository.record_rate_observation(
        "Crawler excavator hire",
        "Crawler excavator hire",
        "day",
        26500,
        source="review_task",
        reviewer="Senior QS",
        metadata={"task_id": "task-rate-2", "section": "Dayworks", "region": "Nyanza"},
    )
    repository.record_alias_suggestion(
        "Vibrating roller",
        "Vibratory roller",
        section_bias="Roadworks",
        reviewer="Senior QS",
        status="approved",
        metadata={"task_id": "task-alias-2"},
    )

    sync_review_artifacts_to_candidate_matches(str(db_path))
    promote_summary = promote_approved_candidates(str(db_path))

    workbook = load_workbook(db_path)
    rate_sheet = workbook["RateLibrary"]
    alias_sheet = workbook["Aliases"]

    rate_descriptions = {str(values[1] or "") for values in rate_sheet.iter_rows(min_row=2, values_only=True)}
    alias_pairs = {(str(values[0] or ""), str(values[1] or "")) for values in alias_sheet.iter_rows(min_row=2, values_only=True)}

    assert promote_summary.promoted >= 2
    assert "Crawler excavator hire" in rate_descriptions
    assert ("Vibrating roller", "Vibratory roller") in alias_pairs
