from openpyxl import Workbook, load_workbook

from src.cli import build_parser
from src.cost_schema import CostDatabase, schema_database_path
from src.models import AppConfig
from src.ingestion import (
    ALIASES_HEADERS,
    RATE_LIBRARY_HEADERS,
    REVIEW_LOG_HEADERS,
    import_priced_boq,
    preview_priced_boq_import,
)


def _make_db(path) -> None:
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)

    rate_sheet = workbook.create_sheet("RateLibrary")
    rate_sheet.append(RATE_LIBRARY_HEADERS)
    rate_sheet.append(
        ["EW001", "Excavate in ordinary soil", "excavate in ordinary soil", "Earthworks", "", "m3", 950, "KES", "Nyanza", "QS", "Earthworks", "8", "library", "", "", "soil", "excavate, soil", "", "", 0, "", True]
    )

    workbook.create_sheet("Aliases").append(ALIASES_HEADERS)
    workbook.create_sheet("SectionMap").append(["trigger_text", "inferred_section", "priority"])
    workbook.create_sheet("BuildUpInputs").append(["input_code", "input_type", "description", "unit", "rate", "region", "source", "active"])
    workbook.create_sheet("BuildUpRecipes").append(["recipe_id", "recipe_name", "output_description", "output_unit", "section", "component_code", "factor", "waste_factor", "notes"])
    workbook.create_sheet("Controls").append(["key", "value"])
    workbook.create_sheet("Rules").append(["key", "value"])
    workbook.create_sheet("ReviewLog").append(REVIEW_LOG_HEADERS)
    workbook.save(path)


def _make_priced_boq(path) -> None:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Bill No. 1 Earthworks"
    ws.append(["Description", "Unit", "Qty", "Rate", "Amount"])
    ws.append(["Earthworks", "", "", "", ""])
    ws.append(["Excavate in ordinary soil", "m3", 25, 980, 24500])
    ws.append(["Selected fill to make up levels", "m3", 10, 1250, 12500])
    ws.append(["Unpriced item", "m3", 5, "", ""])
    workbook.save(path)


def test_import_priced_boq_cli_requires_region() -> None:
    parser = build_parser()
    args = parser.parse_args(
        [
            "import-priced-boq",
            "--db",
            "database/master/qs_database_master.xlsx",
            "--input",
            "boq/priced.xlsx",
            "--region",
            "Nairobi",
        ]
    )
    assert args.command == "import-priced-boq"
    assert args.region == "Nairobi"


def test_import_priced_boq_adds_priced_rows_and_skips_unpriced(tmp_path) -> None:
    db_path = tmp_path / "master.xlsx"
    boq_path = tmp_path / "priced_boq.xlsx"
    _make_db(db_path)
    _make_priced_boq(boq_path)

    summary = import_priced_boq(str(db_path), str(boq_path), region="Nairobi", source_label="Prepared BOQ")

    workbook = load_workbook(db_path, data_only=True)
    rate_sheet = workbook["RateLibrary"]
    rows = list(rate_sheet.iter_rows(min_row=2, values_only=True))
    imported = [row for row in rows if str(row[9] or "") == "Prepared BOQ"]

    assert summary.appended == 2
    assert summary.candidates_created == 0
    assert any("Skipped 1 BOQ row(s) without a usable rate." in note for note in summary.notes)
    assert len(imported) == 2
    assert all(str(row[8] or "") == "Nairobi" for row in imported)


def test_import_priced_boq_is_region_sensitive_for_duplicates(tmp_path) -> None:
    db_path = tmp_path / "master.xlsx"
    boq_path = tmp_path / "priced_boq.xlsx"
    _make_db(db_path)
    _make_priced_boq(boq_path)

    nyanza_summary = import_priced_boq(str(db_path), str(boq_path), region="Nyanza", source_label="Regional BOQ")
    nairobi_summary = import_priced_boq(str(db_path), str(boq_path), region="Nairobi", source_label="Regional BOQ")

    workbook = load_workbook(db_path, data_only=True)
    rate_rows = list(workbook["RateLibrary"].iter_rows(min_row=2, values_only=True))
    candidate_rows = list(workbook["CandidateMatches"].iter_rows(min_row=2, values_only=True)) if "CandidateMatches" in workbook.sheetnames else []

    assert nyanza_summary.candidates_created >= 1
    assert nairobi_summary.appended == 2
    assert any(str(row[8] or "") == "Nairobi" and str(row[1] or "") == "Excavate in ordinary soil" for row in rate_rows)
    assert any(str(row[13] or "") == "Nyanza" for row in candidate_rows)


def test_import_priced_boq_pdf_adds_clear_priced_rows(tmp_path, monkeypatch) -> None:
    db_path = tmp_path / "master.xlsx"
    pdf_path = tmp_path / "priced_boq.pdf"
    _make_db(db_path)
    pdf_path.write_bytes(b"%PDF-1.4")

    monkeypatch.setattr(
        "src.ingestion.extract_text_from_pdf",
        lambda path, config=None, logger=None: "\n".join(
            [
                "EARTHWORKS",
                "Excavate in ordinary soil m3 25 980 24500",
                "Selected fill to make up levels m3 10 1250 12500",
                "Unpriced item m3 5",
            ]
        ),
    )

    summary = import_priced_boq(str(db_path), str(pdf_path), region="Nairobi", source_label="Prepared PDF BOQ")

    workbook = load_workbook(db_path, data_only=True)
    rate_sheet = workbook["RateLibrary"]
    rows = list(rate_sheet.iter_rows(min_row=2, values_only=True))
    imported = [row for row in rows if str(row[9] or "") == "Prepared PDF BOQ"]

    assert summary.appended == 2
    assert summary.candidates_created == 0
    assert len(imported) == 2
    assert all(str(row[8] or "") == "Nairobi" for row in imported)


def test_preview_priced_boq_import_shows_rows_and_aliases(tmp_path, monkeypatch) -> None:
    boq_path = tmp_path / "priced_boq.pdf"
    boq_path.write_bytes(b"%PDF-1.4")

    class FixedProvider:
        def suggest_aliases(self, text: str) -> list[str]:
            return ["soil excavation"] if "Excavate" in text else []

    monkeypatch.setattr(
        "src.ingestion.extract_text_from_pdf",
        lambda path, config=None, logger=None: "\n".join(
            [
                "EARTHWORKS",
                "Excavate in ordinary soil m3 25 980 24500",
                "Selected fill to make up levels m3 10 1250 12500",
                "Unpriced item m3 5",
            ]
        ),
    )
    monkeypatch.setattr("src.ai.embedding_provider.get_embedding_provider", lambda config: FixedProvider())

    preview = preview_priced_boq_import(
        str(boq_path),
        region="Nairobi",
        source_label="Preview PDF BOQ",
        config=AppConfig(data={"ai": {"enabled": True}}),
        ai_assist=True,
    )

    assert preview.total_extracted == 2
    assert preview.skipped_missing_rate == 0
    assert preview.skipped_missing_unit == 0
    assert len(preview.extracted_rows) == 2
    assert preview.extracted_rows[0]["region"] == "Nairobi"
    assert any("AI-assisted alias suggestions" in note for note in preview.notes)
    assert any(row["description"] == "Excavate in ordinary soil" for row in preview.alias_preview)
    assert any("soil excavation" in row["ai_aliases"] for row in preview.alias_preview)


def test_preview_priced_boq_import_shows_append_and_candidate_decisions(tmp_path) -> None:
    db_path = tmp_path / "master.xlsx"
    boq_path = tmp_path / "priced_boq.xlsx"
    _make_db(db_path)
    _make_priced_boq(boq_path)

    preview = preview_priced_boq_import(
        db_path=str(db_path),
        boq_path=str(boq_path),
        region="Nyanza",
        source_label="Regional BOQ",
    )

    assert preview.total_extracted == 2
    assert preview.append_count == 1
    assert preview.candidate_count == 1
    assert preview.duplicate_count == 0
    assert any(row["description"] == "Selected fill to make up levels" for row in preview.append_preview)
    assert any(row["description"] == "Excavate in ordinary soil" for row in preview.candidate_preview)


def test_import_priced_boq_syncs_style_aliases_and_schema(tmp_path, monkeypatch) -> None:
    db_path = tmp_path / "master.xlsx"
    boq_path = tmp_path / "priced_boq.xlsx"
    _make_db(db_path)

    workbook = Workbook()
    ws = workbook.active
    ws.title = "Earthworks"
    ws.append(["Description", "Unit", "Qty", "Rate", "Amount"])
    ws.append(["Excavate in normal soil; depth 1.5-3.0m", "m3", 10, 410, 4100])
    workbook.save(boq_path)

    class FixedProvider:
        model_name = "fixed-test"

        def suggest_aliases(self, text: str) -> list[str]:
            return ["soil excavation"]

        def embed(self, text: str) -> list[float]:
            return [0.1, 0.2, 0.3]

    monkeypatch.setattr("src.ai.embedding_provider.get_embedding_provider", lambda config: FixedProvider())

    summary = import_priced_boq(
        str(db_path),
        str(boq_path),
        region="Nairobi",
        source_label="Style BOQ",
        config=AppConfig(data={"ai": {"enabled": True}}),
        ai_assist=True,
    )

    workbook = load_workbook(db_path, data_only=True)
    alias_rows = list(workbook["Aliases"].iter_rows(min_row=2, values_only=True))
    repository = CostDatabase(schema_database_path(db_path))
    schema_items = repository.fetch_items()
    schema_aliases = repository.fetch_aliases()

    assert summary.appended == 1
    assert any("BOQ-derived alias" in note for note in summary.notes)
    assert any("normalized schema" in note for note in summary.notes)
    assert any(str(row[0] or "") == "Excavate in normal soil" for row in alias_rows)
    assert any(str(row[0] or "") == "soil excavation" for row in alias_rows) is False
    assert len(schema_items) == 1
    assert len(schema_aliases) >= 2
    assert any(alias.alias == "soil excavation" for alias in schema_aliases)
